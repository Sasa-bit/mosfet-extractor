"""
Universal Power-MOSFET Datasheet Parameter Extractor  v1
========================================================
SINGLE FILE — no helper scripts needed.

Built on the same engine as the SiC/Si Schottky diode extractor, retargeted
to power MOSFETs (Infineon CoolMOS/CoolSiC, ST, ON Semi, Toshiba, Nexperia,
Vishay, ROHM, Wolfspeed, IXYS ...).

What it produces (one .xlsx workbook):
  1) "MOSFET Parameters" sheet — every loss-relevant parameter, found or not:
       Required:  V(BR)DSS, ID, RDS(on), VGS(th), Ciss, Coss, Crss, Qg
       Loss:      Qgs, Qgd, td(on), tr, td(off), tf, Eon, Eoff, Eoss, VSD,
                  trr, Qrr
       Thermal:   RthJC, RthJA, RthCS — and ANY parameter whose unit is K/W
                  or °C/W is captured here.
     Typ + Max at every stated temperature, test conditions, loss formula.
  2) "Energy Graphs" sheet — every figure containing Eon / Eoff / Etot
     (switching-energy curves) digitised point-by-point. The X axis may be
     gate resistance RG (Ω), drain current ID (A) or junction temperature
     Tj (°C); the printed X-axis unit is detected and used as the column head.
  3) "Temperature Graphs" sheet — every figure whose variable is a temperature
     (Tj / Tc / Ta / Th): RDS(on) vs Tj, ID/Ptot derating vs Tc, safe-operating
     curves, etc. — digitised the same way.

Both graph sheets embed the original figure image next to each data table so
the digitised numbers can be checked at a glance.

Extraction strategy (unchanged, battle-tested):
  · pdfplumber tables  → structure auto-detect  → symbol/name regex matching
  · word-baseline text fallback (narrow / bilingual / split-subscript sheets)
  · plain-text line fallback
  · universal unit normalisation (mΩ↔Ω, pF↔nF, nC↔µC, ns↔µs, µJ↔mJ, K/W↔°C/W …)
    applied BEFORE range-checking, so a value in any prefix is accepted
  · graph digitisation: PDF vector drawing layer first, embedded-raster-image
    (OCR-calibrated) fallback — for both energy and temperature figures

Usage:
    python mosfet_extractor_v1.py  datasheet.pdf
    python mosfet_extractor_v1.py  datasheet.pdf  --output result.xlsx
    python mosfet_extractor_v1.py  datasheet.pdf  --no-prompt

Install once:
    pip install pdfplumber openpyxl pymupdf opencv-python numpy scipy Pillow pytesseract
    (plus the tesseract-ocr engine for vector-glyph / raster axis labels:
     apt install tesseract-ocr   /   Windows: UB-Mannheim installer)
"""

# ── imports ───────────────────────────────────────────────────────────────────
import re
import math, sys, os, io, argparse, itertools, warnings
import numpy as np
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Excel/XML worksheets cannot store C0 control characters (everything below
# 0x20 except tab/newline/CR). PDF text extraction occasionally emits such a
# char (e.g. \x07 BEL) inside captions or axis labels — it is invisible when
# printed but makes openpyxl raise IllegalCharacterError on cell assignment.
# Strip them at the cell-write boundary so the workbook never sees them.
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE as _XL_BAD_RE
except Exception:                                    # version-safe fallback
    import re as _re_xl
    _XL_BAD_RE = _re_xl.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")

def _xl_safe(v):
    """Return v with worksheet-illegal control chars removed; non-str unchanged."""
    if isinstance(v, str):
        return _XL_BAD_RE.sub("", v)
    return v
from scipy.stats import linregress

warnings.filterwarnings("ignore")

try:
    import cv2
    import fitz
    from PIL import Image
    _HAS_GRAPH = True
except ImportError:
    _HAS_GRAPH = False

try:
    import pytesseract
    pytesseract.get_tesseract_version()
    _HAS_OCR = True
except Exception:
    _HAS_OCR = False

# Infineon uses a private Symbol font in some datasheets; pdfplumber renders
# these characters as "(cid:XXXX)" placeholders. This map decodes them.
_INFINEON_CID_MAP = {
    1089: "0", 1090: "1", 1091: "7", 1092: "2",
    1093: "4", 1094: "5", 1095: "6", 1096: "8",
    # CIDs 1177-1179 are typically non-digit symbols — map to empty string
    1177: "",  1178: "",  1179: "",
}

def _decode_cid_text(text):
    """Replace (cid:N) placeholders with actual characters using Infineon font map."""
    return re.sub(r"\(cid:(\d+)\)",
                  lambda m: _INFINEON_CID_MAP.get(int(m.group(1)), ""),
                  text)

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1 — GRAPH-BASED rd EXTRACTION
# ═════════════════════════════════════════════════════════════════════════════

_DARK   = 80
_CGAP   = 8
_CMIN   = 2
_CJUMP  = 22
_AXFRAC = 0.50

def _p2v(px, p0, p1, v0, v1):
    return v0 + (px-p0)/(p1-p0)*(v1-v0) if p1!=p0 else v0

def _v2p(v, v0, v1, p0, p1):
    return int(p0 + (v-v0)/(v1-v0)*(p1-p0)) if v1!=v0 else p0

def _axbox(gray):
    h, w = gray.shape
    rd = (gray < _DARK).sum(axis=1).astype(float)
    cd = (gray < _DARK).sum(axis=0).astype(float)
    ry = np.where(rd > w*_AXFRAC)[0]
    rx = np.where(cd > h*_AXFRAC)[0]
    if len(ry)<2 or len(rx)<2: return None
    yt,yb = int(ry[0]),int(ry[-1])
    xl,xr = int(rx[0]),int(rx[-1])
    if (xr-xl)<0.15*w or (yb-yt)<0.15*h: return None
    return xl,xr,yt,yb

def _strong_axis_lines(gray, frac=0.5):
    """Pixel positions of strong (mostly-dark) vertical and horizontal lines —
    i.e. the plot-frame / axis borders.  Adjacent dark columns/rows are merged
    into a single line position.  Returns (V_positions, H_positions)."""
    h, w = gray.shape
    cd = (gray < _DARK).sum(axis=0) / float(h)   # dark fraction per column
    rd = (gray < _DARK).sum(axis=1) / float(w)   # dark fraction per row
    def _merge(idx):
        if not idx: return []
        idx = sorted(idx); out, cur = [], [idx[0]]
        for v in idx[1:]:
            if v - cur[-1] <= 3: cur.append(v)
            else: out.append(int(np.mean(cur))); cur = [v]
        out.append(int(np.mean(cur)))
        return out
    V = _merge([x for x in range(w) if cd[x] > frac])
    H = _merge([y for y in range(h) if rd[y] > frac])
    return V, H

def _data_plot_box(gray, x_px, y_px, frac=0.5):
    """Find the TRUE data-plot rectangle (inner axis borders), not the outer
    figure/table frame.  Among the strong border lines, picks the tightest pair
    on each side that still brackets every calibration tick pixel (x_px / y_px).
    This snaps to the real x=min .. x=max axis borders even when OCR only read a
    subset of the ticks, and ignores a wider enclosing figure frame.
    Returns (xl, xr, yt, yb) in image pixels, or None if it can't be resolved."""
    if not x_px or not y_px:
        return None
    V, H = _strong_axis_lines(gray, frac)
    if len(V) < 2 or len(H) < 2:
        return None
    xmin, xmax = min(x_px), max(x_px)
    ymin, ymax = min(y_px), max(y_px)
    left  = [v for v in V if v <= xmin + 3]
    right = [v for v in V if v >= xmax - 3]
    top   = [hh for hh in H if hh <= ymin + 3]
    bot   = [hh for hh in H if hh >= ymax - 3]
    if not (left and right and top and bot):
        return None
    xl, xr, yt, yb = max(left), min(right), max(top), min(bot)
    if (xr - xl) < 40 or (yb - yt) < 40:
        return None
    return int(xl), int(xr), int(yt), int(yb)

def _clusters(ypix, oy, gap=_CGAP, mn=_CMIN):
    if not ypix: return []
    sp=sorted(ypix); cl,cur=[],[sp[0]]
    for y in sp[1:]:
        if y-cur[-1]<=gap: cur.append(y)
        else: cl.append(cur); cur=[y]
    cl.append(cur)
    return [(int(np.mean(c))+oy, len(c)) for c in cl if len(c)>=mn]

def _coldata(gplot, ix0, iy0):
    ph,pw=gplot.shape; d={}
    for cr in range(pw):
        ca=cr+ix0
        yp=np.where(gplot[:,cr]<_DARK)[0].tolist()
        cl=_clusters(yp, iy0)
        if cl: d[ca]=cl
    return d

def _track(sc, sr, cd, p2if, mn, mx, fwd=True, jmp=_CJUMP):
    c=[(sc,sr)]; cr=sr; pc=sc
    cols=sorted(cd.keys())
    srch=[x for x in cols if (x>sc if fwd else x<sc)]
    if not fwd: srch=srch[::-1]
    for ca in srch:
        if abs(ca-pc)>6: break
        cands=[(r,s) for r,s in cd[ca]
               if abs(r-cr)<=jmp and mn<=p2if(r)<=mx*1.05]
        if not cands: pc=ca; continue
        br=min(cands,key=lambda x:abs(x[0]-cr))[0]
        c.append((ca,br)); cr=br; pc=ca
    return c

def _validate(gray, axbox):
    xl,xr,yt,yb=axbox
    inner=gray[yt+3:yb-3,xl+3:xr-3]; ih,iw=inner.shape
    if ih<30 or iw<30: return False
    l70=inner[:,:int(iw*0.70)]
    t60=int(ih*0.40)
    dr=np.where(l70<_DARK)[0]
    if len(dr)==0 or dr.min()>=t60: return False
    return True

# VF-IF page caption patterns — covers all manufacturer styles
_VFIF_POS = re.compile(
    r"I\s*F\s*=\s*f\s*\(\s*V\s*F"
    r"|forward\s+char"
    r"|typ\.?\s+forward"
    r"|forward\s+i.?v\b"
    r"|I\s*F\s+versus\s+V\s*F"
    r"|forward\s+(voltage|current).{0,8}(vs\.?|versus|/).{0,8}(forward|current|voltage)"
    r"|forward\s+voltage\s+(drop|char)"
    r"|typical\s+forward"
    r"|V\s*F\s*[\[/\(].*?[Vv]\s*[\]/\)]"
    r"|I\s*F\s*[\[/\(].*?[Aa]\s*[\]/\)]"
    r"|i-v\s+char|i/v\s+char"
    r"|forward.biased\s+char", re.I)

_VFIF_NEG = re.compile(
    r"E\s*C\s*=\s*f\s*\("
    r"|capacitance\s+stored"
    r"|mathematical\s+equation"
    r"|simplified\s+forward\s+char"
    r"|transient\s+thermal"
    r"|Z\s*th.*?=.*?f\("
    r"|reverse\s+recovery\s+time"
    r"|switching\s+char", re.I)

def _is_vfif_page(doc, pi):
    txt=doc[pi].get_text("text")
    return bool(_VFIF_POS.search(txt)) and not bool(_VFIF_NEG.search(txt))

def _guess_axes(doc, pi, bbox, if_rated):
    page=doc[pi]; ix0,iy0,ix1,iy1=bbox; mg=50
    def nums(axis):
        out=[]
        for blk in page.get_text("dict")["blocks"]:
            if blk["type"]!=0: continue
            bx0,by0,bx1,by1=blk["bbox"]
            ok=(ix0-mg<bx0<ix1+mg and iy1-mg<by0<iy1+mg*3) if axis=="x" else \
               (ix0-mg*2<bx0<ix0+mg*1.5 and iy0-mg<by0<iy1+mg)
            if not ok: continue
            for ln in blk.get("lines",[]):
                for sp in ln.get("spans",[]):
                    for m in re.findall(r"\d+\.?\d*",sp["text"]):
                        try: out.append(float(m))
                        except: pass
        return out
    vf=max((v for v in nums("x") if 1.5<=v<=8),default=3.0)
    ifm=max((v for v in nums("y") if if_rated*1.2<=v<=if_rated*12),default=if_rated*2.0)
    return float(vf), float(ifm)

def _rd_from_image(gray, if_rated, vf_rated, vf_max, if_max, dbg=None):
    ax=_axbox(gray)
    if not ax: return None
    xl,xr,yt,yb=ax
    if (xr-xl)<40 or (yb-yt)<40: return None
    p2vf=lambda c: _p2v(c,xl,xr,0,vf_max)
    p2if=lambda r: _p2v(r,yb,yt,0,if_max)
    vf2p=lambda v: _v2p(v,0,vf_max,xl,xr)
    if2p=lambda i: _v2p(i,0,if_max,yb,yt)
    iy0=yt+3; ix0=xl+3
    cd=_coldata(gray[iy0:yb-3,ix0:xr-3], ix0, iy0)
    if not cd: return None
    ac=vf2p(vf_rated); ar=if2p(if_rated)
    bd=999; sc,sr=ac,ar
    for ca in range(max(ix0,ac-25),min(xr-3,ac+25)):
        for ra,sz in cd.get(ca,[]):
            d=abs(ra-ar)
            if d<bd: bd=d; sc,sr=ca,ra
    mn=if_rated*0.50; mx=if_rated*1.00
    fwd=_track(sc,sr,cd,p2if,mn,mx,True)
    bwd=_track(sc,sr,cd,p2if,mn,mx,False)
    curve=list(reversed(bwd))+fwd[1:]
    vfp=np.array([p2vf(c) for c,r in curve])
    ifp=np.array([p2if(r) for c,r in curve])
    mask=(ifp>=mn)&(ifp<=mx)&(vfp>0.3)
    vl=vfp[mask]; il=ifp[mask]
    if len(il)<6 or len(np.unique(il))<3: return None
    sl,ic,rv,_,_=linregress(il,vl)
    rd=sl; vf0=ic; r2=rv**2
    if not np.isfinite(rd) or rd<=0 or rd>2.0: return None
    if r2<0.80: return None
    if dbg is not None:
        try:
            from PIL import ImageDraw
            draw=ImageDraw.Draw(dbg)
            for c,r in curve:
                if 0<=c<dbg.width and 0<=r<dbg.height:
                    draw.ellipse([c-2,r-2,c+2,r+2],fill='lime')
            vff=[vf0+rd*i for i in [mn,mx]]
            draw.line([vf2p(vff[0]),if2p(mn),vf2p(vff[1]),if2p(mx)],fill='red',width=3)
            draw.ellipse([vf2p(vf_rated)-8,if2p(if_rated)-8,
                          vf2p(vf_rated)+8,if2p(if_rated)+8],outline='yellow',width=3)
        except: pass
    return rd,vf0,r2,len(il)

def _page_images(doc, pi, if_rated, vf_rated, vmax, ifmax, dbg_path):
    page=doc[pi]; best=None; bscore=-999
    for inf in page.get_images(full=True):
        xref,_,w,h=inf[:4]
        if w<150 or h<150: continue
        try:
            raw=doc.extract_image(xref)
            pil=Image.open(io.BytesIO(raw["image"])).convert("RGB")
        except: continue
        gray=cv2.cvtColor(np.array(pil),cv2.COLOR_RGB2GRAY)
        ax=_axbox(gray)
        if ax and not _validate(gray,ax): continue
        bbox=None
        try:
            for item in page.get_image_rects(xref): bbox=(item.x0,item.y0,item.x1,item.y1); break
        except: pass
        if not bbox:
            for blk in page.get_text("dict")["blocks"]:
                if blk.get("type")==1: bbox=blk["bbox"]; break
        _vm=vmax; _im=ifmax
        if bbox and (_vm is None or _im is None):
            av,ai=_guess_axes(doc,pi,bbox,if_rated)
            if _vm is None: _vm=av
            if _im is None: _im=ai
        vpc=[vmax] if vmax else [3.0,3.5,4.0,2.5,5.0]
        ip=round(if_rated*2.0/4)*4
        ipc=[ifmax] if ifmax else sorted(set([ip,round(if_rated*2.5/4)*4,max(8,_im or 8)]))
        for vt in vpc:
            vp=max(0.0,(2.8-vt)*2.0)
            for it in ipc:
                d=pil.copy() if dbg_path else None
                r=_rd_from_image(gray,if_rated,vf_rated,vt,it,dbg=d)
                if r:
                    rd,vf0,r2,n=r
                    ap=2.0*abs(if_rated/it-0.50)
                    sc=r2-vp-ap
                    if sc>bscore:
                        bscore=sc
                        best={"rd":round(rd,5),"vf0":round(vf0,4),"r2":round(r2,5),"n":n,
                              "method":"Graph digitization (VF-IF pixel regression)",
                              "cond":("rd from VF-IF graph 25°C curve "
                                     "{:.1f}-{:.1f}A; anchor {:.3f}V@{:.1f}A R²={:.4f}").format(
                                  if_rated*0.5,if_rated,vf_rated,if_rated,r2)}
                        if dbg_path and d: d.save(dbg_path)
    return best,bscore

def _raster_pages(doc, if_rated, vf_rated, vmax, ifmax, dbg_path):
    best=None; bscore=-999
    for pi in range(doc.page_count):
        mat=fitz.Matrix(3.5,3.5)
        pix=doc[pi].get_pixmap(matrix=mat)
        try: full=Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        except: continue
        pw,ph=full.size
        th=int(ph*0.38); tw=int(pw*0.38)
        for y0 in range(0,ph-th+1,th//3):
            for x0 in range(0,pw-tw+1,tw//3):
                tile=full.crop((x0,y0,x0+tw,y0+th))
                gray=cv2.cvtColor(np.array(tile),cv2.COLOR_RGB2GRAY)
                ax=_axbox(gray)
                if not ax or not _validate(gray,ax): continue
                vpc=[vmax] if vmax else [3.0,3.5,4.0,2.5]
                ip=round(if_rated*2.0/4)*4
                ipc=[ifmax] if ifmax else sorted(set([ip,round(if_rated*2.5/4)*4]))
                for vt in vpc:
                    vp=max(0.0,(2.8-vt)*2.0)
                    for it in ipc:
                        d=tile.copy() if dbg_path else None
                        r=_rd_from_image(gray,if_rated,vf_rated,vt,it,dbg=d)
                        if r:
                            rd,vf0,r2,n=r
                            ap=2.0*abs(if_rated/it-0.50)
                            sc=r2-vp-ap
                            if sc>bscore:
                                bscore=sc
                                best={"rd":round(rd,5),"vf0":round(vf0,4),"r2":round(r2,5),"n":n,
                                      "method":"Graph digitization (page rasterised)",
                                      "cond":("rd from VF-IF graph 25°C curve "
                                             "{:.1f}-{:.1f}A; anchor {:.3f}V@{:.1f}A R²={:.4f}").format(
                                          if_rated*0.5,if_rated,vf_rated,if_rated,r2)}
                                if dbg_path and d: d.save(dbg_path)
    return best,bscore

def extract_rd_from_graph(pdf_path, if_rated, vf_rated,
                          if_max_axis=None, vf_max_axis=None, debug_png=None):
    if not (_HAS_GRAPH and os.path.exists(pdf_path)): return None
    try: doc=fitz.open(pdf_path)
    except: return None
    best=None; bscore=-999
    # Pass 1: caption-matched pages → embedded images
    matched=[pi for pi in range(doc.page_count) if _is_vfif_page(doc,pi)]
    for pi in matched:
        r,sc=_page_images(doc,pi,if_rated,vf_rated,vf_max_axis,if_max_axis,debug_png)
        if r and sc>bscore: bscore=sc; best=r
    # Pass 2: all pages → embedded images
    if best is None:
        for pi in range(doc.page_count):
            if pi in matched: continue
            r,sc=_page_images(doc,pi,if_rated,vf_rated,vf_max_axis,if_max_axis,debug_png)
            if r and sc>bscore: bscore=sc; best=r
    # Pass 3: page rasterisation (vector PDFs — WeEN, STM etc.)
    if best is None:
        r,sc=_raster_pages(doc,if_rated,vf_rated,vf_max_axis,if_max_axis,debug_png)
        if r and sc>bscore: best=r
    doc.close()
    return best

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1B — TEMPERATURE-GRAPH DIGITISATION
#   IF vs T  ·  Ptot vs T  ·  IFSM(Tj)/IFSM(25°C)  ·  relative dynamic params …
# ═════════════════════════════════════════════════════════════════════════════
# Extracts data points from every figure whose X axis is a temperature
# (junction / case / heatsink / ambient).  Works on the PDF *vector drawing
# layer* (exact coordinates).  Axis ticks are read from (1) the text layer
# (ST, onsemi, Nexperia draw them as text), (2) OCR via pytesseract (WeEN,
# Diotec draw them as vector glyphs), (3) an interactive prompt, and only as
# a last resort fall back to normalised 0-1 coordinates.

def _item_to_segments(it):
    """Drawing item → list of centreline/edge segments [((x1,y1),(x2,y2)),...].
    'l' = line; 'c' = bezier (endpoints); 'qu' = stroked quad — thin quads are
    lines drawn with thickness (centreline), fat quads are grid CELLS whose
    four edges form the gridlines (Diotec draws entire graph grids this way)."""
    if it[0] == "l":
        return [((it[1].x, it[1].y), (it[2].x, it[2].y))]
    if it[0] == "c":
        return [((it[1].x, it[1].y), (it[4].x, it[4].y))]
    if it[0] == "re":
        # rectangle (plot box / grid cell) drawn as a single 're' op →
        # decompose into its four edges so the box detector can see them.
        r = it[1]
        x0r, y0r, x1r, y1r = r.x0, r.y0, r.x1, r.y1
        return [((x0r, y0r), (x1r, y0r)), ((x0r, y1r), (x1r, y1r)),
                ((x0r, y0r), (x0r, y1r)), ((x1r, y0r), (x1r, y1r))]
    if it[0] == "qu":
        q = it[1]
        xs = [q.ul.x, q.ur.x, q.ll.x, q.lr.x]
        ys = [q.ul.y, q.ur.y, q.ll.y, q.lr.y]
        w, h = max(xs) - min(xs), max(ys) - min(ys)
        if min(w, h) < 2.5:           # thin quad = a thick-stroked line
            if w >= h:
                return [(((q.ul.x + q.ll.x) / 2, (q.ul.y + q.ll.y) / 2),
                         ((q.ur.x + q.lr.x) / 2, (q.ur.y + q.lr.y) / 2))]
            return [(((q.ul.x + q.ur.x) / 2, (q.ul.y + q.ur.y) / 2),
                     ((q.ll.x + q.lr.x) / 2, (q.ll.y + q.lr.y) / 2))]
        # fat quad = grid cell → its 4 edges
        return [((q.ul.x, q.ul.y), (q.ur.x, q.ur.y)),
                ((q.ll.x, q.ll.y), (q.lr.x, q.lr.y)),
                ((q.ul.x, q.ul.y), (q.ll.x, q.ll.y)),
                ((q.ur.x, q.ur.y), (q.lr.x, q.lr.y))]
    return []

def _merge_collinear(pieces, tol=0.8):
    """Merge collinear axis-aligned pieces [(lo,hi,pos)] into full spans."""
    from collections import defaultdict as _dd
    groups = _dd(list)
    for lo, hi, pos in pieces:
        groups[round(pos / tol)].append((lo, hi, pos))
    out = []
    for g in groups.values():
        g.sort()
        lo, hi = g[0][0], g[0][1]
        pos = sum(p[2] for p in g) / len(g)
        cov = 0.0
        for l2, h2, _ in g:
            if l2 <= hi + 2:
                hi = max(hi, h2)
            else:
                cov += hi - lo
                out.append((lo, hi, pos)); lo, hi = l2, h2
        out.append((lo, hi, pos))
    return out

def _find_plot_boxes(page):
    """Plot grid boxes from groups of equal-span horizontal + vertical lines."""
    from collections import defaultdict as _dd
    Hp, Vp = [], []
    for d in page.get_drawings():
        for it in d["items"]:
            for (xa, ya), (xb, yb) in _item_to_segments(it):
                if abs(ya - yb) < 0.8 and abs(xa - xb) > 4:
                    Hp.append((min(xa, xb), max(xa, xb), (ya + yb) / 2))
                elif abs(xa - xb) < 0.8 and abs(ya - yb) > 4:
                    Vp.append((min(ya, yb), max(ya, yb), (xa + xb) / 2))
    H = [h for h in _merge_collinear(Hp) if h[1] - h[0] > 20]
    V = [v for v in _merge_collinear(Vp) if v[1] - v[0] > 20]
    groups = _dd(list)
    for x0, x1, y in H:
        groups[(round(x0), round(x1))].append(y)
    boxes = []
    for (x0, x1), ys in groups.items():
        if len(ys) < 4 or (x1 - x0) <= 60:
            continue
        # split into y-clusters: vertically stacked graphs can share the same
        # x-extent (one column layout) — a large gap separates the grids
        ys = sorted(ys)
        gaps = [b - a for a, b in zip(ys, ys[1:])]
        med = sorted(gaps)[len(gaps) // 2] if gaps else 0
        clusters, cur = [], [ys[0]]
        for a, b in zip(ys, ys[1:]):
            if (b - a) > max(40.0, 3.0 * med):
                clusters.append(cur); cur = [b]
            else:
                cur.append(b)
        clusters.append(cur)
        for cl in clusters:
            if len(cl) < 4: continue
            y0, y1 = min(cl), max(cl)
            if y1 - y0 <= 60: continue
            nv = sum(1 for (vy0, vy1, vx) in V
                     if x0 - 2 <= vx <= x1 + 2 and vy0 <= y0 + 5 and vy1 >= y1 - 5)
            if nv >= 2:
                boxes.append((float(x0), float(x1), float(y0), float(y1)))
    # --- vertical-extent driven detection -----------------------------------
    # Vertically-stacked plots that share an x-extent but have only sparse
    # horizontal gridlines can get merged/dropped by the y-clustering above.
    # The left/right border verticals cleanly encode each figure's y-span, so
    # recover any figure framed by a left+right edge of matching extent.
    def _round_e(lst):
        return sorted(set((round(a), round(b)) for a, b in lst))
    for (x0, x1), ys in groups.items():
        if (x1 - x0) <= 60 or len(ys) < 3:
            continue
        left = [(vy0, vy1) for (vy0, vy1, vx) in V if abs(vx - x0) <= 3]
        right = [(vy0, vy1) for (vy0, vy1, vx) in V if abs(vx - x1) <= 3]
        if not left or not right:
            continue
        rights = _round_e(right)
        for (ly0, ly1) in _round_e(left):
            if (ly1 - ly0) <= 40:
                continue
            if not any(abs(ry0 - ly0) < 8 and abs(ry1 - ly1) < 8
                       for (ry0, ry1) in rights):
                continue
            nh = sum(1 for y in ys if ly0 - 5 <= y <= ly1 + 5)
            if nh < 3:
                continue
            boxes.append((float(x0), float(x1), float(ly0), float(ly1)))
    # de-duplicate near-identical boxes
    uniq = []
    for b in boxes:
        if not any(all(abs(b[i] - u[i]) < 5 for i in range(4)) for u in uniq):
            uniq.append(b)

    # ── split side-by-side (2-up / n-up) layouts ───────────────────────────
    # Infineon prints two (sometimes three) diagrams in a row.  Their grids sit
    # at the SAME y, so the row's horizontal gridlines are two abutting segments
    # ([x0->div] and [div->x1]) that _merge_collinear joins into one full-width
    # line, collapsing both plots into a single wide box.  A genuine column
    # DIVIDER is an internal full-height vertical at which the gridlines
    # TERMINATE (segment endpoints land on it); an ordinary CENTRE GRIDLINE of a
    # single plot is instead CROSSED by the gridlines.  Splitting only on the
    # "endpoints, not crossings" signature keeps single-plot datasheets intact.
    def _split_columns(box):
        x0, x1, y0, y1 = box
        bh = y1 - y0
        if bh <= 0 or (x1 - x0) <= 120:
            return [box]
        # internal verticals spanning ~the full box height
        divs = sorted(vx for (vy0, vy1, vx) in V
                      if x0 + 12 < vx < x1 - 12
                      and vy0 <= y0 + 0.12 * bh and vy1 >= y1 - 0.12 * bh)
        # collapse a twin "right-border + left-border" pair into one divider
        clustered = []
        for vx in divs:
            if not clustered or vx - clustered[-1] > 12:
                clustered.append(vx)
        real = []
        for vx in clustered:
            endpoint = sum(1 for (hx0, hx1, hy) in Hp
                           if y0 - 3 <= hy <= y1 + 3 and (hx1 - hx0) > 40
                           and (abs(hx0 - vx) < 4 or abs(hx1 - vx) < 4))
            crossing = sum(1 for (hx0, hx1, hy) in Hp
                           if y0 - 3 <= hy <= y1 + 3 and (hx1 - hx0) > 40
                           and hx0 < vx - 4 and hx1 > vx + 4)
            if endpoint >= 2 and endpoint >= crossing:
                real.append(vx)
        if not real:
            return [box]
        xs = [x0] + real + [x1]
        return [(xs[k], xs[k + 1], y0, y1) for k in range(len(xs) - 1)
                if xs[k + 1] - xs[k] > 60]

    split = []
    for b in uniq:
        split.extend(_split_columns(b))
    out = []
    for b in split:
        if not any(all(abs(b[i] - u[i]) < 5 for i in range(4)) for u in out):
            out.append(b)
    return out

def _collect_curve_segments(page, box, margin=2.0):
    """Vector segments of the data curves inside the plot box.

    Axis-aligned segments are normally excluded (grid / axes / ticks), but
    derating curves have exactly-horizontal flat regions (e.g. constant
    rated current up to 75 °C).  Those are rescued by stroke width: data
    curves are drawn noticeably thicker than the grid (e.g. 1.0 pt vs
    0.35 pt in NXP/Nexperia sheets)."""
    x0, x1, y0, y1 = box

    def _inside(ax, ay, bx, by):
        both = (x0 - margin <= ax <= x1 + margin and x0 - margin <= bx <= x1 + margin
                and y0 - margin <= ay <= y1 + margin and y0 - margin <= by <= y1 + margin)
        # A data line drawn flush to (or slightly past) the axes overshoots the
        # box by a few px at each end; rescue it when its midpoint is inside —
        # but only for genuinely long segments, so short axis tick marks that
        # poke just outside the box are NOT mistaken for a flat curve.
        mxp, myp = (ax + bx) / 2.0, (ay + by) / 2.0
        mid = (x0 - margin <= mxp <= x1 + margin and y0 - margin <= myp <= y1 + margin
               and (abs(ax - bx) > 8 or abs(ay - by) > 8))
        return both or mid

    # pass 1: gather everything; measure the grid stroke width
    paths, grid_ws = [], []
    for d in page.get_drawings():
        wd = float(d.get("width") or 0.0)
        for it in d["items"]:
            for seg in _item_to_segments(it):
                (ax, ay), (bx, by) = seg
                if not _inside(ax, ay, bx, by): continue
                h_long = abs(ay - by) < 0.4 and abs(ax - bx) > 8
                v_long = abs(ax - bx) < 0.4 and abs(ay - by) > 8
                paths.append((ax, ay, bx, by, wd, it[0], h_long, v_long))
                if (h_long or v_long) and it[0] == "l" and wd > 0:
                    grid_ws.append(wd)
    grid_w = sorted(grid_ws)[len(grid_ws) // 2] if grid_ws else 0.0

    segs = []
    for ax, ay, bx, by, wd, kind, h_long, v_long in paths:
        if h_long or v_long:
            # rescue thick-stroked flat curve regions (never border/axis lines)
            if h_long:
                on_border = min(abs(ay - y0), abs(ay - y1)) < 1.2
            else:
                on_border = min(abs(ax - x0), abs(ax - x1)) < 1.2
            if not (kind == "l" and grid_w > 0 and wd >= 1.5 * grid_w
                    and not on_border):
                continue
        segs.append((ax, ay, bx, by))
    return segs

def _track_curves(segs, box, n_cols=200, ytol=2.0, maxgap=25):
    """
    Column-sampled curve tracking with closest-first global assignment.
    Handles solid + dashed curves, multiple curves, legend-glyph noise
    (short spurious traces are dropped by the 30%-span filter).
    Returns list of dicts {x_pdf: y_pdf} plus per-curve column-coverage.
    """
    x0, x1, y0, y1 = box
    xs = np.linspace(x0, x1, n_cols)
    step = (x1 - x0) / n_cols
    cols = [[] for _ in range(n_cols)]
    for ax, ay, bx, by in segs:
        lo, hi = (ax, bx) if ax <= bx else (bx, ax)
        i0 = max(int(np.searchsorted(xs, lo - step)), 0)
        i1 = min(int(np.searchsorted(xs, hi + step)), n_cols)
        for i in range(i0, i1):
            xc = xs[i]
            if not (lo - step <= xc <= hi + step): continue
            if abs(bx - ax) > 1e-9:
                t = min(max((xc - ax) / (bx - ax), 0), 1)
                yv = ay + t * (by - ay)
            else:
                yv = (ay + by) / 2
            cols[i].append(yv)
    ccols = []
    for ys in cols:
        ys = sorted(ys); cl = []
        for y in ys:
            if cl and y - cl[-1][-1] <= ytol: cl[-1].append(y)
            else: cl.append([y])
        ccols.append([float(np.mean(c)) for c in cl])
    # active: [curve_idx, last_col, last_y, recent_(col,y)_history]
    curves, active = [], []
    HIST = 6  # points used to estimate local trajectory slope

    def _predict(hist, target_col):
        """Linear extrapolation from recent (col, y) history to target_col.
        Falls back to the last known y when there isn't enough history for a
        slope estimate (curve just started, or has been flat)."""
        if len(hist) < 2:
            return hist[-1][1] if hist else 0.0
        cs = np.array([h[0] for h in hist], dtype=float)
        ys_ = np.array([h[1] for h in hist], dtype=float)
        if cs[-1] == cs[0]:
            return ys_[-1]
        sl, ic = np.polyfit(cs, ys_, 1)
        return sl * target_col + ic

    for i in range(n_cols):
        cands = []
        for ai, a in enumerate(active):
            gap = i - a[1]
            if gap > maxgap: continue
            # Predicted Y uses the curve's recent trajectory (slope), not just
            # its last point.  This is what lets the tracker follow a curve
            # THROUGH a crossing: at the crossing column two curves' last-known
            # Y values are nearly identical, but their predicted (extrapolated)
            # Y values diverge because they're moving in different directions —
            # nearest-to-last-point tracking is ambiguous there, nearest-to-
            # predicted-trajectory is not.
            pred_y = _predict(a[3], i)
            tol = 3.0 + 0.9 * gap
            for j, y in enumerate(ccols[i]):
                d = abs(y - pred_y)
                if d <= tol: cands.append((d, ai, j))
        cands.sort()
        used_a, used_j = set(), set()
        for d, ai, j in cands:
            if ai in used_a or j in used_j: continue
            used_a.add(ai); used_j.add(j)
            y_new = ccols[i][j]
            curves[active[ai][0]][xs[i]] = y_new
            active[ai][1], active[ai][2] = i, y_new
            hist = active[ai][3]; hist.append((i, y_new))
            if len(hist) > HIST: hist.pop(0)
        # convergence sharing: an unmatched curve very close to a taken cluster
        for ai, a in enumerate(active):
            if ai in used_a or i - a[1] > 2: continue
            pred_y = _predict(a[3], i)
            for j, y in enumerate(ccols[i]):
                if abs(y - pred_y) <= ytol * 1.6:
                    curves[a[0]][xs[i]] = y
                    a[1], a[2] = i, y
                    hist = a[3]; hist.append((i, y))
                    if len(hist) > HIST: hist.pop(0)
                    break
        active = [a for a in active if i - a[1] <= maxgap]
        for j, y in enumerate(ccols[i]):
            if j not in used_j:
                curves.append({xs[i]: y})
                active.append([len(curves) - 1, i, y, [(i, y)]])
    out = []
    bh = (y1 - y0) or 1.0
    bw = (x1 - x0) or 1.0

    # ── STITCH fragments split at curve crossings ──────────────────────
    # When two physical curves cross, the nearest-Y assignment above can
    # jump tracks at the crossing, splitting ONE physical curve into two
    # fragments that meet end-to-end (fragment A ends where fragment B
    # begins, or vice-versa) instead of producing one continuous trace.
    # This is independent of manufacturer / graph type — it is a generic
    # consequence of column-sampled tracking near a crossing.  Detect pairs
    # of fragments whose endpoints are close in BOTH x and y (and whose
    # implied connecting slope is consistent with each fragment's own local
    # slope, so we don't accidentally bridge two unrelated curves), and
    # merge them into a single curve dict.
    def _frag_slope_near_end(c, at_end, n=4):
        """Local slope of curve dict `c` near its left (at_end='lo') or
        right (at_end='hi') endpoint, estimated from up to n nearest points."""
        xs_ = sorted(c.keys())
        if len(xs_) < 2:
            return 0.0
        pts = xs_[:n] if at_end == "lo" else xs_[-n:]
        if len(pts) < 2:
            return 0.0
        ys_ = [c[p] for p in pts]
        sl, _ = np.polyfit(pts, ys_, 1)
        return float(sl)

    frags = [dict(c) for c in curves if c]
    merged_any = True
    while merged_any and len(frags) > 1:
        merged_any = False
        best = None  # (cost, i, j, order) order: 'i_then_j' or 'j_then_i'
        for i in range(len(frags)):
            xi = sorted(frags[i].keys())
            if not xi:
                continue
            for j in range(len(frags)):
                if i == j:
                    continue
                xj = sorted(frags[j].keys())
                if not xj:
                    continue
                # candidate: frags[i] ends where frags[j] begins
                gap_x = xj[0] - xi[-1]
                if not (0 <= gap_x <= 6.0):
                    continue
                y_i_end = frags[i][xi[-1]]
                y_j_start = frags[j][xj[0]]
                gap_y = abs(y_j_start - y_i_end)
                if gap_y > 0.10 * bh:
                    continue
                # slope continuity: the connecting segment's slope should be
                # broadly consistent with each fragment's own trend near the
                # join (loose tolerance — crossings bend, they don't reverse
                # direction outright within a few px).
                sl_i = _frag_slope_near_end(frags[i], "hi")
                sl_j = _frag_slope_near_end(frags[j], "lo")
                if gap_x > 1e-6:
                    sl_join = (y_j_start - y_i_end) / gap_x
                else:
                    sl_join = sl_i
                # cost combines positional gap and slope mismatch
                slope_mismatch = abs(sl_join - sl_i) + abs(sl_join - sl_j)
                cost = gap_y + 0.15 * bh * min(slope_mismatch, 4.0) / 4.0
                if best is None or cost < best[0]:
                    best = (cost, i, j)
        if best is not None and best[0] <= 0.14 * bh:
            _, i, j = best
            frags[i].update(frags[j])
            del frags[j]
            merged_any = True
    curves = frags

    for c in curves:
        span = max(c) - min(c)
        if span <= 0.30 * bw:
            continue
        ys = list(c.values())
        yspan = max(ys) - min(ys)
        ymean = sum(ys) / len(ys)
        edge = min(abs(ymean - y0), abs(ymean - y1))
        n_expected = span / step
        coverage = len(c) / max(n_expected, 1)
        # Drop plot-frame / axis-limit lines pinned to the top or bottom of the
        # box — these are not data curves but surface as spurious constant
        # "curves" (e.g. 3.005 / 0.002 on single-curve ST plots, or a flat line
        # mislabelled "Eon"), inflating the curve count.  Two signatures:
        #   (1) a clean near-flat border line (yspan ~0, right at the edge);
        #   (2) a sparse border+tick trace whose mean sits on the edge.
        # Genuine data curves sit well inside the frame (large edge distance),
        # so neither rule can touch them.
        if (yspan < 0.03 * bh and edge < 0.02 * bh) \
                or (edge < 0.04 * bh and coverage < 0.30):
            continue
        # Also drop curves that lie almost entirely on a frame border (within
        # ~5% of the box height from top or bottom edge).  These arise from
        # the vector tracker following the axis border or a near-border gridline.
        if yspan < 0.05 * bh and edge < 0.06 * bh:
            continue
        # NOTE: a "flat + spans most of the X width => gridline" rule was
        # tried here and removed.  It is both unnecessary and actively
        # harmful: _collect_curve_segments (above, earlier in the pipeline)
        # already separates data curves from gridlines by STROKE WIDTH
        # (data curves are drawn noticeably thicker than gridlines in every
        # manufacturer's PDF vector layer), so anything reaching this point
        # has already survived that test and is real data.  A genuinely flat
        # data curve — e.g. Ciss saturating near a constant value across
        # most of the VDS range, which is normal MOSFET physics — has
        # exactly the same (low yspan, wide xspan) signature as a gridline
        # and was being deleted by mistake.
        # Drop sparse traces very close to the top or bottom border — these
        # are axis frame lines that leaked through (not caught by the strict
        # 3% + 2% threshold above because they have slight curvature from
        # tracking noise).  Key signature: extremely close to border AND low
        # coverage (real curves tend to have higher coverage).
        if edge < 0.04 * bh and coverage < 0.50:
            continue
        out.append((c, coverage))

    # ── SUBSET-RETRACE DEDUP ────────────────────────────────────────────────
    # A real artifact has a distinctive geometric signature that's safe to
    # remove without touching genuinely distinct (but converging) curves:
    # the tracker loses a curve partway through (e.g. at a tick label or a
    # crossing), then re-latches onto the SAME physical line for the
    # remainder.  This produces a short fragment whose entire X-range is a
    # SUBSET of a longer curve's X-range, AND whose Y values match the
    # longer curve almost exactly (near-zero pixel difference) over their
    # entire overlap — not just "close", but coincident.
    #
    # This is categorically different from two distinct curves that merely
    # converge over PART of their range (e.g. Ciss/Coss/Crss meeting at high
    # VDS, or Zth duty-cycle curves meeting at long pulse times): those
    # curves each cover most/all of the X range in their own right and only
    # match closely near one end, not coincide everywhere they overlap.
    #
    # Requiring (a) strict X-range containment and (b) near-zero — not just
    # "small" — Y agreement over the ENTIRE overlap makes this safe: it
    # cannot fire on two curves that are merely close, only on one curve
    # that is a literal retrace of another.
    if len(out) > 1:
        keep = [True] * len(out)
        for i in range(len(out)):
            if not keep[i]:
                continue
            xs_i = sorted(out[i][0].keys())
            x0i, x1i = xs_i[0], xs_i[-1]
            for j in range(len(out)):
                if i == j or not keep[j]:
                    continue
                xs_j = sorted(out[j][0].keys())
                x0j, x1j = xs_j[0], xs_j[-1]
                # j's X-range must be properly contained within i's (with a
                # small tolerance), i.e. j starts no earlier and ends no
                # later than i — a genuine subset, not just overlapping.
                if not (x0j >= x0i - 1.0 and x1j <= x1i + 1.0):
                    continue
                if (x1j - x0j) >= (x1i - x0i) - 1.0:
                    continue  # same span, not a proper subset
                shared = [x for x in xs_j if x0i <= x <= x1i]
                if len(shared) < 5:
                    continue
                diffs = [abs(out[i][0].get(x, out[j][0][x]) - out[j][0][x])
                         for x in shared if x in out[i][0]]
                if len(diffs) < 5:
                    continue
                # near-zero (coincident), not just "close" — this is what
                # distinguishes a retrace from a converging-but-distinct
                # curve, which would show real separation over most of the
                # overlap and only approach zero near the convergence point.
                med_diff = sorted(diffs)[len(diffs) // 2]
                p90_diff = sorted(diffs)[int(len(diffs) * 0.9)]
                if med_diff < 0.012 * bh and p90_diff < 0.025 * bh:
                    keep[j] = False  # j is a retrace fragment of i
        out = [o for o, k in zip(out, keep) if k]
    return out

_TEMP_X_PAT = re.compile(
    r"(versus|vs\.?|as\s+a\s+function\s+of)\s+(the\s+)?(initial\s+)?"
    r"((junction|case|heatsink|heat\s*sink|ambient|virtual\s+junction)\s+)?temperature"
    r"|=\s*f\s*\(\s*T\s*(vj|j|c|h|a|amb)?\s*\)"
    r"|current\s+derating|power\s+derating", re.I)
_FIGNO_PAT = re.compile(r"fig(?:ure)?\.?\s*(\d+)", re.I)

# A figure is an ENERGY graph if its caption / axis text mentions switching
# energy (Eon / Eoff / Etot / Esw) or switching loss.  These are digitised
# regardless of what the X axis is (gate resistance, drain current, or Tj).
_ENERGY_PAT = re.compile(
    r"\bE\s*[_(]?\s*(on|off|tot|total|sw|rr|oss|ts)\b"
    r"|\bE(on|off|ts|sw)\b"
    r"|switching\s+energ(y|ies)|switching\s+loss(es)?"
    r"|energy\s+loss(es)?|turn.?on\s+energy|turn.?off\s+energy", re.I)

# A figure is a THERMAL-IMPEDANCE graph (Zth vs pulse-time tp, usually log-log
# with a family of duty-cycle curves) when its caption mentions transient
# thermal impedance / Zth / =f(tp).  These are digitised against tp [s].
_ZTH_PAT = re.compile(
    r"\bZ\s*[_(]?\s*th\b"
    r"|thermal\s+impedance"
    r"|transient\s+thermal"
    r"|=\s*f\s*\(\s*t\s*[_(]?\s*[pP]\b", re.I)

# A figure is a CAPACITANCE graph (Ciss/Coss/Crss vs VDS) when its caption
# mentions capacitance(s) plotted against drain-source voltage.  The bare
# stored-energy Eoss=f(VDS) is NOT a capacitance graph — guard against it via
# _ENERGY_PAT in the classifier so it is routed to the Energy sheet instead.
_CAP_PAT = re.compile(
    r"\bcapacitance(s)?\b"
    r"|\bC\s*[_(]?\s*(iss|oss|rss)\b"
    r"|\bC\s*=\s*f\s*\(\s*V\s*[_(]?\s*DS\b", re.I)

def _is_graph_target(text):
    """True if a caption belongs to one of the digitised figure families:
    temperature-axis, switching/stored energy, thermal impedance (Zth=f(tp)),
    or capacitance (C=f(VDS))."""
    t = text or ""
    return bool(_TEMP_X_PAT.search(t) or _ENERGY_PAT.search(t)
                or _ZTH_PAT.search(t) or _CAP_PAT.search(t))

def _norm_unit(u):
    """Normalise a raw parenthesised unit token to a clean display unit."""
    t = (u or "").strip()
    low = t.lower()
    if low in ("c", "\u00b0c", "degc", "\u00b0 c"):     return "\u00b0C"
    if low in ("ohm", "ohms", "\u03a9", "\u2126"):       return "\u03a9"
    if low in ("mohm", "m\u03a9", "m\u2126"):            return "m\u03a9"
    if low in ("k", "kelvin"):                            return "K"
    return t

def _fig_x_unit(caption, ax_title=None, x_title=None):
    """Printed unit of the X axis.

    Priority: (1) the unit printed in parentheses in the x-axis title — this is
    authoritative and works for energy graphs (RG in \u03a9, ID in A) as well as
    temperature graphs;  (2) temperature detection from the caption / axis title
    \u2192 \u00b0C;  (3) caption keyword fallback (gate resistance \u2192 \u03a9, drain current \u2192 A).
    """
    for t in (x_title, ax_title):
        m = re.search(r"\(\s*([^)]{1,6}?)\s*\)\s*$", (t or "").strip())
        if m:
            u = _norm_unit(m.group(1))
            if re.fullmatch(r"m?\u03a9|\u00b0C|K|A|mA|V|mV|ns|\u00b5s|nC|\u00b5C|%|W", u, re.I):
                return u
    if (_TEMP_X_PAT.search(caption or "")
            or (ax_title and _XTITLE_TEMP.search(ax_title))
            or re.search(r"versus\s+.*temperature|=\s*f\s*\(\s*T", caption or "", re.I)):
        return "\u00b0C"
    cl = (caption or "").lower()
    if "gate resistance" in cl: return "\u03a9"
    if "drain current"  in cl: return "A"
    # Thermal impedance: X = pulse time tp in seconds; capacitance: X = VDS in V.
    if _ZTH_PAT.search(caption or "") or re.search(r"=\s*f\s*\(\s*t\s*[_(]?\s*p\b", cl):
        return "s"
    if _CAP_PAT.search(caption or "") and not _ENERGY_PAT.search(caption or ""):
        return "V"
    if _ENERGY_PAT.search(caption or "") and re.search(
            r"oss|stored\s+energy|output\s+capacitance", cl):
        return "V"
    src = " ".join(t for t in (ax_title, x_title) if t)
    m = re.search(r"\(\s*(m?\u03a9|ohm|mA|A|mV|V|ns|\u00b5s|nC)\s*\)", src, re.I)
    return _norm_unit(m.group(1)) if m else ""

def _classify_fig(caption):
    cl = caption.lower()
    if _ENERGY_PAT.search(caption):            return "energy"
    # Thermal impedance (Zth=f(tp)) is checked before capacitance/relative so a
    # caption like "transient thermal impedance" never falls through to generic.
    if _ZTH_PAT.search(caption):               return "thermal_z"
    # Capacitance (C=f(VDS)) — only when it is NOT an energy graph (Eoss already
    # returned above), so "Coss stored energy" cannot be mistaken for it.
    if _CAP_PAT.search(caption):               return "capacitance"
    if "relative" in cl:                       return "relative"
    if "power" in cl and "dissipation" in cl:  return "power"
    if "on-resistance" in cl or "on resistance" in cl or "rds(on)" in cl: return "resistance"
    if "derating" in cl or re.search(r"\bcurrent\b", cl): return "current"
    return "generic"

def _deriv_x_name(caption, x_title=None):
    cap = caption or ""
    # The X variable is the quantity AFTER "as a function of" / "versus" / "vs"
    # — the part before it is the Y quantity.  Captions like "drain current as
    # a function of case temperature" must give X=Tc, not X=ID.  Match keywords
    # on this RHS clause; fall back to the whole caption when there is no split.
    parts = re.split(r"\s+(?:versus|vs\.?|as\s+a\s+function\s+of)\s+",
                     cap, flags=re.I)
    x_clause = parts[-1] if len(parts) > 1 else cap
    src = (x_title or "") + " " + x_clause
    cl = src.lower()
    # Robust descriptive-phrase matches first — these survive even when the
    # axis-title band is jumbled together with an in-plot conditions line
    # (the clean caption still carries the phrase).
    if "gate resistance" in cl: return "RG"
    if "drain current"  in cl: return "ID"
    # Thermal impedance is plotted against pulse time tp [s]; capacitance against
    # drain-source voltage VDS [V].  Detect these before the temperature fallback
    # so neither axis is mislabelled "Tj".
    if _ZTH_PAT.search(cl) or re.search(r"=\s*f\s*\(\s*t\s*[_(]?\s*p\b", cl):
        return "tp (pulse)"
    if re.search(r"=\s*f\s*\(\s*v\s*[_(]?\s*ds\b", cl) or (
            _CAP_PAT.search(cap) and not _ENERGY_PAT.search(cap)):
        return "VDS"
    if "mount" in cl or re.search(r"\bT\s*[_,]?\s*[mM]\b", src):                  return "TM (mount)"
    if "heatsink" in cl or "heat sink" in cl or re.search(r"\bT\s*[_,]?\s*[hH]\b", src): return "Th (heatsink)"
    if "case" in cl or re.search(r"\bT\s*[_,]?\s*[cC]\b", src):                    return "Tc (case)"
    if "ambient" in cl or re.search(r"\bT\s*[_,]?\s*[aA](mb)?\b", src):            return "Ta (ambient)"
    if "junction" in cl or "temperature" in cl:                                    return "Tj (junction)"
    # Bare-symbol axis title ("RG (\u03a9)", "ID (A)") — only trusted when the title
    # is NOT contaminated by '=' style test conditions.
    if x_title and "=" not in x_title:
        mt = re.match(r"\s*(R\s*G|I\s*D|V\s*GS|V\s*DS)\b", x_title, re.I)
        if mt:
            return re.sub(r"\s+", "", mt.group(1)).upper()
    # Energy graphs are plotted against RG / ID / VDS, never temperature.  The
    # RG ("gate resistance") and ID ("drain current") cases are handled above;
    # output-capacitance stored-energy (Eoss) is plotted vs VDS.  Fall back to a
    # neutral label instead of mislabelling an energy axis "Tj".  Energy wording
    # usually sits in the Y part of the caption, so search the whole caption.
    full = (x_title or "") + " " + cap
    if _ENERGY_PAT.search(full):
        if re.search(r"\bE\s*[_(]?\s*oss\b|stored\s+energy|output\s+capacitance",
                     full, re.I):
            return "VDS"
        return "X"
    return "Tj (junction)"

def _deriv_y_name(kind, caption, y_title=None):
    if y_title: return y_title
    if kind == "power":      return "Ptot (W)"
    if kind == "current":    return "IF (A)"
    if kind == "thermal_z":  return "ZthJC (K/W)"
    if kind == "capacitance":return "C (pF)"
    # Infineon / function-style caption "<Y>=f(<X>)" — the token immediately
    # before "=f(" is the plotted quantity (RDS(on), VDSS, EAS, …).
    mfun = re.search(r"([A-Za-z][\w()/,.\-]*?)\s*=\s*f\s*\(", caption or "")
    if mfun:
        yq = mfun.group(1).strip(" .,-")
        if yq and len(yq) <= 12:
            return f"{yq} (as printed on Y axis)"
    # use the caption text before "versus"/"as a function of"
    m = re.split(r"\s+(?:versus|vs\.?|as\s+a\s+function\s+of)\s+", caption, flags=re.I)
    lead = re.sub(r"^fig(?:ure)?\.?\s*\d+\.?\s*", "", m[0], flags=re.I).strip()
    return (lead[:60] + " (ratio / as printed on Y axis)") if lead else "Y (as printed)"

def _infineon_diagram_captions(page):
    """Infineon datasheets label plots 'Diagram N: <title>' (header above the
    plot) with a 'Y=f(X); conditions' function line below it, and render the
    sub/superscripts of that function on a separate baseline. Reconstruct a
    clean caption per diagram ('Diagram N: <title> Y=f(X); ...') so the normal
    temperature/energy qualification, classification and digitisation work.
    Returns [(diagram_no, caption, bbox)] like _figure_captions."""
    try:
        raw = page.get_text("words")  # (x0, y0, x1, y1, text, ...)
    except Exception:
        return []
    if not raw:
        return []
    def _wd(w):
        return {"x0": w[0], "x1": w[2], "top": w[1], "bottom": w[3], "text": w[4]}
    words = [_wd(w) for w in raw]
    mid = page.rect.width / 2.0
    out = []
    for lo, hi in ((0, mid), (mid, page.rect.width)):
        col = [w for w in words if lo <= (w["x0"] + w["x1"]) / 2 < hi]
        if not col:
            continue
        col.sort(key=lambda w: ((w["top"] + w["bottom"]) / 2, w["x0"]))
        lines = []
        for w in col:
            cy = (w["top"] + w["bottom"]) / 2
            if lines and abs(cy - lines[-1]["cy"]) <= 5:
                L = lines[-1]; L["ws"].append(w)
                L["cy"] = (L["cy"] * L["n"] + cy) / (L["n"] + 1); L["n"] += 1
            else:
                lines.append({"cy": cy, "n": 1, "ws": [w]})
        for L in lines:
            L["text"] = _join_subscripts(L["ws"]); L["y"] = L["cy"]
        titles = [L for L in lines if re.match(r"\s*Diagram\s*\d", L["text"])]
        for ti, L in enumerate(titles):
            mt = re.match(r"\s*Diagram\s*(\d+)\s*:?\s*(.*)", L["text"])
            if not mt:
                continue
            dno = int(mt.group(1)); title = mt.group(2).strip()
            y_next = titles[ti + 1]["y"] if ti + 1 < len(titles) else 1e9
            func = ""
            for L2 in lines:
                if L["y"] < L2["y"] < y_next and re.search(r"=\s*f\s*\(", L2["text"]):
                    func = L2["text"]; break
            cap = collapse(f"Diagram {dno}: {title} {func}".strip())
            ws = L["ws"]
            bb = [min(w["x0"] for w in ws), min(w["top"] for w in ws),
                  max(w["x1"] for w in ws), max(w["bottom"] for w in ws)]
            out.append((dno, cap, bb))
    out.sort(key=lambda c: c[0])
    return out

def _figure_captions(page):
    """
    Caption lines starting with 'Figure N.' / 'Fig. N.' joined with their
    wrapped continuation lines.  Returns [(fig_no, caption, bbox)].
    Handles both caption-above-figure (ST/onsemi) and below (WeEN/Nexperia).
    """
    raw = []
    try:
        d = page.get_text("dict")
    except Exception:
        return []
    # Infineon 'Diagram N:' layout — use the dedicated reconstructor.
    # Infineon's text layer separates words with C0 control chars (e.g. \x03),
    # so the raw page text reads "Diagram\x031:\x03Power..." and a plain
    # "Diagram\s*\d+:" search misses it (\x03 is not \s).  Sanitise the control
    # chars to spaces before testing, otherwise delegation silently fails and
    # only the one generic temperature caption per page is recovered.
    try:
        _pg = re.sub(r"[\x00-\x1f]+", " ", page.get_text() or "")
        if re.search(r"Diagram\s*\d+\s*:", _pg):
            inf = _infineon_diagram_captions(page)
            if inf:
                return inf
    except Exception:
        pass
    for b in d.get("blocks", []):
        if b.get("type") != 0: continue
        for l in b.get("lines", []):
            txt = " ".join(s.get("text", "") for s in l.get("spans", [])).strip()
            if txt:
                raw.append((txt, list(l["bbox"])))
    raw.sort(key=lambda r: (round(r[1][1], 1), r[1][0]))
    caps = []
    used = set()
    for i, (txt, bb) in enumerate(raw):
        m = _FIGNO_PAT.match(txt)
        if not m or i in used: continue
        # ── Filter out inline figure references ─────────────────────────
        # Datasheets often contain "(see Figure 14. Test circuit...)" inside
        # table cells or test-condition text.  When the PDF text layer breaks
        # this across lines, "Figure 14. Test circuit..." appears as a
        # separate line that _FIGNO_PAT.match() catches.  Detect this by:
        #   (1) Checking whether the preceding text ends with "(see", "see",
        #       "(refer to", "and" (continuation of another Figure ref), etc.
        #   (2) Checking whether "Figure N." appears inside parentheses in the
        #       full page text (reconstructed from nearby raw text).
        is_inline_ref = False
        fig_tag = m.group(0)  # e.g. "Figure 14"
        for j2 in range(max(0, i - 6), i):
            t2, b2 = raw[j2]
            # same column and within a few px vertically above
            xov = min(b2[2], bb[2]) - max(b2[0], bb[0])
            vgap = bb[1] - b2[3]
            if xov > 10 and -2 <= vgap <= 10:
                t2_clean = t2.rstrip().rstrip(")").rstrip()
                if re.search(r"\(\s*see\s*$|see\s*$|\(refer\s+to\s*$|\(ref\.?\s*$"
                             r"|and\s*$", t2_clean, re.I):
                    is_inline_ref = True
                    break
                # Also check if this preceding line itself contains "(see"
                # and has no closing ")" — the figure ref is a continuation
                if re.search(r"\(\s*see\s", t2, re.I) and ")" not in t2.split("see")[-1]:
                    is_inline_ref = True
                    break
            # also check same-line text that appears before "Figure N."
            same_line = abs(b2[1] - bb[1]) < 3.5
            if same_line and b2[2] <= bb[0] + 5:
                if re.search(r"\(\s*see\s*$|see\s*$|and\s*$", t2.rstrip(), re.I):
                    is_inline_ref = True
                    break
        # Broader check: reconstruct a text window around this line from the
        # raw text, and if "Figure N" appears after an unmatched "(" → inline
        if not is_inline_ref:
            window_parts = []
            for j2 in range(max(0, i - 8), min(len(raw), i + 3)):
                t2, b2 = raw[j2]
                xov = min(b2[2], bb[2]) - max(b2[0], bb[0])
                vgap = abs(bb[1] - b2[1])
                if xov > 5 and vgap < 40:
                    window_parts.append(t2)
            window = " ".join(window_parts)
            # check for unmatched "(" before "Figure N"
            pos = window.find(fig_tag)
            if pos > 0:
                before = window[:pos]
                n_open = before.count("(") - before.count(")")
                if n_open > 0:
                    is_inline_ref = True
        if is_inline_ref:
            used.add(i)
            continue
        fig_no = int(m.group(1)); cap = txt; cbb = bb[:]
        last_bb = bb
        # old NXP/Nexperia layout: "Fig.2" and its caption text are separate
        # spans on the SAME visual line → join horizontally first
        for j in range(i + 1, len(raw)):
            t2, b2 = raw[j]
            if j in used or _FIGNO_PAT.match(t2): continue
            same_line = abs(b2[1] - bb[1]) < 3.5
            hgap = b2[0] - cbb[2]
            if same_line and -2 <= hgap <= 30:
                cap += " " + t2
                cbb = [cbb[0], min(cbb[1], b2[1]), max(cbb[2], b2[2]), max(cbb[3], b2[3])]
                last_bb = b2; used.add(j)
        for j in range(i + 1, len(raw)):
            t2, b2 = raw[j]
            xov = min(b2[2], cbb[2]) - max(b2[0], cbb[0])
            if xov <= 10:
                continue        # different column of a 2-column figure grid
            if _FIGNO_PAT.match(t2):
                break           # next figure in the SAME column
            vgap = b2[1] - last_bb[3]
            if vgap > 10:
                break
            # Continuation lines can vertically overlap the previous line by a
            # few px when leading is tight (e.g. Nexperia wraps "... as a
            # function | of junction temperature ..." with ~2px overlap).  A
            # strict >= -2 dropped the 2nd line, losing the word "temperature"
            # and mis-classifying the whole figure.  -6 keeps true wrapped
            # lines while the same-column (xov) + line-pitch guards prevent
            # grabbing unrelated text.
            if vgap >= -6 and len(cap) < 240:
                cap += " " + t2
                cbb = [min(cbb[0], b2[0]), cbb[1], max(cbb[2], b2[2]), b2[3]]
                last_bb = b2; used.add(j)
        caps.append((fig_no, collapse(cap), cbb))
    # ── fallback: captions WITHOUT a "Figure N." prefix (Diotec, Vishay…) ───
    cap_texts = " ".join(c[1] for c in caps).lower()
    for i, (txt, bb) in enumerate(raw):
        if i in used or _FIGNO_PAT.match(txt): continue
        if not _TEMP_X_PAT.search(txt): continue
        if len(txt) > 8 and txt.lower() in cap_texts: continue
        cap = txt; cbb = bb[:]
        for j in range(i + 1, min(i + 3, len(raw))):
            t2, b2 = raw[j]
            xov = min(b2[2], cbb[2]) - max(b2[0], cbb[0])
            if xov > 10 and 0 <= b2[1] - cbb[3] < 10 and not _TEMP_X_PAT.search(t2):
                cbb = [min(cbb[0], b2[0]), cbb[1], max(cbb[2], b2[2]), b2[3]]
        caps.append((None, collapse(cap), cbb))
    return caps

def _pair_caption_box(captions, boxes):
    """
    Pair figure captions with plot boxes.

    Datasheets place captions consistently on ONE side of their figures
    (Vishay/WeEN: below; ST/onsemi/Infineon: above).  Two complications make
    naive edge-based pairing fail, both seen on ST datasheets:

      • In-plot text (axis labels like "E (µJ)", watermarks "AMxxxxx", and
        condition lines "ID=38A …") sits just inside the frame, so the
        caption's bounding box can grow DOWNWARD into its own plot.  Strict
        "box fully above/below the caption box" tests then reject the correct
        pairing (e.g. Figure 13 was left unpaired entirely).
      • _find_plot_boxes emits two slightly different boxes for the same plot
        (grid-line vs border-extent paths), so the next row's plot may pair
        more cheaply and the whole document flips to the wrong side
        (e.g. Figure 10's caption grabbing Figure 8's energy plot).

    Fix: anchor each caption at its TOP edge — the "Figure N. <title>" line,
    which is clean even when the box grew into the plot — and classify a box
    as ABOVE/BELOW by its vertical CENTRE relative to that anchor, measuring
    the gap from the title line to the box's near edge.  Two document-level
    hypotheses ("plot below caption" for caption-above layouts, "plot above
    caption" for caption-below layouts) are scored; the one pairing more
    captions wins (tie → smaller total gap).
    """
    def _assign(direction):
        cand = []
        for ci, (fn, cap, cbb) in enumerate(captions):
            cap_top, cap_bot = cbb[1], cbb[3]
            cxc = (cbb[0] + cbb[2]) / 2.0
            for bi, b in enumerate(boxes):
                x0, x1, y0, y1 = b
                xov = min(x1, cbb[2]) - max(x0, cbb[0])
                # Overlap test: either the caption covers a fair fraction of the
                # box width (centred captions), OR the caption is mostly CONTAINED
                # within the box's x-span (short, left-aligned Infineon sub-labels
                # such as "Ptot=f(TC)" that sit at the bottom-left of their plot).
                cap_w = max(1.0, cbb[2] - cbb[0])
                contained = xov >= 0.6 * cap_w
                if xov < 0.25 * (x1 - x0) and not contained: continue
                bcy = (y0 + y1) / 2.0
                if direction in ("below", "any") and bcy > cap_top:
                    vgap = y0 - cap_top              # plot BELOW the title line
                elif direction in ("above", "any") and bcy < cap_top:
                    # Plot ABOVE the caption.  Anchor the gap at the STABLE
                    # "Fig N." title line (cap_top), not cap_bot: a caption
                    # whose bbox grew tall — e.g. by absorbing the in-plot
                    # text (watermark / axis labels) of the NEXT figure below
                    # it — would otherwise see its own plot as >140px away and
                    # reject it, then steal the lower figure's plot instead
                    # (Nexperia p9: Fig 18 grabbing Fig 20's Zth plot).
                    vgap = cap_top - y1
                else:
                    continue
                if vgap > 140: continue
                cx_mis = abs((x0 + x1) / 2 - cxc)
                cand.append((abs(vgap) + 0.8 * cx_mis, ci, bi))
        cand.sort()
        uc, ub, pairs, tot = set(), set(), [], 0.0
        for score, ci, bi in cand:
            if ci in uc or bi in ub: continue
            uc.add(ci); ub.add(bi)
            pairs.append((ci, bi)); tot += score
        return pairs, tot

    below, t_b = _assign("below")
    above, t_a = _assign("above")
    if below or above:
        if len(below) != len(above):
            return below if len(below) > len(above) else above
        return below if t_b <= t_a else above
    return _assign("any")[0]

def _merge_number_words(words):
    """Merge adjacent words on the same line ('1 000' → '1000')."""
    words = sorted(words, key=lambda w: (round(w[1], 1), w[0]))
    out, i = [], 0
    while i < len(words):
        x0, y0, x1, y1, t = words[i]
        j = i + 1
        while (j < len(words) and abs(words[j][1] - y0) < 2
               and 0 <= words[j][0] - x1 <= 4):
            t += words[j][4]; x1 = words[j][2]; j += 1
        out.append((x0, y0, x1, y1, t)); i = j
    return out

def _text_axis_labels(page, clip, axis="x"):
    """Numeric tick labels from the PDF *text layer* inside `clip`."""
    try:
        ws = [(w[0], w[1], w[2], w[3], w[4]) for w in page.get_text("words")
              if clip.x0 <= (w[0] + w[2]) / 2 <= clip.x1
              and clip.y0 <= (w[1] + w[3]) / 2 <= clip.y1]
    except Exception:
        return []
    out = []
    for x0, y0, x1, y1, t in _merge_number_words(ws):
        t = t.replace(",", ".").strip()
        if not re.fullmatch(r"-?\d+(\.\d+)?", t): continue
        out.append((float(t), (x0 + x1) / 2 if axis == "x" else (y0 + y1) / 2))
    return out

def _ocr_axis_labels(page, clip, zoom=8, axis="x"):
    """OCR numeric tick labels inside `clip` (for vector-glyph datasheets)."""
    if not _HAS_OCR: return []
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        data = pytesseract.image_to_data(
            img, config="--psm 11 -c tessedit_char_whitelist=0123456789.-",
            output_type=pytesseract.Output.DICT)
    except Exception:
        return []
    out = []
    for i, t in enumerate(data["text"]):
        t = (t or "").strip()
        if not re.fullmatch(r"-?\d+(\.\d+)?", t): continue
        cx = clip.x0 + (data["left"][i] + data["width"][i] / 2) / zoom
        cy = clip.y0 + (data["top"][i] + data["height"][i] / 2) / zoom
        out.append((float(t), cx if axis == "x" else cy))
    return out

def _robust_axis_fit(labels, lo_px, hi_px):
    """RANSAC linear fit value = m*px + c over noisy tick labels."""
    best = None
    n = len(labels)
    for i in range(n):
        for j in range(i + 1, n):
            v1, p1 = labels[i]; v2, p2 = labels[j]
            if abs(p2 - p1) < 0.2 * abs(hi_px - lo_px) or v1 == v2: continue
            m = (v2 - v1) / (p2 - p1); c = v1 - m * p1
            span = abs(m * (hi_px - lo_px))
            tol = max(0.02 * span, 0.02)
            inl = [(v, p) for (v, p) in labels if abs(m * p + c - v) <= tol]
            if best is None or len(inl) > len(best):
                best = inl
    if not best or len(best) < 3:
        return None
    vs = np.array([v for v, p in best]); ps = np.array([p for v, p in best])
    m, c = np.polyfit(ps, vs, 1)
    return float(m), float(c), len(best)

def _calibrate_axis(page, box, axis, interactive, caption):
    """One axis: text → OCR → LOG decade labels → log fallback → user prompt → normalised.
    Returns ((m, c), method, is_log); value = m*px+c, or 10**(m*px+c) if log."""
    x0, x1, y0, y1 = box
    if axis == "x":
        clip = fitz.Rect(x0 - 22, y1 + 1, x1 + 26, y1 + 22); lo, hi = x0, x1
    else:
        clip = fitz.Rect(x0 - 42, y0 - 8, x0 - 1, y1 + 8);   lo, hi = y1, y0
    text_lbl = _text_axis_labels(page, clip, axis)
    fit = _robust_axis_fit(text_lbl, lo, hi)
    if fit: return (fit[0], fit[1]), "text labels", False
    ocr_lbl = _ocr_axis_labels(page, clip, axis=axis)
    fit = _robust_axis_fit(ocr_lbl, lo, hi)
    if fit: return (fit[0], fit[1]), "OCR", False
    # logarithmic axis: decade labels 10^k (text layer, exponent superscripts).
    # X-log is normally off (most X axes are linear and text/OCR labels above
    # already succeeded), but thermal-impedance plots use a log tp[s] X axis
    # and capacitance plots use a log VDS[V] X axis, so enable X-decade
    # reading for those captions to avoid false log fits on other graph types.
    _is_log_cap = (_ZTH_PAT.search(caption or "")
                   or _CAP_PAT.search(caption or "")
                   or re.search(r"\bE\s*[_(]?\s*oss\b|stored\s+energy", caption or "", re.I))
    if axis == "x" and not _is_log_cap:
        dec = []
    else:
        dec = _decade_axis_labels(page, clip, axis=axis)
    fit = _robust_axis_fit(dec, lo, hi) if len(dec) >= 3 else None
    if fit: return (fit[0], fit[1]), "log decade labels", True
    # OCR decade labels: for raster figure images the text layer has nothing;
    # use image-based OCR that specifically reconstructs "10^k" notation.
    if _is_log_cap:
        ocr_dec = _ocr_decade_axis_labels(page, clip, axis=axis)
        fit = _robust_axis_fit(ocr_dec, lo, hi) if len(ocr_dec) >= 3 else None
        if fit: return (fit[0], fit[1]), "OCR decade labels", True
    # Log-fit fallback: when axis labels are plain numbers (1, 10, 100, 1000)
    # without "10^k" notation, the linear fits above fail but a fit in log10
    # space succeeds.  STRICT filter: only keep values that are exact powers of
    # 10 (|log10(v) − round(log10(v))| < 0.05) so spurious OCR reads (e.g. 3,
    # 4, 5 from exponent digits, or random page numbers) cannot trigger a false
    # log calibration.  Also require ≥1 decade span and ≥3 qualifying points.
    def _try_log_fit(labels, method_name):
        if not labels:
            return None, None
        log_pts = []
        for (v, p) in labels:
            if v <= 0:
                continue
            lv = math.log10(v)
            if abs(lv - round(lv)) < 0.05:      # must be a near-exact decade
                log_pts.append((round(lv), p))
        if len(log_pts) < 3:
            return None, None
        log_span = max(pt[0] for pt in log_pts) - min(pt[0] for pt in log_pts)
        if log_span < 1.0:                        # must span ≥1 decade
            return None, None
        f2 = _robust_axis_fit(log_pts, lo, hi)
        return (f2, method_name) if f2 else (None, None)
    # Only apply log fallback for known log-scale graph types (capacitance, Zth)
    if _is_log_cap:
        for lbl_set, meth in ((text_lbl, "text labels (log)"), (ocr_lbl, "OCR (log)")):
            f2, mn = _try_log_fit(lbl_set, meth)
            if f2: return (f2[0], f2[1]), mn, True
    # Wider Y-axis clip fallback: Infineon two-up layouts place the shared Y
    # axis labels far to the left of the right-subbox edge; the standard 42px
    # clip misses them.  Try 100px and 160px widths before giving up.
    if axis == "y":
        for extra in (58, 118):   # adds 58 or 118 px → total 100 or 160 px
            wide_clip = fitz.Rect(x0 - 42 - extra, y0 - 8, x0 - 1, y1 + 8)
            t2 = _text_axis_labels(page, wide_clip, axis)
            fit = _robust_axis_fit(t2, lo, hi)
            if fit: return (fit[0], fit[1]), "text labels (Y wide)", False
            o2 = _ocr_axis_labels(page, wide_clip, axis=axis)
            fit = _robust_axis_fit(o2, lo, hi)
            if fit: return (fit[0], fit[1]), "OCR (Y wide)", False
            if _is_log_cap:
                d2 = _decade_axis_labels(page, wide_clip, axis=axis)
                fit = _robust_axis_fit(d2, lo, hi) if len(d2) >= 3 else None
                if fit: return (fit[0], fit[1]), "log decade labels (Y wide)", True
                od2 = _ocr_decade_axis_labels(page, wide_clip, axis=axis)
                fit = _robust_axis_fit(od2, lo, hi) if len(od2) >= 3 else None
                if fit: return (fit[0], fit[1]), "OCR decade labels (Y wide)", True
                for lbl_set, meth in ((t2, "text labels (log, Y wide)"),
                                      (o2, "OCR (log, Y wide)")):
                    f2, mn = _try_log_fit(lbl_set, meth)
                    if f2: return (f2[0], f2[1]), mn, True
    if interactive and sys.stdin.isatty():
        try:
            ax_name = "X" if axis == "x" else "Y"
            print(f'\n      [graphs] Cannot auto-read the {ax_name}-axis scale of:')
            print(f'               "{caption[:90]}"')
            rng = input(f"               Enter {ax_name}-axis min,max as printed (e.g. 25,175) or blank to skip: ").strip()
            if rng:
                a, b = [float(v) for v in re.split(r"[,;\s]+", rng)[:2]]
                if axis == "x":
                    m = (b - a) / (x1 - x0); c = a - m * x0
                else:
                    m = (b - a) / (y0 - y1); c = a - m * y1
                return (m, c), "user-entered", False
        except Exception:
            pass
    if axis == "x":
        m = 1.0 / (x1 - x0); c = -m * x0
    else:
        m = 1.0 / (y0 - y1); c = -m * y1
    return (m, c), "NORMALISED 0-1", False

def _x_axis_title(page, box):
    """The x-axis title text printed below the tick labels
    (e.g. 'Mount Temperature (\u00b0C)', 'Tamb (\u00b0C)')."""
    x0, x1, y0, y1 = box
    try:
        ws = [w for w in page.get_text("words")
              if x0 - 10 <= (w[0] + w[2]) / 2 <= x1 + 10
              and y1 + 9 <= (w[1] + w[3]) / 2 <= y1 + 32
              and not re.fullmatch(r"-?\d+(\.\d+)?", w[4])]
        ws.sort(key=lambda w: w[0])
        return collapse(" ".join(w[4] for w in ws))[:60]
    except Exception:
        return ""

_XTITLE_TEMP = re.compile(
    r"temperature|\bT\s*[._-]?\s*(amb|a|c|j|vj|m|mt|case|h)?\s*\(\s*\u00b0?\s*C\s*\)", re.I)

def _y_unit_hint(page, box):
    """Axis unit token ('mA', 'µA', 'W', '%', …) printed near the Y axis."""
    x0, x1, y0, y1 = box
    try:
        for w in page.get_text("words"):
            cx, cy = (w[0] + w[2]) / 2, (w[1] + w[3]) / 2
            if (x0 - 45) <= cx <= (x0 + 35) and (y0 - 30) <= cy <= (y0 + 0.6 * (y1 - y0)):
                m = re.fullmatch(r"\(?(mA|µA|uA|kA|A|mV|kV|V|mW|W|pF|nF|%|"
                                 r"µJ|uJ|mJ|nJ|J|nC|µC|uC|mΩ|Ω|mohm|ohm)\)?",
                                 w[4].strip(), re.I)
                if m: return _norm_unit(m.group(1))
    except Exception:
        pass
    return None


def _inplot_phrases(page, box):
    """Group in-plot text words into phrases for curve labels / conditions.
    The right margin is widened slightly beyond the box to catch labels
    placed just outside the plot frame (common for capacitance Ciss/Coss/Crss
    labels and thermal impedance D=0.5 etc.), but ONLY within the plot's own
    vertical band — a wide top margin would also catch a neighbouring
    figure's caption/title text in stacked (2x2 grid) layouts, contaminating
    this figure's conditions with unrelated text."""
    x0, x1, y0, y1 = box
    try:
        ws = [(w[0], w[1], w[2], w[3], w[4]) for w in page.get_text("words")
              if (x0 - 4 <= (w[0] + w[2]) / 2 <= x1 + 22
                  and y0 <= (w[1] + w[3]) / 2 <= y1 + 2)
              or (x0 - 4 <= (w[0] + w[2]) / 2 <= x1 + 4
                  and y0 - 16 <= (w[1] + w[3]) / 2 < y0)]
    except Exception:
        return []
    ws.sort(key=lambda w: (round(w[1], 1), w[0]))
    phrases, i = [], 0
    while i < len(ws):
        x0p, y0p, x1p, y1p, t = ws[i]
        j = i + 1
        while (j < len(ws) and abs(ws[j][1] - y0p) < 2.5
               and 0 <= ws[j][0] - x1p <= 6):
            t += (" " if ws[j][0] - x1p > 1 else "") + ws[j][4]
            x1p = ws[j][2]; y1p = max(y1p, ws[j][3]); j += 1
        phrases.append({"text": collapse(t), "cx": (x0p + x1p) / 2,
                        "cy": (y0p + y1p) / 2, "bb": (x0p, y0p, x1p, y1p)})
        i = j
    return phrases

def _harvest_inplot_text(page, box, caption, curves_TV):
    """
    Sort the text drawn inside the plot into:
      • per-curve labels  (short symbols sitting on/next to a curve)
      • figure-wide conditions (anything with '=', units, 'Reference', …)
      • x / y axis titles
    curves_TV: list of (px_x_array, px_y_array) in PDF coordinates.
    """
    x0, x1, y0, y1 = box
    cond_parts, labels = [], {}
    x_title = y_title = None
    cap_l = caption.lower()
    for ph in _inplot_phrases(page, box):
        t = ph["text"]
        if not t or re.fullmatch(r"-?\d+(\.\d+)?", t): continue
        if re.fullmatch(r"[A-Z]{2,5}\d{2,5}[A-Z]?(\s*-\s*\d+)?", t): continue   # figure drawing codes
        if re.fullmatch(r"[A-Za-z]{2,4}\d{3,6}[vV]\d+", t): continue            # ST watermark codes (AM10399v1)
        if re.search(r"handbook|halfpage|full\s*pagewidth|dbook", t, re.I): continue  # layout markers
        if re.fullmatch(r"(axis\s*title|1st\s*line|2nd\s*line|series\s*\d*)", t, re.I): continue  # Excel-chart template text
        if t.lower() in cap_l and len(t) > 10:  continue      # caption echo
        # axis titles
        if re.search(r"\(\s*°?\s*C\s*\)", t) and ph["cy"] > y1 - 0.2 * (y1 - y0) and len(t) <= 12:
            x_title = t; continue
        if (("/" in t and "=" not in t)
                or re.search(r"\((W|A|V|pF|nF|ns|nC|µA|mA)\)\s*$", t)) \
                and ph["cy"] < y0 + 0.25 * (y1 - y0):
            y_title = t; continue
        hard_cond = (len(t) > 28
                     or re.search(r"reference|waveform|single\s+pulse|measured|"
                                  r"typical\s+values|prior\s+to", t, re.I))
        if not hard_cond:
            # try to attach to the nearest curve at this x — equation-style
            # labels ("Tj = 125 °C", "RthJA = 130 °C/W") must sit CLOSER to
            # the curve than plain symbols to count as that curve's label
            best_ci, best_d = None, (12.0 if "=" in t else 18.0)
            for ci, (TX, TY) in enumerate(curves_TV):
                if not (TX[0] - 22 <= ph["cx"] <= TX[-1] + 25): continue
                xq = min(max(ph["cx"], TX[0]), TX[-1])
                ycurve = float(np.interp(xq, TX, TY))
                d = abs(ycurve - ph["cy"])
                if d < best_d: best_d, best_ci = d, ci
            if best_ci is not None:
                labels.setdefault(best_ci, []).append(t)
                continue
            if "=" not in t and len(t) > 14:
                pass        # long plain text → condition below
        cond_parts.append(t)
    # second pass: unlabelled curves take the nearest remaining short
    # '='-phrase within a wider radius (e.g. "RthJA = 130 \u00b0C/W")
    leftovers = [c for c in cond_parts if "=" in c and len(c) <= 22]
    for ci, (TX, TY) in enumerate(curves_TV):
        if ci in labels or not leftovers: continue
        best, bd = None, 30.0
        for ph2 in _inplot_phrases(page, box):
            if ph2["text"] not in leftovers: continue
            if not (TX[0] - 10 <= ph2["cx"] <= TX[-1] + 10): continue
            d = abs(float(np.interp(ph2["cx"], TX, TY)) - ph2["cy"])
            if d < bd: bd, best = d, ph2["text"]
        if best:
            labels[ci] = [best]
            leftovers.remove(best)
            cond_parts = [c for c in cond_parts if c != best]
    # exactly ONE curve still unlabelled and exactly ONE leftover phrase whose
    # symbol resembles an assigned label (RthJM↔RthJA, Tj=…↔Tj=…) → pair them
    unlab = [ci for ci in range(len(curves_TV)) if ci not in labels]
    if len(unlab) == 1 and len(leftovers) == 1 and labels:
        lhs_new = leftovers[0].split("=")[0].strip()[:3].lower()
        if any(v[0].split("=")[0].strip()[:3].lower() == lhs_new
               for v in labels.values()):
            labels[unlab[0]] = [leftovers[0]]
            cond_parts = [c for c in cond_parts if c != leftovers[0]]
    # dedupe conditions, keep order
    seen, conds = set(), []
    for c in cond_parts:
        if c not in seen:
            seen.add(c); conds.append(c)
    return labels, "; ".join(conds)[:180], x_title, y_title

def _render_figure_png(page, box, cap_bb, all_caps=None, zoom=2.5):
    """Crop to include the plot box, its axis labels, AND the figure caption
    (title line with figure number).

    The crop includes:
      - The plot box with generous axis-label margins
      - The caption/title line (Figure N. Title...) for context
      - Any in-plot condition text or parameter values
    This gives the reader immediate context: which figure it is, what it shows,
    and what conditions apply.

    To stay clean in tight 2x2 grids, the crop is clamped vertically so it
    never reaches into another figure's plot area.
    """
    x0, x1, y0, y1 = box
    # Determine whether caption is above or below the plot
    cap_above = cap_bb[1] < y0  # caption top is above plot top
    if cap_above:
        # Caption is above the plot — include it in the top margin
        ry0 = min(cap_bb[1] - 4, y0 - 16)   # start from caption top
        ry1 = y1 + 30                         # X-axis tick labels + title
    else:
        # Caption is below the plot — include it in the bottom margin
        ry0 = y0 - 16                         # top axis title / watermark
        ry1 = max(cap_bb[3] + 4, y1 + 30)    # extend to caption bottom
    rx0 = x0 - 50      # Y-axis tick labels + (rotated) Y-axis title
    rx1 = x1 + 16      # small right margin
    # Clamp against other captions so we don't leak into adjacent figures
    if all_caps:
        for _c in all_caps:
            cb = _c[2]
            cx0, cy0, cx1, cy1 = cb[0], cb[1], cb[2], cb[3]
            if min(cx1, x1) - max(cx0, x0) < 0.2 * (x1 - x0):
                continue                       # different column
            cyc = (cy0 + cy1) / 2.0
            # Only clamp against OTHER figure captions, not our own
            own_cap = (abs(cb[0] - cap_bb[0]) < 5 and abs(cb[1] - cap_bb[1]) < 5)
            if own_cap:
                continue
            if cyc <= y0 and cy1 < cap_bb[1]:  # another caption above us
                ry0 = max(ry0, cy1 + 2)
            elif cyc >= y1 and cy0 > cap_bb[3]:  # another caption below us
                ry1 = min(ry1, cy0 - 2)
    rx0 = max(rx0, 0); ry0 = max(ry0, 0)
    rx1 = min(rx1, page.rect.x1); ry1 = min(ry1, page.rect.y1)
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom),
                              clip=fitz.Rect(rx0, ry0, rx1, ry1))
        return pix.tobytes("png"), pix.width, pix.height
    except Exception:
        return None, 0, 0

def _ocr_axis_from_image(pil_img, x0, x1, y0, y1, axis, zoom=4):
    """OCR numeric tick labels just outside the plot box of a raster figure.
    Returns a robust linear fit (m, c) mapping pixel→value, or None."""
    if not _HAS_OCR:
        return None
    import numpy as _np
    W, H = pil_img.size
    if axis == "x":
        crop = pil_img.crop((max(0, x0 - 6), min(H, y1 + 1),
                             min(W, x1 + 6), min(H, y1 + 26)))
    else:
        crop = pil_img.crop((max(0, x0 - 46), max(0, y0 - 6),
                             max(0, x0 - 1), min(H, y1 + 6)))
    cw, ch = crop.size
    if cw < 4 or ch < 4:
        return None
    crop = crop.resize((cw * zoom, ch * zoom))
    try:
        data = pytesseract.image_to_data(
            crop, config="--psm 11 -c tessedit_char_whitelist=0123456789.-",
            output_type=pytesseract.Output.DICT)
    except Exception:
        return None
    pts = []
    for i, txt in enumerate(data["text"]):
        t = (txt or "").strip()
        if not re.fullmatch(r"-?\d+(\.\d+)?", t):
            continue
        try:
            val = float(t)
        except ValueError:
            continue
        if axis == "x":
            cx = (data["left"][i] + data["width"][i] / 2) / zoom + max(0, x0 - 6)
            pts.append((val, cx))
        else:
            cy = (data["top"][i] + data["height"][i] / 2) / zoom + max(0, y0 - 6)
            pts.append((val, cy))
    if len(pts) < 2:
        return None
    lo, hi = (x0, x1) if axis == "x" else (y1, y0)
    return _robust_axis_fit(pts, lo, hi)


def extract_raster_temp_graphs(pdf_path, taken_keys=None):
    """
    Fallback for datasheets whose temperature graphs are EMBEDDED RASTER
    IMAGES (e.g. Infineon CoolSiC).  For each image whose nearby sub-caption /
    figure caption matches the temperature pattern, the plot box is found with
    _axbox, curves are tracked column-by-column, and the axes are OCR-calibrated.
    Returns the same figure-dict list shape as extract_derating_curves().
    """
    figs = []
    if not (_HAS_GRAPH and os.path.exists(pdf_path)):
        return figs
    taken_keys = taken_keys or set()
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return figs

    for pi in range(doc.page_count):
        page = doc[pi]
        # gather caption-like lines (Figure N + sub-captions like "Ptot = f(TC)")
        caps = _figure_captions(page)
        # only proceed if at least one temperature caption exists on this page
        if not any(_is_graph_target(c[1]) for c in caps):
            continue
        try:
            infos = page.get_image_info(xrefs=True)
        except Exception:
            infos = page.get_image_info()
        for inf in infos:
            bbox = inf.get("bbox")
            if not bbox:
                continue
            ix0, iy0, ix1, iy1 = bbox
            iw, ih = ix1 - ix0, iy1 - iy0
            if iw < 90 or ih < 90:
                continue                       # logos / icons
            # caption for this image: nearest temperature caption sitting just
            # BELOW this image (Infineon prints "Ptot = f(TC)" under the plot).
            # Only "caption below image" is accepted so a caption between two
            # stacked figures is not also claimed by the figure beneath it.
            best_cap, best_d = None, 1e9
            for (fn, ctext, cbb) in caps:
                if not _is_graph_target(ctext):
                    continue
                cxc = (cbb[0] + cbb[2]) / 2
                if not (ix0 - 20 <= cxc <= ix1 + 20):
                    continue
                if cbb[1] < iy1 - 6:        # caption must be below the image
                    continue
                d = cbb[1] - iy1
                if 0 <= d < 60 and d < best_d:
                    best_d, best_cap = d, (fn, ctext, cbb)
            if not best_cap:
                continue
            fig_no, caption, cbb = best_cap
            # attach a "Figure N" number from the closest Figure caption in column
            if fig_no is None:
                for (fn2, c2, b2) in caps:
                    if fn2 is None:
                        continue
                    if abs((b2[0] + b2[2]) / 2 - (ix0 + ix1) / 2) < 0.5 * iw \
                            and abs(b2[1] - cbb[1]) < 40:
                        fig_no = fn2; break

            key = (pi, round(ix0), round(iy0))
            if key in taken_keys:
                continue

            # render the image region (plus a margin so the axis tick labels,
            # which Infineon places OUTSIDE the image bbox, are captured too)
            zoom = 3.0
            mxn = 0.16 * iw
            myn = 0.13 * ih
            pw_pg, ph_pg = page.rect.width, page.rect.height
            clip = fitz.Rect(max(0, ix0 - mxn), max(0, iy0 - myn),
                             min(pw_pg, ix1 + mxn), min(ph_pg, iy1 + myn))
            try:
                pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip)
                pil = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
            except Exception:
                continue
            gray = cv2.cvtColor(np.array(pil), cv2.COLOR_RGB2GRAY)
            W, H = pil.size
            # OCR the entire rendered figure once; numeric tokens are the axis
            # ticks.  Their geometry both calibrates the axes AND locates the
            # plot box (more robust than border detection, which can latch onto
            # the figure's table frame).
            if not _HAS_OCR:
                continue
            try:
                data = pytesseract.image_to_data(
                    pil, config="--psm 11 -c tessedit_char_whitelist=0123456789.-",
                    output_type=pytesseract.Output.DICT)
            except Exception:
                continue
            xticks, yticks = [], []   # (value, pixel_center)
            for i, txt in enumerate(data["text"]):
                t = (txt or "").strip()
                if not re.fullmatch(r"-?\d{1,4}(\.\d+)?", t):
                    continue
                try: val = float(t)
                except ValueError: continue
                cx = data["left"][i] + data["width"][i] / 2
                cy = data["top"][i] + data["height"][i] / 2
                xticks.append((val, cx, cy))
                yticks.append((val, cy, cx))
            if len(xticks) < 3:
                continue

            def _axis_band(ticks, perp_tol, want="x"):
                """Pick the tick band that spreads most along the axis
                direction (x ticks spread horizontally, y ticks vertically),
                then keep the longest strictly-increasing value run so legend
                numbers / stray reads from adjacent figures are dropped."""
                best, best_spread = None, -1.0
                for _, _pos, perp in ticks:
                    near = [t for t in ticks if abs(t[2] - perp) < perp_tol]
                    if len(near) < 3:
                        continue
                    spread = max(t[1] for t in near) - min(t[1] for t in near)
                    if spread > best_spread:
                        best_spread, best = spread, near
                if not best:
                    return []
                uniq = sorted({round(t[1]): t for t in best}.values(),
                              key=lambda t: t[1])
                pts = [(t[0], t[1]) for t in uniq]
                if len(pts) >= 4:
                    vals = [pp[0] for pp in pts]
                    inc = sum(1 for a, b in zip(vals, vals[1:]) if b > a)
                    dec = sum(1 for a, b in zip(vals, vals[1:]) if b < a)
                    if dec > inc:
                        pts = pts[::-1]
                    best_run, cur = [pts[0]], [pts[0]]
                    for pp in pts[1:]:
                        if pp[0] > cur[-1][0]:
                            cur.append(pp)
                        else:
                            if len(cur) > len(best_run): best_run = cur
                            cur = [pp]
                    if len(cur) > len(best_run): best_run = cur
                    if len(best_run) >= 3:
                        pts = best_run
                return pts

            xpts = _axis_band(xticks, 0.025 * H, "x")
            ypts = _axis_band(yticks, 0.04 * W, "y")
            if len(xpts) < 3 or len(ypts) < 3:
                continue
            xfit = _robust_axis_fit(xpts, min(p[1] for p in xpts), max(p[1] for p in xpts))
            yfit = _robust_axis_fit(ypts, min(p[1] for p in ypts), max(p[1] for p in ypts))
            if not xfit or not yfit:
                continue
            mxp, cxp = xfit[0], xfit[1]
            myp, cyp = yfit[0], yfit[1]
            # plot box in pixels.  Tick extremes give only the span actually
            # OCR-read; when a few ticks are missed (e.g. Infineon Ptot=f(TC),
            # where only 25/50/75 are recognised) that box collapses to a corner
            # of the graph.  So prefer the TRUE data-plot rectangle detected from
            # the axis border lines, snapped around the ticks we did read; only
            # fall back to the tick-extreme box if that detection fails.
            #
            # Use only ticks that are INLIERS of the (RANSAC) calibration line:
            # a stray OCR read below/above the axis (e.g. a misread "-55" under
            # the x-axis) would otherwise drag the box past the real frame and
            # pull in below-axis label rows as phantom flat curves.
            def _xspan(vmin, vmax):
                return max(1.0, abs(vmax - vmin))
            xv_span = _xspan(min(p[0] for p in xpts), max(p[0] for p in xpts))
            yv_span = _xspan(min(p[0] for p in ypts), max(p[0] for p in ypts))
            xtol = max(2.0, 0.06 * xv_span)
            ytol = max(2.0, 0.06 * yv_span)
            xpx_in = [p[1] for p in xpts if abs(p[0] - (mxp * p[1] + cxp)) <= xtol]
            ypx_in = [p[1] for p in ypts if abs(p[0] - (myp * p[1] + cyp)) <= ytol]
            if len(xpx_in) < 2: xpx_in = [p[1] for p in xpts]
            if len(ypx_in) < 2: ypx_in = [p[1] for p in ypts]
            xl = int(min(xpx_in)); xr = int(max(xpx_in))
            yt = int(min(ypx_in)); yb = int(max(ypx_in))
            _db = _data_plot_box(gray, xpx_in, ypx_in)
            if _db:
                xl, xr, yt, yb = _db
            if (xr - xl) < 40 or (yb - yt) < 40:
                continue

            # curves inside the plot box (exclude a few px near the frame)
            plot = gray[max(0, yt + 1):min(H, yb - 1), max(0, xl + 1):min(W, xr - 1)]
            ph, pw = plot.shape
            # Blank out full-width horizontal lines (the top/bottom frame and any
            # solid-dark gridlines): they span the whole width and would appear as
            # a flat phantom "curve" in every column.  Genuine derating/power
            # curves are monotonic and never run dark across most of the width, so
            # a high dark-fraction row is structural, not data.
            if pw > 0:
                row_dark = (plot < _DARK).sum(axis=1) / float(pw)
                for ry in np.where(row_dark > 0.55)[0]:
                    plot[ry, :] = 255
            col_pts = {}
            for cc in range(pw):
                ink = np.where(plot[:, cc] < _DARK)[0].tolist()
                cl = _clusters(ink, 0)
                if cl:
                    col_pts[cc] = [r for r, _w in cl]
            if len(col_pts) < pw * 0.3:
                continue

            # number of curves AND their left-edge seed positions, found by
            # POOLING cluster rows over the left strip and grouping them into
            # y-levels.  Pooling is robust to dotted curves (whose dots are sparse
            # in any single column but accumulate into a clear level) and to close
            # neighbours (kept apart as long as the level gap stays small), which a
            # per-column modal count cannot handle.
            xs = sorted(col_pts.keys())
            if not xs:
                continue

            def _cluster1d(vals, gap):
                vals = sorted(vals)
                out, cur = [], [vals[0]]
                for v in vals[1:]:
                    if v - cur[-1] <= gap:
                        cur.append(v)
                    else:
                        out.append((float(np.mean(cur)), len(cur))); cur = [v]
                out.append((float(np.mean(cur)), len(cur)))
                return out

            n_strip = max(5, int(len(xs) * 0.06))
            strip = xs[:n_strip]
            pooled = []
            for c in strip:
                pooled.extend(col_pts[c])
            lvl_gap = max(8, int(0.025 * (yb - yt)))
            levels = _cluster1d(pooled, lvl_gap) if pooled else []
            lvl_support = max(3, int(0.22 * n_strip))
            seeds = sorted([row for row, cnt in levels if cnt >= lvl_support])
            if not seeds:
                seeds = sorted([row for row, _ in levels])[:1] or [sorted(col_pts[xs[0]])[0]]
            n_curves = max(1, min(len(seeds), 8))
            seed_rows = seeds[:n_curves]

            # follow each curve left→right from its seed level.  Curves never
            # cross, so each track takes the nearest row within a gap tolerance;
            # the tolerance is kept below the smallest inter-curve spacing so a
            # solid curve is never captured by a momentarily-absent neighbour
            # (e.g. a gap in the dotted curve).
            tracks = [dict() for _ in range(n_curves)]
            if n_curves >= 2:
                min_sep = min(b - a for a, b in zip(seed_rows, seed_rows[1:]))
            else:
                min_sep = (yb - yt)
            gap_tol = max(10.0, min(0.06 * (yb - yt), 0.55 * min_sep))
            prev = list(seed_rows)
            for c in xs:
                rows_here = sorted(col_pts.get(c, []))
                if not rows_here:
                    continue
                for ti2 in range(n_curves):
                    r = min(rows_here, key=lambda rr: abs(rr - prev[ti2]))
                    if abs(r - prev[ti2]) <= gap_tol:
                        tracks[ti2][c] = r
                        prev[ti2] = r
            kind = _classify_fig(caption)
            # fold in the "Figure N <title>" caption (Infineon prints the
            # descriptive title there: "Power dissipation" / "Max. forward
            # current") which classifies more reliably than "Ptot = f(TC)"
            fig_title = ""
            for (fn3, c3, b3) in caps:
                if fn3 == fig_no and fn3 is not None and not _TEMP_X_PAT.search(c3):
                    fig_title = re.sub(r"^figure\s*\d+\s*", "", c3, flags=re.I)
                    break
            merged_cap = (caption + " " + fig_title).strip()
            if kind == "generic":
                low = merged_cap.lower()
                if re.search(r"\bp\s*tot|dissipation", low):     kind = "power"
                elif re.search(r"\bi\s*f\b|forward\s+current", low): kind = "current"
            tmin_v, tmax_v = mxp * xl + cxp, mxp * xr + cxp
            tlo, thi = min(tmin_v, tmax_v), max(tmin_v, tmax_v)
            stp = 5.0 if (thi - tlo) > 60 else max(round((thi - tlo) / 20, 1), 0.5)
            grid = np.round(np.arange(np.ceil(tlo / stp) * stp,
                                      thi + stp * 0.25, stp), 2)

            curves_out = []
            for ti2, tr in enumerate(tracks):
                if len(tr) < pw * 0.12:
                    continue
                cols = sorted(tr.keys())
                Tv = np.array([mxp * (c + xl + 1) + cxp for c in cols])
                Yv = np.array([myp * (tr[c] + yt + 1) + cyp for c in cols])
                order = np.argsort(Tv)
                Tv, Yv = Tv[order], Yv[order]
                vals = []
                for t in grid:
                    if Tv[0] - 1e-6 <= t <= Tv[-1] + 1e-6:
                        vals.append(round(float(np.interp(t, Tv, Yv)), 3))
                    else:
                        vals.append(None)
                label = (f"Curve {ti2 + 1}" if n_curves > 1 else "Curve")
                curves_out.append({"label": label, "style": "solid", "vals": vals})
            if not curves_out:
                continue
            # Y-range from calibration: clamp out-of-range artefacts (legend box
            # outlines, stray ink) to None so the table never shows impossible
            # values; the embedded figure image remains the reference.
            yvals_all = [mxp * 0 for _ in (0,)]  # placeholder to keep names
            y_at_top = myp * yt + cyp
            y_at_bot = myp * yb + cyp
            ylo_cal, yhi_cal = min(y_at_top, y_at_bot), max(y_at_top, y_at_bot)
            span = yhi_cal - ylo_cal
            ylo_cal -= 0.05 * span; yhi_cal += 0.05 * span
            for cv in curves_out:
                cv["vals"] = [v if (v is None or ylo_cal <= v <= yhi_cal) else None
                              for v in cv["vals"]]
            # drop curves that ended up essentially empty after clamping
            curves_out = [cv for cv in curves_out
                          if sum(1 for v in cv["vals"] if v is not None) >= 3]
            # NOTE: a "75%-of-points-agree" dedup pass was tried here and
            # removed — duty-cycle / capacitance curve families that
            # legitimately converge over part of their range were being
            # collapsed into one curve.  The seed-level clustering above
            # (lvl_gap / lvl_support) already separates genuinely distinct
            # curves at their left edge, which is the reliable signal.
            if not curves_out:
                continue

            # harvest in-image legend labels (D = 0.1 etc.) via OCR of plot area
            conditions = collapse(re.sub(r"^figure\s*\d+\s*", "", caption, flags=re.I))
            y_unit = None
            try:
                hdr = page.get_text("words")
                for w in hdr:
                    if (ix0 - 40) <= w[0] <= (ix0 + 30) and (iy0 - 6) <= w[1] <= (iy0 + 0.5 * ih):
                        mm = re.fullmatch(r"\(?(W|A|V|pF|nF|mA|\u00b5A|K/W)\)?", w[4].strip())
                        if mm: y_unit = mm.group(1); break
            except Exception:
                pass

            png, pw2, ph2 = None, iw, ih
            try:
                png = pix.tobytes("png"); pw2, ph2 = pix.width, pix.height
            except Exception:
                pass

            _xunit = _fig_x_unit(merged_cap, None, None)
            if _xunit == "\u00b0C":
                x_name = _deriv_x_name(merged_cap, None)
            else:
                x_name = "RG / ID (X axis as printed)"
            y_name = (f"Ptot ({y_unit or 'W'})" if kind == "power" else
                      f"IF ({y_unit or 'A'})" if kind == "current" else
                      f"E ({y_unit or 'µJ'})" if kind == "energy" else
                      _deriv_y_name(kind, merged_cap, None))
            disp_cap = (fig_title or caption).strip()
            figs.append({
                "fig_no": fig_no, "kind": kind, "page": pi + 1, "x_unit": _xunit,
                "caption": disp_cap[:170], "x_name": x_name, "y_name": y_name,
                "calib": "X/Y: OCR from raster image (approximate \u2014 verify against embedded figure)",
                "conditions": conditions,
                "temps": [round(float(t), 2) for t in grid],
                "curves": curves_out, "png": png, "png_w": pw2, "png_h": ph2,
                "_raster": True})
            taken_keys.add(key)
    doc.close()
    return figs


def extract_derating_curves(pdf_path, interactive=True):
    """
    Digitise every figure whose X axis is a temperature: current derating,
    power dissipation vs T, surge-current ratio vs Tj, relative dynamic
    parameters vs Tj, etc.

    Returns a list of figure dicts:
      {fig_no, kind, page, caption, x_name, y_name, calib, conditions,
       temps:[...], curves:[{label, style, vals:[...]}], png, png_w, png_h}
    """
    figs = []
    if not _HAS_GRAPH:
        return figs
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return figs
    for pi in range(doc.page_count):
        page = doc[pi]
        captions = _figure_captions(page)
        targets = [(ci, c) for ci, c in enumerate(captions)
                   if _is_graph_target(c[1])]
        if not targets:
            continue
        boxes = _find_plot_boxes(page)
        if not boxes:
            continue
        pairs = dict(_pair_caption_box(captions, boxes))
        for ci, (fig_no, caption, cbb) in [(ci, c) for ci, c in enumerate(captions)]:
            if ci not in pairs:
                continue
            box = boxes[pairs[ci]]
            ax_title = _x_axis_title(page, box)
            # qualify by caption text OR by the x-axis title printed under
            # the graph (Tj / Tc / Tamb / TM / "... Temperature (\u00b0C)")
            if not (_is_graph_target(caption)
                    or (ax_title and _XTITLE_TEMP.search(ax_title))):
                continue
            x0b, x1b, y0b, y1b = box
            segs = _collect_curve_segments(page, box)
            tracked = _track_curves(segs, box)
            if not tracked: continue
            pct_y = False
            try:
                for w in page.get_text("words"):
                    if "%" in w[4] and (x0b - 30) <= w[0] <= (x0b + 20) \
                            and (y0b - 25) <= w[1] <= (y0b + 0.4 * (y1b - y0b)):
                        pct_y = True; break
            except Exception:
                pass
            (mx, cxx), xmeth, xlog = _calibrate_axis(page, box, "x", interactive, caption)
            (my, cyy), ymeth, ylog = _calibrate_axis(page, box, "y", interactive, caption)
            calib = f"X: {xmeth} | Y: {ymeth}"
            y_unit = _y_unit_hint(page, box)
            kind = _classify_fig(caption)
            # PDF-coordinate polylines for label matching
            curves_px = []
            for c, coverage in tracked:
                ks = sorted(c)
                curves_px.append((np.array(ks), np.array([c[k] for k in ks]), coverage))
            curve_lbls, conditions, x_title, y_title = _harvest_inplot_text(
                page, box, caption, [(t[0], t[1]) for t in curves_px])
            # convert to engineering units
            conv = []
            for idx, (KX, KY, coverage) in enumerate(curves_px):
                T = mx * KX + cxx
                V = my * KY + cyy
                if xlog:
                    T = 10.0 ** T
                if ylog:
                    V = 10.0 ** V
                conv.append((T, V, coverage, curve_lbls.get(idx)))
            # ── post-tracking curve deduplication (engineering units) ─────
            # After conversion, a few "curves" can still be near-duplicates
            # caused by the vector tracker following both sides of a very
            # thick line, or a closely-spaced gridline that survived the
            # pixel-domain filters.  This is now a SAFETY NET (the main fix
            # is the slope-aware column tracker above, which already keeps
            # crossing curves correctly separated) — so the tolerance here is
            # intentionally tight to avoid merging genuinely distinct curves
            # that happen to be close over part of their range (e.g. two
            # curves crossing, or converging asymptotically).
            #
            # For log-scale Y axes (capacitance, Zth) comparisons are done in
            # LOG space: a fixed pF/K·W difference is huge at the low end and
            # tiny at the high end, so a linear-scale tolerance either merges
            # real low-value curves or never catches real duplicates at the
            # high end.  Comparing log10(V) puts both ends on equal footing.
            if len(conv) > 1:
                def _curves_similar(c1, c2, ytol_frac=0.02):
                    """Check if two curves are near-duplicates (tight tolerance,
                    safety net only — see note above)."""
                    T1, V1, _, _ = c1
                    T2, V2, _, _ = c2
                    tlo = max(T1[0], T2[0])
                    thi = min(T1[-1], T2[-1])
                    if thi <= tlo:
                        return False
                    tpts = np.linspace(tlo, thi, 20)
                    v1s = np.interp(tpts, T1, V1)
                    v2s = np.interp(tpts, T2, V2)
                    if ylog:
                        v1s = np.log10(np.clip(v1s, 1e-12, None))
                        v2s = np.log10(np.clip(v2s, 1e-12, None))
                        allv = np.concatenate([np.log10(np.clip(V1, 1e-12, None)),
                                               np.log10(np.clip(V2, 1e-12, None))])
                    else:
                        allv = np.concatenate([V1, V2])
                    yrange = max(float(np.max(allv) - np.min(allv)), 1e-9)
                    diffs = np.abs(v1s - v2s) / yrange
                    # require near-PERFECT agreement across the ENTIRE shared
                    # range (95%+) — two curves that merely cross or converge
                    # briefly will fail this; only true duplicates (parallel
                    # traces of one physical line) pass.
                    return np.mean(diffs < ytol_frac) > 0.95

                keep = [True] * len(conv)
                for i_c in range(len(conv)):
                    if not keep[i_c]:
                        continue
                    for j_c in range(i_c + 1, len(conv)):
                        if not keep[j_c]:
                            continue
                        if _curves_similar(conv[i_c], conv[j_c]):
                            # keep whichever has higher coverage / more points
                            if conv[j_c][2] > conv[i_c][2]:
                                keep[i_c] = False
                            else:
                                keep[j_c] = False
                conv = [c for c, k in zip(conv, keep) if k]

            if not conv:
                continue
            tmin = min(t[0][0] for t in conv); tmax = max(t[0][-1] for t in conv)
            if xlog and tmin > 0 and tmax > tmin:
                # log-spaced sample grid (e.g. thermal impedance tp[s] decades)
                grid = np.round(np.logspace(np.log10(tmin), np.log10(tmax), 25), 10)
            elif "NORMALISED" in xmeth:
                grid = np.round(np.linspace(tmin, tmax, 21), 3)
            else:
                stp = 5.0 if (tmax - tmin) > 60 else max(round((tmax - tmin) / 20, 1), 0.5)
                grid = np.arange(np.ceil(tmin / stp) * stp, tmax + stp * 0.25, stp)
                grid = np.round(grid, 2)
            conv.sort(key=lambda t: -t[1][0])
            # ── SAFE label-assisted cleanup ──────────────────────────────
            # Earlier dedup (above, before unit conversion) already merges
            # near-identical curves.  This second check only removes a curve
            # when it is provably a near-duplicate of a LABELLED curve (e.g.
            # the tracker produced two parallel traces of the same physical
            # line — one picked up the label, one didn't).  It must NEVER
            # delete a curve just because it lacks a text label: real curve
            # families (Zth duty-cycle sets, multi-temperature sweeps) are
            # often mostly unlabelled and that is expected, not an error.
            labelled_idx = [i for i, (_, _, _, lbl) in enumerate(conv) if lbl]
            if labelled_idx and len(conv) > len(labelled_idx):
                def _shape_dup(c1, c2, ytol_frac=0.04):
                    T1, V1 = c1[0], c1[1]; T2, V2 = c2[0], c2[1]
                    tlo, thi = max(T1[0], T2[0]), min(T1[-1], T2[-1])
                    if thi <= tlo:
                        return False
                    tpts = np.linspace(tlo, thi, 15)
                    v1s = np.interp(tpts, T1, V1); v2s = np.interp(tpts, T2, V2)
                    if ylog:
                        v1s = np.log10(np.clip(v1s, 1e-12, None))
                        v2s = np.log10(np.clip(v2s, 1e-12, None))
                        yr = max(float(np.max(np.log10(np.clip(V1, 1e-12, None)))
                                       - np.min(np.log10(np.clip(V1, 1e-12, None)))),
                                 float(np.max(np.log10(np.clip(V2, 1e-12, None)))
                                       - np.min(np.log10(np.clip(V2, 1e-12, None)))), 1e-9)
                    else:
                        yr = max(max(V1) - min(V1), max(V2) - min(V2), 1e-9)
                    return np.mean(np.abs(v1s - v2s) / yr < ytol_frac) > 0.80
                keep2 = [True] * len(conv)
                for li in labelled_idx:
                    for ui in range(len(conv)):
                        if ui == li or not keep2[ui] or conv[ui][3]:
                            continue
                        if _shape_dup(conv[li], conv[ui]):
                            keep2[ui] = False
                conv = [c for c, k in zip(conv, keep2) if k]
            # Hard cap: if still >8 curves, keep only the ones with best coverage
            if len(conv) > 8:
                conv.sort(key=lambda t: -t[2])  # sort by coverage descending
                conv = conv[:8]
                conv.sort(key=lambda t: -t[1][0])  # re-sort by Y position

            # ── Capacitance graph: enforce Ciss > Coss > Crss label ordering ──
            # Spatial-proximity label matching can assign the wrong name when
            # curves cross each other (Coss and Crss converge at high VDS).
            # The physics is unambiguous: Ciss ≥ Coss ≥ Crss at every VDS, so
            # re-sort by median Y in log space and reassign names by position.
            if kind == "capacitance" and 2 <= len(conv) <= 3:
                conv.sort(key=lambda t: -float(
                    np.median(np.log10(np.clip(t[1], 1e-12, None)))))
                cap_names = ["Ciss", "Coss", "Crss"][:len(conv)]
                # Always reassign by position: Ciss is biggest, Crss smallest
                conv = [(t[0], t[1], t[2], [cap_names[ci_cap]])
                        for ci_cap, t in enumerate(conv)]

            curves_out = []
            for ci2, (T, V, coverage, lbl) in enumerate(conv, 1):
                vals = []
                for t in grid:
                    if T[0] - 1e-6 <= t <= T[-1] + 1e-6:
                        v = float(np.interp(t, T, V))
                        vals.append(float(f"{v:.4g}") if ylog else round(v, 3))
                    else:
                        vals.append(None)
                style = "dashed/dotted" if coverage < 0.85 else "solid"
                if lbl:
                    label = " ".join(lbl)[:40]
                elif kind == "power" and len(conv) == 2:
                    label = "Ptot typ (solid)" if style == "solid" else "Ptot max (dashed)"
                elif len(conv) == 1:
                    label = "Curve"
                else:
                    label = (f"Curve {ci2} (top)" if ci2 == 1 else
                             f"Curve {ci2} (bottom)" if ci2 == len(conv) else f"Curve {ci2}")
                curves_out.append({"label": label, "style": style, "vals": vals})
            png, pw, ph = _render_figure_png(page, box, cbb, all_caps=captions)
            figs.append({
                "fig_no": fig_no, "kind": kind, "page": pi + 1,
                "caption": caption[:170], "_vector": True,
                "x_name": _deriv_x_name(caption + " " + (ax_title or ""), x_title or ax_title),
                "y_name": ("% of rated value" if pct_y else
                           (f"IR ({y_unit})" if "reverse current" in caption.lower() and y_unit and not y_title else
                            f"IF ({y_unit})" if kind == "current" and y_unit and not y_title else
                            f"Ptot ({y_unit})" if kind == "power" and y_unit and not y_title else
                            f"E ({y_unit or 'µJ'})" if kind == "energy" and not y_title else
                            f"ZthJC ({y_unit or 'K/W'})" if kind == "thermal_z" and not y_title else
                            f"C ({y_unit or 'pF'})" if kind == "capacitance" and not y_title else
                            f"RDS(on) ({y_unit})" if kind == "resistance" and y_unit and not y_title else
                            _deriv_y_name(kind, caption, y_title))),
                "x_unit": _fig_x_unit(caption, ax_title, x_title),
                "calib": calib, "conditions": conditions,
                "temps": [(float(f"{t:.4g}") if xlog else round(float(t), 2)) for t in grid],
                "curves": curves_out,
                "png": png, "png_w": pw, "png_h": ph})
    doc.close()
    # ── raster fallback: temperature graphs that are embedded images ────────
    try:
        have = {(f["page"], f.get("fig_no")) for f in figs}
        for rf in extract_raster_temp_graphs(pdf_path):
            if (rf["page"], rf.get("fig_no")) not in have:
                figs.append(rf)
    except Exception as _e:
        pass
    # stable order: by page, then figure number
    figs.sort(key=lambda f: (f["page"], f.get("fig_no") or 999))
    return figs

# ── rd from the VF-IF forward-characteristics graph (VECTOR, log-IF aware) ──
_FWD_CAP = re.compile(
    r"forward\s+characteristic|durchlasskennlinien"
    r"|forward\s+(voltage\s+drop|current).{0,45}(versus|vs\.?|as\s+a\s+function\s+of)"
    r".{0,45}(forward\s+)?(current|voltage)"
    r"|I\s*F?\s*=\s*f\s*\(\s*V", re.I)

def _decade_axis_labels(page, clip, axis="y"):
    """Tick labels on a log axis: '10'+superscript exponent pairs, plain
    numbers.  Returns [(log10(value), center_coord)] for robust linear fit.
    For axis='x', center_coord is x-center; for 'y', it is y-center."""
    ws = [w for w in page.get_text("words")
          if clip.x0 <= (w[0] + w[2]) / 2 <= clip.x1
          and clip.y0 <= (w[1] + w[3]) / 2 <= clip.y1]
    out, used = [], set()
    def _pos(w):
        return (w[0] + w[2]) / 2 if axis == "x" else (w[1] + w[3]) / 2
    for i, w in enumerate(ws):
        if i in used: continue
        t = w[4].strip()
        if t == "10":
            exp = 1
            for j, w2 in enumerate(ws):
                if j == i or j in used: continue
                # Superscript sits to the right of "10" and at/above its baseline
                if re.fullmatch(r"[+-]?\d+", w2[4].strip()) \
                        and -2 <= w2[0] - w[2] <= 10 and w2[1] <= w[1] + 5:
                    exp = int(w2[4]); used.add(j); break
            out.append((float(exp), _pos(w))); used.add(i)
        elif re.fullmatch(r"10[+-]?\d+", t):              # merged "10-1"/"102"/"10-6"
            m = re.match(r"10([+-]?\d+)", t)
            out.append((float(m.group(1)), _pos(w))); used.add(i)
        elif re.fullmatch(r"1|1\.0", t):
            out.append((0.0, _pos(w))); used.add(i)
        elif re.fullmatch(r"0\.1", t):
            out.append((-1.0, _pos(w))); used.add(i)
        elif re.fullmatch(r"0\.01", t):
            out.append((-2.0, _pos(w))); used.add(i)
        elif re.fullmatch(r"0\.001", t):
            out.append((-3.0, _pos(w))); used.add(i)
        elif re.fullmatch(r"0\.0001", t):
            out.append((-4.0, _pos(w))); used.add(i)
        elif re.fullmatch(r"100|1000|10000|100000", t):
            out.append((math.log10(float(t)), _pos(w))); used.add(i)
    return out

def _ocr_decade_axis_labels(page, clip, axis="x", zoom=8):
    """OCR-based decade label reader for log axes on RASTER figure images.
    Detects '10' tokens followed by a nearby superscript exponent digit, plus
    plain power-of-10 values (0.001, 0.01, 0.1, 1, 10, 100, ...).
    Returns [(log10_value, center_coord)] identical to _decade_axis_labels."""
    if not (_HAS_OCR and _HAS_GRAPH):
        return []
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), clip=clip)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        data = pytesseract.image_to_data(
            img, config="--psm 11 -c tessedit_char_whitelist=0123456789.-",
            output_type=pytesseract.Output.DICT)
    except Exception:
        return []
    tokens = []
    for i, txt in enumerate(data["text"]):
        t = (txt or "").strip()
        if not re.fullmatch(r"-?\d+(\.\d+)?", t):
            continue
        try:
            val = float(t)
        except ValueError:
            continue
        cx = clip.x0 + (data["left"][i] + data["width"][i] / 2) / zoom
        cy = clip.y0 + (data["top"][i] + data["height"][i] / 2) / zoom
        pos = cx if axis == "x" else cy
        lx = data["left"][i]; ly = data["top"][i]
        rw = data["width"][i]; rh = data["height"][i]
        tokens.append({"val": val, "pos": pos, "lx": lx, "ly": ly,
                       "rw": rw, "rh": rh})
    out, used = [], set()
    for i, tok in enumerate(tokens):
        if i in used:
            continue
        v = tok["val"]
        if abs(v - 10.0) < 0.5:
            # Look for an adjacent exponent token (superscript to the right)
            exp = 1
            for j, tok2 in enumerate(tokens):
                if j == i or j in used:
                    continue
                # x-gap in image pixels between right edge of "10" and left of exp
                gap_x = tok2["lx"] - (tok["lx"] + tok["rw"])
                # superscript sits at or above the baseline of "10"
                at_or_above = tok2["ly"] <= tok["ly"] + 0.4 * tok["rh"]
                if -2 <= gap_x <= 12 * zoom and at_or_above:
                    try:
                        e = int(tok2["val"])
                        # treat negative exponent: if the value token is negative
                        if tok2["val"] < 0:
                            e = int(tok2["val"])
                        exp = e; used.add(j); break
                    except Exception:
                        pass
            out.append((float(exp), tok["pos"])); used.add(i)
        else:
            # Plain power-of-10 value (0.1, 0.01, 100, 1000, …)
            if v > 0:
                lv = math.log10(v)
                if abs(lv - round(lv)) < 0.05:
                    out.append((round(lv), tok["pos"])); used.add(i)
    return out

def _rd_from_vector_vfif(pdf_path, if_rated):
    """
    Digitise the forward-characteristics (VF on X, IF on Y — linear or LOG)
    curve from the vector layer and compute rd as the secant slope between
    IF = IF_rated/2 and IF = IF_rated on the 25 °C curve (the curve with the
    highest VF at rated current — Schottky/ultrafast VF tempco is negative
    in that region).
    """
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return None
    best = None
    for pi in range(doc.page_count):
        page = doc[pi]
        caps = [(fn, c, bb) for (fn, c, bb) in _figure_captions(page)
                if _FWD_CAP.search(c)]
        # also accept caption-less: any line matching the pattern
        if not caps:
            try:
                d = page.get_text("dict")
                for b in d.get("blocks", []):
                    if b.get("type") != 0: continue
                    for l in b.get("lines", []):
                        t = " ".join(s.get("text", "") for s in l.get("spans", []))
                        if _FWD_CAP.search(t):
                            caps.append((None, collapse(t), list(l["bbox"])))
            except Exception:
                pass
        if not caps: continue
        boxes = _find_plot_boxes(page)
        if not boxes: continue
        pairs = dict(_pair_caption_box(caps, boxes))
        for ci, (fn, cap, cbb) in enumerate(caps):
            if ci not in pairs: continue
            box = boxes[pairs[ci]]
            x0, x1, y0, y1 = box
            segs = _collect_curve_segments(page, box)
            tracked = _track_curves(segs, box)
            if not tracked: continue
            xfit = _robust_axis_fit(
                _text_axis_labels(page, fitz.Rect(x0 - 22, y1 + 1, x1 + 26, y1 + 16), "x"),
                x0, x1)
            if not xfit:
                xfit = _robust_axis_fit(
                    _ocr_axis_labels(page, fitz.Rect(x0 - 22, y1 + 1, x1 + 26, y1 + 16),
                                     axis="x"), x0, x1)
            if not xfit: continue
            mx, cxx = xfit[0], xfit[1]
            ystrip = fitz.Rect(x0 - 36, y0 - 8, x0 - 1, y1 + 8)
            # log axis first (decade labels), then linear
            ylog = _robust_axis_fit(_decade_axis_labels(page, ystrip), y1, y0)
            ylin = None if ylog else _robust_axis_fit(
                _text_axis_labels(page, ystrip, "y"), y1, y0)
            if not ylog and not ylin: continue
            cands = []
            for c, coverage in tracked:
                ks = sorted(c)
                VF = np.array([mx * k + cxx for k in ks])
                ypix = np.array([c[k] for k in ks])
                if ylog:
                    IF = 10.0 ** (ylog[0] * ypix + ylog[1])
                else:
                    IF = ylin[0] * ypix + ylin[1]
                _iu = _y_unit_hint(page, box)
                IF = IF * {"mA": 1e-3, "µA": 1e-6, "uA": 1e-6,
                           "kA": 1e3}.get(_iu, 1.0)
                # keep curves that span the anchor current range
                i1, i2 = if_rated / 2.0, if_rated
                if IF.min() <= i1 * 0.9 and IF.max() >= i2 * 0.98:
                    order = np.argsort(IF)
                    IFs, VFs = IF[order], VF[order]
                    vf1 = float(np.interp(i1, IFs, VFs))
                    vf2 = float(np.interp(i2, IFs, VFs))
                    cands.append((vf2, vf1, coverage, IFs, VFs))
            if not cands: continue
            cands.sort(key=lambda t: -t[0])
            # Curve selection:
            #   · dotted/dashed curves are the OTHER temperature → restrict to
            #     solid curves when any exist (Diotec: 25°C solid, 125°C dotted)
            #   · 3+ solid curves: usually (high-T typ, 25°C typ, 25°C MAX) →
            #     second-highest VF is the 25°C TYPICAL curve (NXP/Nexperia)
            #   · otherwise highest VF = 25°C curve (negative VF tempco)
            solid = [c for c in cands if c[2] >= 0.85]
            pool = solid or cands
            # de-duplicate traces of the SAME physical curve (dotted curves
            # can track as two parallel traces with identical VF)
            dedup = []
            for c in pool:
                if not any(abs(c[0] - u[0]) < max(0.012, 0.02 * u[0]) for u in dedup):
                    dedup.append(c)
            pool = dedup
            pick = 1 if len(pool) >= 3 else 0
            vf2, vf1, coverage, IFs, VFs = pool[pick]
            i1, i2 = if_rated / 2.0, if_rated
            rd = (vf2 - vf1) / (i2 - i1)
            if not (0.0005 <= rd <= 10.0): continue
            sel = (IFs >= i1) & (IFs <= i2)
            r2 = 1.0
            if sel.sum() >= 4:
                p = np.polyfit(IFs[sel], VFs[sel], 1)
                pred = np.polyval(p, IFs[sel])
                ss_r = float(np.sum((VFs[sel] - pred) ** 2))
                ss_t = float(np.sum((VFs[sel] - VFs[sel].mean()) ** 2))
                r2 = 1 - ss_r / ss_t if ss_t > 0 else 1.0
            best = {"rd": rd, "r2": r2,
                    "vf0": vf2 - rd * i2,
                    "cond": (f"from VF-IF graph (vector), 25\u00b0C "
                             f"{'typical ' if pick == 1 else ''}curve, "
                             f"IF = {i1:g}\u2192{i2:g} A"),
                    "method": "vector VF-IF digitisation"}
            break
        if best: break
    doc.close()
    return best

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 1C — WORD-BASELINE TEXT SCANNER  (universal narrow-format fallback)
# ═════════════════════════════════════════════════════════════════════════════
# Some manufacturers (Diotec, some Vishay/onsemi appnote-style sheets) use
# narrow bilingual tables that pdfplumber cannot reconstruct, and draw symbol
# subscripts as separate text spans ("I" + "FAV"), which scrambles the plain
# text layer.  This scanner rebuilds true visual lines from the PDF *word*
# layer (grouped by baseline, sorted by x) — "IFAV = 10 A   VRRM = 100 V"
# comes back intact — and then parses  SYMBOL/NAME → (<|Typ.|~) VALUE UNIT
# with footnote markers ("5)") stripped and metric prefixes normalised.

_UNIT_ACCEPT = {   # param → {unit_text: factor_to_param_unit}
    "V_DS":    {"V": 1, "kV": 1000},
    "I_D":     {"A": 1, "mA": 0.001},
    "R_DS(on)":{"mΩ": 1, "Ω": 1000, "mohm": 1, "ohm": 1000, "mOhm": 1},
    "V_GS(th)":{"V": 1, "mV": 0.001},
    "C_iss":   {"pF": 1, "nF": 1000},
    "C_oss":   {"pF": 1, "nF": 1000},
    "C_oss(er)":  {"pF": 1, "nF": 1000},
    "C_oss(eff)": {"pF": 1, "nF": 1000},
    "C_rss":   {"pF": 1, "nF": 1000},
    "Q_g":     {"nC": 1, "µC": 1000, "uC": 1000},
    "Q_gs":    {"nC": 1, "µC": 1000, "uC": 1000},
    "Q_gd":    {"nC": 1, "µC": 1000, "uC": 1000},
    "t_d(on)": {"ns": 1, "µs": 1000, "us": 1000},
    "t_r":     {"ns": 1, "µs": 1000, "us": 1000},
    "t_d(off)":{"ns": 1, "µs": 1000, "us": 1000},
    "t_f":     {"ns": 1, "µs": 1000, "us": 1000},
    "E_on":    {"µJ": 1, "uJ": 1, "mJ": 1000, "nJ": 0.001},
    "E_off":   {"µJ": 1, "uJ": 1, "mJ": 1000, "nJ": 0.001},
    "E_oss":   {"µJ": 1, "uJ": 1, "mJ": 1000, "nJ": 0.001},
    "V_SD":    {"V": 1, "mV": 0.001},
    "t_rr":    {"ns": 1, "µs": 1000, "us": 1000},
    "Q_rr":    {"nC": 1, "µC": 1000, "uC": 1000},
    "R_thJC":  {"K/W": 1, "°C/W": 1, "C/W": 1},
    "R_thJA":  {"K/W": 1, "°C/W": 1, "C/W": 1},
    "R_thCS":  {"K/W": 1, "°C/W": 1, "C/W": 1},
}

_UNIT_TOKEN = re.compile(
    r"^(µA|uA|mA|nA|kA|A|kV|mV|V|K/W|°C/W|C/W|mΩ|Ω|mohm|ohm|mOhm|"
    r"pF|nF|µF|ns|µs|us|ms|ps|nC|µC|uC|mC|pC|mJ|µJ|uJ|nJ|W|%)\)?$")
_NUM_TOKEN = re.compile(r"^[<>≤≥~]?\s*([+-]?\d+(?:[.,]\d+)?)(?:/([+-]?\d+(?:[.,]\d+)?))?$")
_FOOT = re.compile(r"^\d\)$")

def baseline_lines(pdf_path):
    """Rebuild visual text lines from the fitz word layer for every page.
    Returns list-of-pages, each a list of dicts {text, words:[(x0,x1,t)]}."""
    pages = []
    if not _HAS_GRAPH:
        return pages
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return pages
    for page in doc:
        try:
            ws = sorted(page.get_text("words"), key=lambda w: (w[3], w[0]))
        except Exception:
            pages.append([]); continue
        lines, cur, cur_y = [], [], None
        for w in ws:
            yb = w[3]
            if cur_y is None or abs(yb - cur_y) <= 2.5:
                cur.append(w); cur_y = yb if cur_y is None else (cur_y + yb) / 2
            else:
                lines.append(cur); cur, cur_y = [w], yb
        if cur: lines.append(cur)
        out = []
        for ln in lines:
            ln = sorted(ln, key=lambda w: w[0])
            out.append({"text": " ".join(w[4] for w in ln),
                        "words": [(w[0], w[2], w[4]) for w in ln]})
        pages.append(out)
    doc.close()
    return pages

def _line_value_candidates(words):
    """[(x, value, unit, kind)] — kind in {'typ','max','min'}; footnotes stripped."""
    out = []
    toks = [t for t in words if not _FOOT.match(t[2])]
    for i, (x0, x1, t) in enumerate(toks):
        kind = "typ"
        tt = t
        if tt[:1] in "<≤":  kind, tt = "max", tt[1:].strip()
        elif tt[:1] in ">≥": kind, tt = "min", tt[1:].strip()
        elif tt[:1] == "~":  kind, tt = "typ", tt[1:].strip()
        m = _NUM_TOKEN.match(tt)
        if not m: continue
        # marker in the PREVIOUS token ("Typ.", "<", "max.")
        if kind == "typ" and i > 0:
            prev = toks[i - 1][2].rstrip(".").lower()
            if prev in ("<", "≤"): kind = "max"
            elif prev in (">", "≥"): kind = "min"
            elif prev in ("max", "max."): kind = "max"
        # unit: rest of this token after the number, or the NEXT token
        unit = None
        tail = tt[m.end():].strip()
        um = _UNIT_TOKEN.match(tail) if tail else None
        if um: unit = um.group(1)
        elif i + 1 < len(toks):
            um = _UNIT_TOKEN.match(toks[i + 1][2])
            if um: unit = um.group(1)
        if not unit: continue
        vals = [m.group(1)]
        if m.group(2): vals.append(m.group(2))
        for v in vals:
            out.append((x0, float(v.replace(",", ".")), unit, kind))
    return out

def scan_baseline_lines(pages_lines, p):
    """Universal line scanner over rebuilt visual lines."""
    accept = _UNIT_ACCEPT.get(p["param"])
    if not accept: return []
    records, multi = [], p.get("multi_cond", False)
    for lines in pages_lines:
        for ln in lines:
            lc = collapse(ln["text"])
            if len(lc) > 220: continue
            sym_x, at_cond = None, ""
            for (x0, x1, t) in ln["words"]:
                base, _, at = t.partition("@")
                if len(base) <= 10 and smatch(compact(base), p["symbol_res"]):
                    sym_x, at_cond = x0, (at or "")
                    break
            name_hit = nmatch(lc, p["name_res"])
            if sym_x is None and not name_hit: continue
            cands = [(x, v * accept[u], u, k)
                     for (x, v, u, k) in _line_value_candidates(ln["words"])
                     if u in accept]
            cands = [c for c in cands if in_range(p["param"], str(c[1]))]
            if not cands: continue
            if sym_x is not None:
                after = [c for c in cands if c[0] > sym_x]
                cands = after or cands
            at_t = get_temps(at_cond)
            temps = at_t if at_t else get_temps(lc)
            tc = max(temps) if temps else None
            ih = bool(tc and tc >= 90)
            cond = clean_cond(lc)
            if at_cond: cond = clean_cond(at_cond) + "; " + cond
            # group: one record per condition-line; first typ + first max
            tv = next((c for c in cands if c[3] == "typ"), None)
            mv = next((c for c in cands if c[3] == "max"), None)
            if p.get("use_min") and not tv:
                tv = next((c for c in cands if c[3] == "min"), None)
            if not tv and not mv: continue
            extra = []
            if tv and len([c for c in cands if c[3] == "typ"]) > 1:
                extra = [c for c in cands if c[3] == "typ"][1:]
            records.append({"typ": fmt_val(str(tv[1])) if tv else "—",
                            "max": fmt_val(str(mv[1])) if mv else "—",
                            "cond": cond[:130], "temp_c": tc, "is_high": ih,
                            "extracted_unit": None})
            for (x, v, u, k) in extra:        # e.g. "IFSM = 180/200 A"
                records.append({"typ": fmt_val(str(v)), "max": "—",
                                "cond": cond[:130], "temp_c": tc, "is_high": ih,
                                "extracted_unit": None})
            if not multi:
                return records
    return records

def scan_diotec_vf(pages_lines, p):
    """
    Diotec characteristics layout:
        VF [V] 1)  @ IF [A]  @ Tj      VF [V] 1)  @ IF [A]  @ Tj
        Typ. 0.51   25°C   < 0.68  25°C
              5                10
        Typ. 0.45  125°C   < 0.62 125°C
    Typ values share the left IF current, '<' (max) values the right one.
    """
    for lines in pages_lines:
        for li, ln in enumerate(lines):
            if not re.search(r"V\s*F?\s*\[V\].*@\s*I\s*F?\s*\[A\]", ln["text"]):
                continue
            typs, maxs, ifs = [], [], []
            for ln2 in lines[li + 1: li + 6]:
                words = [t for t in ln2["words"] if not _FOOT.match(t[2])]
                texts = [w[2] for w in words]
                temps_x = [(w[0], int(m.group(1))) for w in words
                           if (m := re.match(r"^(\d{2,3})°C$", w[2]))]
                for i, t in enumerate(texts):
                    if t.rstrip(".").lower() == "typ" and i + 1 < len(texts):
                        m = re.match(r"^(\d+(?:\.\d+)?)$", texts[i + 1])
                        if m: typs.append((words[i + 1][0], float(m.group(1)), temps_x))
                    if t in ("<", "≤") and i + 1 < len(texts):
                        m = re.match(r"^(\d+(?:\.\d+)?)$", texts[i + 1])
                        if m: maxs.append((words[i + 1][0], float(m.group(1)), temps_x))
                    m2 = re.match(r"^<(\d+(?:\.\d+)?)$", t)
                    if m2: maxs.append((words[i][0], float(m2.group(1)), temps_x))
                if all(re.match(r"^\d+(\.\d+)?$", t) for t in texts) and 1 <= len(texts) <= 3:
                    ifs.extend(float(t) for t in texts)
            if not typs and not maxs: continue
            recs = []
            if_typ = f"IF = {ifs[0]:g} A" if ifs else ""
            if_max = f"IF = {ifs[1]:g} A" if len(ifs) > 1 else if_typ
            def _t_for(x, temps_x):
                if not temps_x: return None
                return min(temps_x, key=lambda t: abs(t[0] - x))[1]
            for (x, v, tx) in typs:
                if not in_range("VF", str(v)): continue
                tc = _t_for(x, tx)
                recs.append({"typ": fmt_val(str(v)), "max": "—",
                             "cond": if_typ or "—", "temp_c": tc,
                             "is_high": bool(tc and tc >= 90), "extracted_unit": None})
            for (x, v, tx) in maxs:
                if not in_range("VF", str(v)): continue
                tc = _t_for(x, tx)
                recs.append({"typ": "—", "max": fmt_val(str(v)),
                             "cond": if_max or "—", "temp_c": tc,
                             "is_high": bool(tc and tc >= 90), "extracted_unit": None})
            if recs: return recs
    return []

# ── Multi-device family tables (Vishay etc.) ────────────────────────────────
_DEV_TOKEN = re.compile(r"^[A-Z]{1,5}\d[A-Z0-9./\-]{1,12}$")
_DEV_STOP = {"UNIT", "SYMBOL", "TYP", "TYP.", "MAX", "MAX.", "MIN", "MIN.",
             "PARAMETER", "VALUE", "CONDITIONS", "TEST", "NOTES"}

def find_device_columns(all_tables):
    """
    Detect family-datasheet tables whose header row carries one column per
    device (Vishay: PARAMETER | SYMBOL | AS1FD | AS1FG | … | UNIT).
    Returns (devices, variants) where variants maps a compacted symbol to
    [(device, value_string), …] for rows whose per-device values DIFFER.
    """
    devices, variants = [], {}
    for tb in all_tables:
        rows = tb["data"]
        if not rows: continue
        # the device header can sit below a banner row ("MAXIMUM RATINGS…")
        hdr_i, hdr, dev_cols = None, None, []
        for ri in range(min(3, len(rows))):
            cand = [collapse(c) for c in (rows[ri] or [])]
            if len(cand) < 4: continue
            dc = [(i, h) for i, h in enumerate(cand)
                  if h and _DEV_TOKEN.match(h.upper()) and h.upper() not in _DEV_STOP]
            if len(dc) >= 2:
                hdr_i, hdr, dev_cols = ri, cand, dc
                break
        if hdr_i is None: continue
        if not devices:
            devices = [h for _, h in dev_cols]
        sym_col = next((i for i, h in enumerate(hdr) if "SYMBOL" in h.upper()), None)
        for rr in rows[hdr_i + 1:]:
            if not rr: continue
            sym = compact(rr[sym_col]) if sym_col is not None and sym_col < len(rr) else ""
            if not sym:
                # symbol may live in any short cell
                for c in rr:
                    cc = compact(c)
                    if cc and len(cc) <= 8 and re.match(r"^[A-Za-zθ]", cc):
                        sym = cc; break
            vals = []
            for i, dname in dev_cols:
                v = collapse(rr[i]) if i < len(rr) else ""
                vals.append((dname, v))
            filled = [v for _, v in vals if v and not is_blank(v)]
            distinct = {v for v in filled}
            if sym and len(filled) >= 2 and len(distinct) >= 2:
                # spanning cells (same value for all devices) stay single-row
                variants.setdefault(sym, vals)
    return devices, variants

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 2 — PARAMETER CATALOGUE
# ═════════════════════════════════════════════════════════════════════════════



def _rd_from_vfif_graph(pdf_path, if_rated, vf_rated, debug_png=None):
    """
    Extract rd from a VF-IF forward characteristics graph in the PDF.
    Universal — works for Infineon, STM, WeEN, and all other manufacturers.
    Handles any orientation/nomenclature: IF vs VF, VF vs IF, forward char, etc.
    """
    if not _HAS_GRAPH or not os.path.exists(pdf_path):
        return None
    try:
        doc = fitz.open(pdf_path)
    except Exception:
        return None

    # Page caption patterns that indicate a VF-IF graph page
    VFIF_PAT = re.compile(
        r"forward.{0,20}current.{0,30}forward.{0,20}volt"
        r"|forward.{0,20}volt.{0,30}forward.{0,20}curr"
        r"|I[_\s]*F\s*[=,f]\s*f\s*\(\s*V[_\s]*F"
        r"|forward.{0,15}char"
        r"|V[_\s]*F.{0,10}FORWARD.{0,10}VOLT"
        r"|I[_\s]*F.{0,10}FORWARD.{0,10}CURR",
        re.I)

    best_result = None
    best_score  = -999.0

    for pi in range(doc.page_count):
        page = doc[pi]
        txt  = "".join(ch for ch in (page.get_text("text") or "") if ord(ch) >= 32)
        if not VFIF_PAT.search(txt):
            continue

        # Auto-detect axis ranges from page text
        vf_max_auto = 2.5
        for m in re.finditer(r"\b([2-5]\.0|[2-5])\b", txt):
            try:
                v = float(m.group(1))
                if 1.5 <= v <= 5.0:
                    vf_max_auto = v; break
            except: pass
        if_max_auto = round(if_rated * 2.0 / 10) * 10
        for n in re.findall(r"\b(40|30|60|100|50|80|20)\b", txt):
            try:
                v = float(n)
                if if_rated * 1.5 <= v <= if_rated * 5:
                    if_max_auto = v; break
            except: pass

        # Rasterise page at 4x
        mat = fitz.Matrix(4.0, 4.0)
        pix = page.get_pixmap(matrix=mat)
        try:
            full = Image.open(io.BytesIO(pix.tobytes("png"))).convert("RGB")
        except Exception:
            continue
        pw, ph = full.size

        # Try left half and right half (most datasheets have 2-column graph layout)
        # Also try full width and other splits
        quadrants = [
            (0,     ph//2, pw//2, ph),    # bottom-left
            (0,     0,     pw//2, ph//2), # top-left
            (pw//2, ph//2, pw,    ph),    # bottom-right
            (pw//2, 0,     pw,    ph//2), # top-right
            (0,     ph//4, pw//2, 3*ph//4), # left-centre
        ]

        for (x0, y0, x1, y1) in quadrants:
            tile = full.crop((x0, y0, x1, y1))
            gray = cv2.cvtColor(np.array(tile), cv2.COLOR_RGB2GRAY)
            axbox = _axbox(gray)
            if not axbox: continue
            xl, xr, yt, yb = axbox
            if (xr - xl) < 150 or (yb - yt) < 150: continue

            # Build column data
            ix0 = xl+3; iy0 = yt+3
            cd = _coldata(gray[iy0:yb-3, ix0:xr-3], ix0, iy0)
            if len(cd) < 40: continue

            for vfm in [vf_max_auto, 2.5, 3.0, 3.5, 2.0, 4.0]:
                ifm = if_max_auto

                px2vf = lambda c: _p2v(c, xl, xr, 0.0, vfm)
                px2if = lambda r: _p2v(r, yb, yt, 0.0, ifm)

                # Anchor: find nearest dark pixel to (vf_rated, if_rated)
                ac = _v2p(vf_rated, 0.0, vfm, xl, xr)
                ar = _v2p(if_rated, 0.0, ifm, yb, yt)
                bd = 999; sc, sr = ac, ar
                for ca in range(max(ix0, ac-60), min(xr-3, ac+60)):
                    for ra, sz in cd.get(ca, []):
                        d = abs(ra - ar)
                        if d < bd: bd=d; sc,sr=ca,ra
                if bd > 100: continue

                IF_MIN = if_rated * 0.5
                IF_MAX = if_rated * 1.0
                fwd = _track(sc, sr, cd, px2if, IF_MIN, IF_MAX, True)
                bwd = _track(sc, sr, cd, px2if, IF_MIN, IF_MAX, False)
                curve = list(reversed(bwd)) + fwd[1:]

                vf_pts = np.array([px2vf(c) for c,r in curve])
                if_pts = np.array([px2if(r) for c,r in curve])
                mask = ((if_pts >= IF_MIN) & (if_pts <= IF_MAX)
                        & (vf_pts > 0.3) & (vf_pts < vfm * 0.95))
                vfl = vf_pts[mask]; ifl = if_pts[mask]

                if len(ifl) < 8 or len(np.unique(ifl)) < 4: continue

                sl, ic, rv, _, _ = linregress(ifl, vfl)
                rd = sl; vf0 = ic; r2 = rv**2

                if not np.isfinite(rd) or rd <= 0 or rd > 1.0: continue
                if r2 < 0.85: continue

                ap = 2.0 * abs(if_rated / ifm - 0.5)
                score = r2 - ap
                if score > best_score:
                    best_score = score
                    best_result = {
                        "rd":    round(rd, 5),
                        "vf0":   round(vf0, 4),
                        "r2":    round(r2, 5),
                        "n_pts": len(ifl),
                        "method": f"VF-IF graph digitisation (p{pi+1})",
                        "cond":  (f"rd from VF-IF graph 25 degC curve; "
                                  f"IF={IF_MIN:.0f}-{IF_MAX:.0f}A; "
                                  f"anchor {vf_rated:.2f}V@{if_rated:.0f}A; "
                                  f"R2={r2:.4f} n={len(ifl)}"),
                    }
                    if debug_png:
                        try:
                            from PIL import ImageDraw
                            dbg = tile.copy(); draw = ImageDraw.Draw(dbg)
                            for c,r in curve:
                                if 0<=c<dbg.width and 0<=r<dbg.height:
                                    draw.ellipse([c-2,r-2,c+2,r+2], fill="lime")
                            vff = [vf0+rd*i for i in [IF_MIN, IF_MAX]]
                            draw.line([_v2p(vff[0],0.0,vfm,xl,xr), _v2p(IF_MIN,0.0,ifm,yb,yt),
                                       _v2p(vff[1],0.0,vfm,xl,xr), _v2p(IF_MAX,0.0,ifm,yb,yt)],
                                      fill="red", width=3)
                            dbg.save(debug_png)
                        except: pass

    doc.close()
    return best_result

PARAMS=[
    # ── Required Parameters ────────────────────────────────────────────────
    dict(param="V_DS",full="Drain-Source Breakdown Voltage",unit="V",
         section="Required Parameters",
         note="Device voltage rating",
         symbol_res=[r"^V\(?BR\)?DSS$",r"^V\(?BR\)?\(?DSS\)?$",r"^VBRDSS$",
                     r"^VDSS$",r"^V\(?DSS\)?$",r"^VDS$",r"^V\(?DS\)?$",
                     r"^BVDSS$",r"^V\(?BR\)?$"],
         name_res=[r"drain.?source\s+breakdown\s+voltage",
                   r"breakdown\s+voltage",
                   r"drain.?source\s+voltage",
                   r"drain.?to.?source\s+(breakdown\s+)?voltage"],
         use_min=True,multi_cond=False),
    dict(param="I_D",full="Continuous Drain Current",unit="A",
         section="Required Parameters",
         note="Sets conduction-loss operating point",
         symbol_res=[r"^ID$",r"^I\(?D\)?$",r"^ID25$",r"^IDcont$",
                     r"^ID\(?cont\)?$",r"^IDcontinuous$",r"^IDDC$",r"^I\(?DDC\)?$",
                     r"^IDC$",r"^I\(?DC\)?$"],
         name_res=[r"continuous\s+(dc\s+)?drain\s+current",
                   r"drain\s+current.*(continuous|dc)",
                   r"^drain\s+current$",r"continuous\s+current"],
         use_min=False,multi_cond=True),
    dict(param="R_DS(on)",full="Drain-Source On-Resistance",unit="m\u03a9",
         section="Required Parameters",
         note="Conduction loss: P = I\u00b2\u00b7RDS(on)",
         symbol_res=[r"^RDS\(?on\)?$",r"^R\(?DS\)?\(?on\)?$",r"^RDSon$",
                     r"^RDS$",r"^Ron$",r"^R\(?on\)?$"],
         name_res=[r"drain.?source\s+on.?(state\s+)?resistance",
                   r"static\s+drain.?source\s+on.?resistance",
                   r"on.?(state\s+)?resistance"],
         use_min=False,multi_cond=True),
    dict(param="V_GS(th)",full="Gate Threshold Voltage",unit="V",
         section="Required Parameters",
         note="Turn-on threshold; sets gate-drive requirement",
         symbol_res=[r"^VGS\(?th\)?$",r"^V\(?GS\)?\(?th\)?$",r"^VGSth$",
                     r"^Vth$",r"^V\(?th\)?$",r"^VTH$",r"^VGS\(?TO\)?$"],
         name_res=[r"gate\s+threshold\s+voltage",
                   r"gate.?source\s+threshold\s+voltage",
                   r"threshold\s+voltage"],
         use_min=False,multi_cond=False),
    dict(param="C_iss",full="Input Capacitance",unit="pF",
         section="Required Parameters",
         note="Gate drive: Pgate = Qg\u00b7VGS\u00b7fsw",
         symbol_res=[r"^Ciss$",r"^C\(?iss\)?$",r"^CISS$"],
         name_res=[r"input\s+capacitance"],
         use_min=False,multi_cond=False),
    dict(param="C_oss",full="Output Capacitance",unit="pF",
         section="Required Parameters",
         note="Switching: Poss = \u00bd\u00b7Coss\u00b7VDS\u00b2\u00b7fsw",
         symbol_res=[r"^Coss$",r"^C\(?oss\)?$",r"^COSS$"],
         name_res=[r"(?<!related )(?<!effective )(?<!energy )(?<!time )"
                   r"output\s+capacitance"],
         use_min=False,multi_cond=False),
    # Effective output capacitance comes in two flavours that some makers list
    # alongside the plain Coss (ON Semi: Coss(eff.) + Coss(er.); Vishay: Co(tr)
    # + Co(er)).  They are NOT the plain Coss and must each get their OWN row,
    # never collapsed into the Coss row.  Symbol match is authoritative; the
    # energy-related one is claimed first so the bare "effective" name falls to
    # C_oss(eff).
    dict(param="C_oss(er)",full="Output Capacitance (energy related)",unit="pF",
         section="Required Parameters",
         note="Energy-equivalent Coss for switching-loss estimation",
         symbol_res=[r"^Coss\(?er\.?\)?$",r"^Co\(?er\.?\)?$",
                     r"^Cosser$",r"^Coer$"],
         name_res=[r"energy[\s\-]*related\s+output\s+capacitance",
                   r"output\s+capacitance[,\s]+energy[\s\-]*related",
                   r"effective\s+output\s+capacitance[,\s]+energy[\s\-]*related"],
         use_min=False,multi_cond=False),
    dict(param="C_oss(eff)",full="Effective Output Capacitance (time related)",unit="pF",
         section="Required Parameters",
         note="Charge/time-equivalent Coss (Co(tr) / Coss(eff.))",
         symbol_res=[r"^Coss\(?eff\.?\)?$",r"^Co\(?eff\.?\)?$",
                     r"^Coss\(?tr\.?\)?$",r"^Co\(?tr\.?\)?$",
                     r"^Cosseff$",r"^Cotr$"],
         name_res=[r"time[\s\-]*related\s+output\s+capacitance",
                   r"output\s+capacitance[,\s]+time[\s\-]*related",
                   r"effective\s+output\s+capacitance(?!.*energy)"],
         use_min=False,multi_cond=False),
    dict(param="C_rss",full="Reverse Transfer Capacitance",unit="pF",
         section="Required Parameters",
         note="Miller effect / switching speed",
         symbol_res=[r"^Crss$",r"^C\(?rss\)?$",r"^CRSS$"],
         name_res=[r"reverse\s+transfer\s+capacitance",
                   r"miller\s+capacitance"],
         use_min=False,multi_cond=False),
    dict(param="Q_g",full="Total Gate Charge",unit="nC",
         section="Required Parameters",
         note="Gate drive: Pgate = Qg\u00b7VGS\u00b7fsw",
         symbol_res=[r"^Qg$",r"^Q\(?g\)?$",r"^QG$",r"^Qgtot$",
                     r"^Qg\(?tot(al)?\)?$",r"^Qgtotal$"],
         name_res=[r"total\s+gate\s+charge",r"^gate\s+charge$",
                   r"gate\s+charge\s+total"],
         use_min=False,multi_cond=True),
    # ── Loss Analysis Parameters ───────────────────────────────────────────
    dict(param="Q_gs",full="Gate-Source Charge",unit="nC",
         section="Loss Analysis Parameters",
         note="Switching: pre-Miller gate charge",
         symbol_res=[r"^Qgs$",r"^Q\(?gs\)?$",r"^QGS$",r"^Qgs1$",r"^Qgs2$",
                     r"^QGS\(?pl\)?$",r"^Qgs\(?pl\)?$",r"^QGSpl$"],
         name_res=[r"gate[\s\-]?(to[\s\-]?)?source\s+charge",
                   r"plateau.*gate.*source"],
         use_min=False,multi_cond=False),
    dict(param="Q_gd",full="Gate-Drain (Miller) Charge",unit="nC",
         section="Loss Analysis Parameters",
         note="Switching loss: controls turn-on/off transition",
         symbol_res=[r"^Qgd$",r"^Q\(?gd\)?$",r"^QGD$"],
         name_res=[r"gate[\s\-]?(to[\s\-]?)?drain\s+charge",r"miller\s+charge",
                   r"gate.?drain\s+\(?miller\)?\s+charge"],
         use_min=False,multi_cond=False),
    dict(param="t_d(on)",full="Turn-On Delay Time",unit="ns",
         section="Loss Analysis Parameters",
         note="Switching: part of total turn-on time",
         symbol_res=[r"^td\(?on\)?$",r"^t\(?d\(?on\)?\)?$",r"^tdon$",
                     r"^td,on$"],
         name_res=[r"turn.?on\s+delay\s+time",r"turn.?on\s+delay"],
         use_min=False,multi_cond=False),
    dict(param="t_r",full="Rise Time",unit="ns",
         section="Loss Analysis Parameters",
         note="Switching: Psw \u221d (tr+tf)\u00b7VDS\u00b7ID\u00b7fsw",
         symbol_res=[r"^tr$",r"^t\(?r\)?$",r"^trise$",r"^t\(?rise\)?$"],
         name_res=[r"^rise\s+time$",r"\brise\s+time\b"],
         use_min=False,multi_cond=False),
    dict(param="t_d(off)",full="Turn-Off Delay Time",unit="ns",
         section="Loss Analysis Parameters",
         note="Switching: part of total turn-off time",
         symbol_res=[r"^td\(?off\)?$",r"^t\(?d\(?off\)?\)?$",r"^tdoff$",
                     r"^td,off$"],
         name_res=[r"turn.?off\s+delay\s+time",r"turn.?off\s+delay"],
         use_min=False,multi_cond=False),
    dict(param="t_f",full="Fall Time",unit="ns",
         section="Loss Analysis Parameters",
         note="Switching: Psw \u221d (tr+tf)\u00b7VDS\u00b7ID\u00b7fsw",
         symbol_res=[r"^tf$",r"^t\(?f\)?$",r"^tfall$",r"^t\(?fall\)?$"],
         name_res=[r"^fall\s+time$",r"\bfall\s+time\b"],
         use_min=False,multi_cond=False),
    dict(param="E_on",full="Turn-ON Switching Energy",unit="\u00b5J",
         section="Loss Analysis Parameters",
         note="Turn-on loss: Pon = Eon\u00b7fsw",
         symbol_res=[r"^Eon$",r"^E\(?on\)?$",r"^EON$"],
         name_res=[r"turn.?on\s+(switching\s+)?energy",
                   r"switch.?on\s+energy"],
         use_min=False,multi_cond=True),
    dict(param="E_off",full="Turn-OFF Switching Energy",unit="\u00b5J",
         section="Loss Analysis Parameters",
         note="Turn-off loss: Poff = Eoff\u00b7fsw",
         symbol_res=[r"^Eoff$",r"^E\(?off\)?$",r"^EOFF$"],
         name_res=[r"turn.?off\s+(switching\s+)?energy",
                   r"switch.?off\s+energy"],
         use_min=False,multi_cond=True),
    dict(param="E_oss",full="Output Capacitance Stored Energy",unit="\u00b5J",
         section="Loss Analysis Parameters",
         note="Hard-switching loss per cycle (\u00bd\u00b7Coss\u00b7VDS\u00b2)",
         symbol_res=[r"^Eoss$",r"^E\(?oss\)?$",r"^EOSS$",
                     r"^Eoss@?\d*V?$",r"^E@?\d{2,4}V?oss$"],
         name_res=[r"output\s+(capacitance\s+)?(stored\s+)?energy",
                   r"coss\s+stored\s+energy",
                   r"energy\s+stored\s+in\s+coss",
                   r"stored\s+energy.*coss"],
         use_min=False,multi_cond=False),
    dict(param="V_SD",full="Body Diode Forward Voltage",unit="V",
         section="Loss Analysis Parameters",
         note="Dead-time conduction loss in half-bridge",
         symbol_res=[r"^VSD$",r"^V\(?SD\)?$",r"^VFSD$",r"^VF$",r"^V\(?F\)?$",
                     r"^VDS\(?diode\)?$"],
         name_res=[r"(body\s+)?diode\s+forward\s+voltage",
                   r"source.?drain\s+(diode\s+)?voltage",
                   r"diode\s+forward\s+(on.?)?voltage",
                   r"forward\s+(on.?)?voltage.*diode"],
         use_min=False,multi_cond=True),
    dict(param="t_rr",full="Reverse Recovery Time",unit="ns",
         section="Loss Analysis Parameters",
         note="Body-diode recovery loss in hard-switching",
         symbol_res=[r"^trr$",r"^t\(?rr\)?$",r"^t_?rr$"],
         name_res=[r"reverse\s+recovery\s+time"],
         use_min=False,multi_cond=False),
    dict(param="Q_rr",full="Reverse Recovery Charge",unit="nC",
         section="Loss Analysis Parameters",
         note="Body-diode recovery loss: Prr = Qrr\u00b7VDS\u00b7fsw",
         symbol_res=[r"^Qrr$",r"^Q\(?rr\)?$",r"^QRR$",
                     r"^Qr$",r"^Q\(?r\)?$"],
         name_res=[r"reverse\s+recovery\s+charge",
                   r"recovered\s+charge"],
         use_min=False,multi_cond=False),
    # ── Thermal Parameters (any K/W or \u00b0C/W parameter) ─────────────────
    dict(param="R_thJC",full="Thermal Resistance Junction-to-Case",unit="K/W",
         section="Thermal Parameters",
         note="Junction temperature rise: \u0394Tj = Ptot \u00b7 RthJC",
         symbol_res=[r"^Rth\(?j[\-\u2013\u2011\u2212]?c\)?$",r"^RthJC$",
                     r"^R\(?th\)?\(?jc\)?$",r"^\u03b8\(?j[\-\u2013\u2011\u2212]?c\)?$",
                     r"^R\u03b8JC$",r"^Zth\(?j[\-\u2013\u2011\u2212]?c\)?$",r"^ZthJC$"],
         name_res=[r"thermal\s+res.*junction.{0,8}case",
                   r"junction.{0,6}(?:to.{0,4})?case\s+thermal",
                   r"junction\s+to\s+case"],
         use_min=False,multi_cond=False),
    dict(param="R_thJA",full="Thermal Resistance Junction-to-Ambient",unit="K/W",
         section="Thermal Parameters",
         note="Free-air operation: \u0394Tj = Ptot \u00b7 RthJA (no heatsink)",
         symbol_res=[r"^Rth\(?j[\-\u2013\u2011\u2212]?a(mb)?\)?$",r"^RthJA$",
                     r"^R\(?th\)?\(?ja\)?$",r"^\u03b8\(?j[\-\u2013\u2011\u2212]?a\)?$",
                     r"^R\u03b8JA$",r"^ZthJA$"],
         name_res=[r"thermal\s+res.*junction.{0,20}ambient",
                   r"junction\s+to\s+ambient"],
         use_min=False,multi_cond=False),
    dict(param="R_thCS",full="Thermal Resistance Case-to-Sink",unit="K/W",
         section="Thermal Parameters",
         note="Case-to-heatsink interface (with thermal paste/TIM)",
         symbol_res=[r"^Rth\(?c[\-\u2013\u2011\u2212]?s(ink)?\)?$",r"^RthCS$",
                     r"^R\(?th\)?\(?cs\)?$",r"^\u03b8\(?c[\-\u2013\u2011\u2212]?s\)?$",
                     r"^RthCH$",r"^Rth\(?c[\-\u2013\u2011\u2212]?h(s)?\)?$"],
         name_res=[r"thermal\s+res.*case.{0,20}(sink|heat.?sink)",
                   r"case\s+to\s+(sink|heat.?sink)"],
         use_min=False,multi_cond=False),
]


UNIT_RANGES={
    "V_DS":(20,2000),"I_D":(0.1,2000),"R_DS(on)":(0.1,200000),
    "V_GS(th)":(0.3,12),"C_iss":(1,5e6),"C_oss":(0.1,5e6),
    "C_oss(er)":(0.1,5e6),"C_oss(eff)":(0.1,5e6),"C_rss":(0.05,5e6),
    "Q_g":(0.1,5000),"Q_gs":(0.02,3000),"Q_gd":(0.02,3000),
    "t_d(on)":(0.05,1e5),"t_r":(0.05,1e5),"t_d(off)":(0.05,1e5),"t_f":(0.05,1e5),
    "E_on":(0.001,1e6),"E_off":(0.001,1e6),"E_oss":(0.001,1e6),
    "V_SD":(0.2,6.0),"t_rr":(0.5,1e5),"Q_rr":(0.05,1e6),
    "R_thJC":(0.01,500),"R_thJA":(0.05,5000),"R_thCS":(0.01,500),
}
SECTION_ORDER={
    "Required Parameters":0,"Loss Analysis Parameters":1,
    "Thermal Parameters":2,
}


# ═
# ═════════════════════════════════════════════════════════════════════════════
# SECTION 3 — UTILITIES + TABLE PARSING
# ═════════════════════════════════════════════════════════════════════════════

_GLYPH_FIX = {"(cid:2)": "\u00b5", "\uf06d": "\u00b5", "\uf0b0": "\u00b0",
              "(cid:176)": "\u00b0", "\uf0a3": "\u2264", "(cid:163)": "\u2264"}
def deglyph(s):
    s = str(s or "")
    for k, v in _GLYPH_FIX.items():
        if k in s: s = s.replace(k, v)
    return s
def compact(s): return re.sub(r"[\s_]","",deglyph(s))
def collapse(s): return re.sub(r"\s+"," ",deglyph(s)).strip()
def is_blank(v): return collapse(v) in ("","-","\u2011","\u2013","\u2014","n.a.","N/A","na",".","−","n.c.","NC","*")
def get_temps(s):
    # A temperature is "<n> °C" but NOT the °C inside a thermal-resistance unit
    # "°C/W" — otherwise "0.52 °C/W" yields a bogus temp of 52 and "40 °C/W"
    # yields 40.  The negative lookahead (?!\s*/) rejects the unit form.
    return [int(t) for t in re.findall(
        r"(\d{2,3})\s*[°\uf0b0]?\s*[Cc]\b(?!\s*/)", deglyph(s))]
def first_num(s):
    """
    Extract the first meaningful number from a cell string.
    Handles dual-device notation: '2 x 30 A' → returns '30' (per-unit value).
    '2 x 30' means 2 diodes × 30 A each; we want 30 (per diode).
    """
    s = str(s or "")
    # Dual-device pattern: "N x M" or "N X M" — return M (the per-unit value)
    m_dual = re.search(r'(\d+)\s*[xX×]\s*([0-9]+\.?[0-9]*)', s)
    if m_dual:
        return m_dual.group(2)
    # thousands separated by a space or thin space: "1 000" → "1000"
    s = re.sub(r"(?<=\d)[\s\u202f\u00a0](?=\d{3}\b)", "", s)
    m = re.search(r"[+-]?\d+\.?\d*(?:[eE][+-]?\d+)?", s)
    return m.group(0) if m else None
def fmt_val(v):
    if v is None: return "—"
    try:
        f=float(v); return str(int(f)) if f==int(f) else f"{f:g}"
    except (ValueError,OverflowError): return str(v)

def _safe_str(v):
    """Strip control chars that Excel/openpyxl cannot accept."""
    if v is None: return v
    return ''.join(ch for ch in str(v) if ord(ch) >= 32 or ch in '\t\n ')



def _decode_infineon_switching_cond(raw, page_text="", preceding_text=""):
    """
    Decode garbled Infineon switching condition strings (Symbol-font gibberish).

    Uses TWO text sources for temperature:
    1. preceding_text: text from the page BEFORE this table (most reliable)
    2. page_text: fallback if preceding_text not provided

    For dual-temp sections "175°C/125°C":
      diF/dt=1000 A/µs → first temperature (175°C)
      diF/dt=400  A/µs → second temperature (125°C)
    """
    raw_s = "".join(ch for ch in str(raw or "") if ord(ch) >= 32)
    # Fire ONLY on Infineon-specific signatures — never on the generic word
    # "switch" alone ("when switched from IF = 10 mA…" is a normal Nexperia
    # trr condition, not garbled Symbol-font text).
    _infineon_sig = ("IKW" in raw_s
                     or ("A/µs" in raw_s and "switch" in raw_s)
                     or ("Cs=" in raw_s and "Ls=" in raw_s))
    if not _infineon_sig:
        return None

    # diF/dt from readable tail
    if "0A/µs" in raw_s:
        difdt = "1000"
    elif "A/µs" in raw_s:
        difdt = "400"
    else:
        difdt = None

    # Use the most recent Tvj header PRECEDING this table
    tvj = None
    search_text = preceding_text if preceding_text else page_text
    if search_text:
        pg_clean = "".join(ch for ch in search_text if ord(ch) >= 32)
        # Find ALL Tvj= patterns in order; use the LAST one (most recent header)
        hdr_matches = list(re.finditer(
            r"[Tt]vj\s*[=\s]+(\d{2,3})(?:\s*°?C?\s*/\s*(\d{2,3}))?\s*°?C",
            pg_clean))
        if hdr_matches:
            last = hdr_matches[-1]
            tvj_primary   = last.group(1)
            tvj_secondary = last.group(2) if last.group(2) else None
            # Dual-section rule: 1000A/µs → primary temp, 400A/µs → secondary
            if tvj_secondary and difdt == "1000":
                tvj = tvj_primary + "°C"
            elif tvj_secondary and difdt == "400":
                tvj = tvj_secondary + "°C"
            else:
                tvj = tvj_primary + "°C"

    parts = []
    if tvj: parts.append(f"Tvj={tvj}")
    parts.append("VR=400V; IF=20.0A")
    if difdt: parts.append(f"diF/dt={difdt} A/µs")
    parts.append("L\u03c3=30nH; C\u03c3=40pF; switch IKW50N65H5")
    return "; ".join(parts)


def _extract_unit_from_page(pdf_path, page_num, param_symbol, table_row_y=None):
    """
    Extract the unit for a parameter by scanning the PDF character stream.
    This is needed for tables where pdfplumber does not extract a unit column.
    Returns the unit string (e.g. 'µC', 'ns', 'µA') or None.
    """
    # Known unit map for standard symbols - fallback when extraction fails
    KNOWN_UNITS = {
        'trr': 'ns', 'Qrr': 'µC', 'Irrm': 'A', 'dirr/dt': 'A/µs',
        'VF': 'V', 'IF': 'A', 'IR': 'µA', 'Rth': 'K/W',
        'Cj': 'pF', 'Cd': 'pF', 'Qc': 'nC', 'Qr': 'nC',
    }
    for sym, unit in KNOWN_UNITS.items():
        if param_symbol.startswith(sym):
            return unit
    return None


def clean_cond(s, n=155):
    """Clean a test-condition string for Excel output.
    Strips Infineon Symbol-font control chars (\x02 etc.), figure refs,
    subscript artefacts, then normalises spacing and renames garbled symbols.
    """
    if not s: return ""
    # Strip ALL control chars (Infineon Symbol-font \x02, others)
    s = ''.join(ch for ch in str(s) if ord(ch) >= 32 or ch in '\t\n ')
    # Remove figure/note refs
    s = re.sub(r"\(Note\s*\d+\)|\(\d+\)", "", s, flags=re.I)
    s = re.sub(r"\bFig\.?\s*\d+\b", "", s, flags=re.I)
    # Collapse whitespace
    s = re.sub(r"[\n\r]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Strip trailing lone-letter subscript artefacts ("T = 25 °C j" → "T = 25 °C")
    s = re.sub(r"\s+\b[A-Za-z]\b\s*$", "", s)
    s = re.sub(r"\s+\b[A-Za-z]\b\s*;", "; ", s)
    # Rename pdfplumber-garbled symbols to readable form
    # Fix pdfplumber subscript-split VR=VRRM notation: "V = V R RRM" -> "VR = VRRM"
    s = re.sub(r"\bV\s*=\s*V\s*R\s+RRM\b", "VR = VRRM", s, flags=re.I)
    s = re.sub(r"\bV\s*=\s*V\s*R\b(?!\s*R)", "VR = VR", s, flags=re.I)
    s = re.sub(r"\bI\s*=\s*([0-9.]+)\s*A\b", r"IF = \1 A", s)
    s = re.sub(r"\bT\s*=\s*([0-9]+)\s*°?\s*C\b", r"Tj = \1 °C", s)
    s = re.sub(r"\bV\s*=\s*([0-9.]+)\s*V\b", r"VR = \1 V", s)
    s = re.sub(r"\s*;\s*", "; ", s)
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s[:n]


def smatch(sc,sr): return any(re.search(p,sc,re.I) for p in sr)
def nmatch(t,nr):  return any(re.search(p,t,re.I)  for p in nr)
def _vnum(n):
    try: v=float(n); return ("." in n) or (len(n)<=5 and 0<v<100000)
    except: return False
def in_range(k,v):
    if v is None: return True
    rng=UNIT_RANGES.get(k)
    if not rng: return True
    try: return rng[0]<=float(v)<=rng[1]
    except: return True

# factors to convert a value in the given (prefixed) unit to the parameter's
# base unit used by UNIT_RANGES and the catalogue (V, A, ns, pF, µJ, nC, K/W, µA)
_BASE_UNIT = {"V_DS":"V","I_D":"A","R_DS(on)":"mΩ","V_GS(th)":"V",
              "C_iss":"pF","C_oss":"pF","C_oss(er)":"pF","C_oss(eff)":"pF","C_rss":"pF","Q_g":"nC","Q_gs":"nC",
              "Q_gd":"nC","t_d(on)":"ns","t_r":"ns","t_d(off)":"ns","t_f":"ns",
              "E_on":"µJ","E_off":"µJ","E_oss":"µJ","V_SD":"V",
              "t_rr":"ns","Q_rr":"nC","R_thJC":"K/W","R_thJA":"K/W","R_thCS":"K/W"}


_UNIT_FACTOR = {
    "V":   {"v":1.0,"mv":1e-3,"kv":1e3},
    "A":   {"a":1.0,"ma":1e-3,"ka":1e3,"µa":1e-6,"ua":1e-6},
    "ns":  {"ns":1.0,"µs":1e3,"us":1e3,"ms":1e6,"ps":1e-3},
    "nC":  {"nc":1.0,"µc":1e3,"uc":1e3,"pc":1e-3,"mc":1e6},
    "pF":  {"pf":1.0,"nf":1e3,"µf":1e6,"uf":1e6},
    "µJ": {"µj":1.0,"uj":1.0,"mj":1e3,"nj":1e-3,"j":1e6},
    "K/W": {"k/w":1.0,"°c/w":1.0,"c/w":1.0},
    "mΩ": {"mΩ":1.0,"mω":1.0,"mohm":1.0,"Ω":1e3,"ω":1e3,"ohm":1e3,"r":1e3,"kΩ":1e6,"kohm":1e6},
}

def to_base(param, value, unit_text):
    """Convert `value` expressed in `unit_text` to the parameter's base unit.
    Returns the (possibly unchanged) numeric value as a float, or the original
    on any failure so behaviour is unchanged when units are absent/unknown."""
    if value is None: return value
    base = _BASE_UNIT.get(param)
    if not base or not unit_text: return value
    tok = re.sub(r"[^\w/\u00b0\u00b5]", "", deglyph(str(unit_text)).strip().lower())
    tok = tok.split("@")[0]
    fac = _UNIT_FACTOR.get(base, {}).get(tok)
    if fac is None: return value
    try: return float(value) * fac
    except (TypeError, ValueError): return value

def _join_subscripts(ws):
    """Join a list of pdfplumber word dicts into readable text, re-attaching
    sub/superscript fragments (Infineon renders 'V_DS' as a baseline 'V' plus a
    slightly-lowered 'DS' that otherwise sorts to the end of the line).

    Also de-interleaves two stacked text baselines that the line-clustering
    collapses together — Infineon's tiny multi-line Note/Test-Condition cells
    print e.g. "TC=25°C" and "TC=100°C" only ~2-3px apart, which otherwise sort
    character-by-character into "TTCC==2150°0C°C".
    """
    if not ws:
        return ""
    ws = sorted(ws, key=lambda w: ((w["top"] + w["bottom"]) / 2, w["x0"]))
    lines = []
    for w in ws:
        cy = (w["top"] + w["bottom"]) / 2
        if lines and abs(cy - lines[-1]["cy"]) <= 5:
            L = lines[-1]; L["ws"].append(w)
            L["cy"] = (L["cy"] * L["n"] + cy) / (L["n"] + 1); L["n"] += 1
        else:
            lines.append({"cy": cy, "n": 1, "ws": [w]})

    def _emit(lws):
        """Render one baseline's words left-to-right, re-attaching subscripts."""
        lws = sorted(lws, key=lambda w: w["x0"])
        base = min(w["top"] for w in lws)
        toks = []
        for w in lws:
            t = w["text"]
            is_sub = ((w["top"] - base) > 1.0 and len(t) <= 7
                      and not re.search(r"[=<>]", t)
                      and re.fullmatch(r"[A-Za-z0-9,()/\-.]+", t) is not None)
            if is_sub and toks:
                toks[-1] = toks[-1] + t
            else:
                toks.append(t)
        return " ".join(toks)

    out = []
    for ln in lines:
        lws = ln["ws"]
        # A true single line has near-monotonic, well-separated x positions.
        # Two condition lines printed a couple of px apart instead REPEAT the
        # same x columns (one glyph per baseline), so many words collide in x.
        # Detect that and split the cluster by glyph 'top' so each physical line
        # is read intact rather than zipped.  The collision guard keeps normal
        # single lines (and their subscripts) on one baseline.
        xs = sorted(w["x0"] for w in lws)
        collisions = sum(1 for a, b in zip(xs, xs[1:]) if abs(a - b) < 2.0)
        if len(lws) >= 6 and collisions >= 3:
            # Read in (cy, x) order and start a new baseline whenever x resets
            # to the left — the hallmark of a second stacked line beginning.
            sw = sorted(lws, key=lambda w: ((w["top"] + w["bottom"]) / 2, w["x0"]))
            groups, cur, maxx = [], [], -1e9
            for w in sw:
                if cur and w["x0"] < maxx - 6:
                    groups.append(cur); cur = []; maxx = -1e9
                cur.append(w); maxx = max(maxx, w["x0"])
            if cur:
                groups.append(cur)
            if len(groups) >= 2:
                out.extend(_emit(g) for g in groups)
                continue
        out.append(_emit(lws))
    return " ".join(out).strip()

def _augment_table_sidecols(tbl_obj, inner, words):
    """Rebuild the Parameter-name column (left of the ruled grid) and the
    Note/Test-condition column (right of the grid) that pdfplumber's lattice
    parser frequently drops when those outer columns lack a vertical border
    (typical of Infineon / Toshiba characteristics tables).

    Each table row's vertical band is clamped to the next row's top so that a
    tall merged symbol/parameter cell does not vacuum up the conditions that
    belong to the sub-rows beneath it. Returns the original table unchanged
    when no side content is found (so datasheets whose grid already spans the
    full width are unaffected)."""
    try:
        bx0, _, bx1, _ = tbl_obj.bbox
        rows = list(getattr(tbl_obj, "rows", []) or [])
        if not rows or len(rows) != len(inner):
            return inner
        tops = sorted(set(round(r.bbox[1], 1) for r in rows))
        def _next_top(yt):
            for t in tops:
                if t > yt + 2:
                    return t
            return None
        def _harvest(x_test, yt, yb):
            ws = [w for w in words if x_test(w)
                  and yt - 1 <= (w["top"] + w["bottom"]) / 2 <= yb + 1]
            return _join_subscripts(ws)
        pnames, tight, full = [], [], []
        for rg in rows:
            yt, yb_full = rg.bbox[1], rg.bbox[3]
            nt = _next_top(yt)
            yb_tight = min(yb_full, nt) if nt else yb_full
            L = lambda w: w["x1"] <= bx0 + 2
            R = lambda w: w["x0"] >= bx1 - 2
            pnames.append(_harvest(L, yt, yb_tight))
            tight.append(_harvest(R, yt, yb_tight))
            # full (un-clamped) band — captures a merged condition cell that is
            # vertically centred over a group of parameter rows.
            full.append(_harvest(R, yt, yb_full))
        n = len(rows)
        # Start from the tight per-row condition; where empty (a row whose own
        # band held no text because the shared cell is centred elsewhere) fall
        # back to the full-band grab.
        cond = [tight[i] or full[i] for i in range(n)]
        # Propagate a shared condition down/up through immediately-adjacent rows
        # that came up empty (merged cell spanning several parameter rows).
        for i in range(1, n):
            if not cond[i]:
                cond[i] = cond[i - 1]
        for i in range(n - 2, -1, -1):
            if not cond[i]:
                cond[i] = cond[i + 1]
        # Upgrade a fragment to its longer adjacent superset, so every row of a
        # multi-line merged condition shows the complete test condition.
        for _ in range(3):
            for i in range(n):
                for j in (i - 1, i + 1):
                    if (0 <= j < n and cond[i] and cond[j]
                            and len(cond[j]) > len(cond[i]) and cond[i] in cond[j]):
                        cond[i] = cond[j]
        if not any(pnames) and not any(cond):
            return inner
        return [[pnames[i], *list(inner[i]), cond[i]] for i in range(n)]
    except Exception:
        return inner

def detect_structure(rows):
    if not rows: return None
    ncols=max((len(r) for r in rows[:6] if r),default=0)
    if ncols==0: return None
    roles={}
    for row in rows[:3]:
        if not row: continue
        cells=[collapse(c).lower() for c in row]
        if sum(1 for c in cells if re.match(r'^[+\-]?\d',c))>0: continue
        for ci,cl in enumerate(cells):
            if not cl: continue
            pure=not re.search(r'[()_]',cl) or cl in ('min.','typ.','max.')
            if re.search(r"\bmin\.?\b",cl) and "min" not in roles and pure: roles["min"]=ci
            if re.search(r"\btyp\.?\b|typ\.\s*value",cl) and "typ" not in roles and pure: roles["typ"]=ci
            if re.search(r"\bmax\.?\b|max\.\s*value",cl) and "max" not in roles and pure: roles["max"]=ci
            if re.search(r"\bunit\b",cl) and "unit" not in roles: roles["unit"]=ci
            if re.search(r"\bcond|test\s*cond|note|condition",cl) and "cond" not in roles: roles["cond"]=ci
            if re.search(r"\bsymbol\b",cl) and "sym" not in roles: roles["sym"]=ci
            if re.search(r"\bparameter\b|\bcharacteristic\b|\bdescription\b",cl) and "param_col" not in roles: roles["param_col"]=ci
            if re.search(r"\bvalue\b|\brating\b|\blimit\b",cl) and "val" not in roles: roles["val"]=ci
    skip=0
    for row in rows[:4]:
        if not row: continue
        if sum(1 for c in [collapse(x) for x in row] if re.match(r"^[+-]?\d",c))==0 and any(c for c in [collapse(x) for x in row]): skip+=1
        else: break
    if not any(k in roles for k in ("min","typ","max","val")):
        if ncols==7:   roles.update({"sym":0,"param_col":1,"cond":2,"min":3,"typ":4,"max":5,"unit":6})
        elif ncols>=8: roles.update({"sym":0,"param_col":1,"cond":2,"min":4,"typ":5,"max":6,"unit":7})
        elif ncols==6: roles.update({"sym":0,"param_col":1,"min":2,"typ":3,"max":4,"unit":5})
        elif ncols==5:
            # Check if col1 looks like a conditions string (long, mixed content)
            # vs a numeric column (short, mostly digits)
            col1_samples = [str(rows[r][1] or '') for r in range(max(skip,0), min(skip+3, len(rows))) if rows[r] and len(rows[r])>1]
            col1_len = max((len(s) for s in col1_samples), default=0)
            if col1_len > 15:  # long string = conditions column
                roles.update({"sym":0,"cond":1,"min":2,"typ":3,"max":4})
            else:
                roles.update({"sym":0,"min":1,"typ":2,"max":3,"unit":4})
        elif ncols==4: roles.update({"sym":0,"val":1,"unit":2})
        elif ncols==3: roles.update({"sym":0,"val":1,"unit":2})
    # Detect a SECOND condition column: the column immediately after cond
    # that also appears to hold conditions (not a numeric col, not unit).
    # This handles datasheets like STM where conditions span TWO columns
    # (e.g. col2=temperature, col3=forward current).
    if "cond" in roles:
        cond_idx = roles["cond"]
        next_idx = cond_idx + 1
        # Check if next_idx exists and is NOT already assigned a role
        assigned = set(v for k,v in roles.items() if k != "skip")
        if next_idx not in assigned and next_idx < ncols:
            # Scan data rows to see if this column often contains text (not numbers)
            text_count = 0
            for row in rows[roles.get("skip",0):roles.get("skip",0)+8]:
                if not row or next_idx >= len(row): continue
                cv = str(row[next_idx] or "").strip()
                if cv and not re.match(r'^[0-9\.\-\+]', cv):
                    text_count += 1
            if text_count >= 1:
                roles["cond2"] = next_idx

    roles["skip"]=skip; return roles

def gcell(row,idx):
    if idx is None or row is None or idx>=len(row): return ""
    return collapse(row[idx])

def row_matches(row,p,st):
    if st.get("sym") is not None:
        ck=compact(gcell(row,st["sym"]))
        if ck and smatch(ck,p["symbol_res"]): return True
    if st.get("param_col") is not None:
        pc=gcell(row,st["param_col"])
        if pc and nmatch(pc,p["name_res"]): return True
        # Quick-reference tables (Infineon 'Key performance parameters') place
        # the SYMBOL in the Parameter column rather than a separate Symbol
        # column — test the symbol patterns against it as well.
        if pc and smatch(compact(pc),p["symbol_res"]): return True
    fk=compact(" ".join(str(c or "") for c in row))
    if smatch(fk,p["symbol_res"]): return True
    fs=collapse(" ".join(str(c or "") for c in row))
    if nmatch(fs,p["name_res"]): return True
    return False


def _split_stacked(cell_str):
    """Split Infineon-style stacked cell values like "1.60\n1.65" into a list."""
    if cell_str is None: return [None]
    s = str(cell_str).strip()
    if "\n" in s:
        parts = [p.strip() for p in s.split("\n") if p.strip() and p.strip() != "-"]
        if len(parts) > 1:
            return parts
    return [s]


def row_values(row, st, p, carry_cond2=None, carry_vr=None):
    """
    Extract (typ, max, cond, new_carry) from a table row.

    Handles the STM-style split-condition layout where test conditions span
    two columns: col_cond = temperature, col_cond2 = forward current.
    Continuation rows (col_cond2 = None) carry forward the current condition
    from the preceding row.

    Also reads all text columns between param_col and the first numeric column
    to build the full condition string — this catches datasheets where
    conditions are embedded in the parameter name column or extra text columns.
    """
    def g(k):
        v = gcell(row, st.get(k))
        return None if is_blank(v) else first_num(v)

    # ── Build condition from ALL text columns in the row ─────────────────────
    # Start with the designated cond column
    cond_a = clean_cond(gcell(row, st.get("cond")) if st.get("cond") is not None else "")

    # Secondary condition column (split-cond style, e.g. STM)
    cond2_raw = gcell(row, st.get("cond2")) if st.get("cond2") is not None else ""
    cond_b = clean_cond(cond2_raw) if not is_blank(cond2_raw) else ""

    # Carry forward secondary condition from previous row when not present
    if not cond_b and carry_cond2:
        cond_b = carry_cond2

    # Also check if conditions are embedded in parameter_col (some STM absolute-ratings tables
    # put "Per diode, δ = 0.5" in the same column as the parameter name continuation)
    param_extra = ""
    if st.get("param_col") is not None:
        pc = gcell(row, st["param_col"])
        if pc and not is_blank(pc):
            # If it looks like a condition (contains "per", "δ", "=", temp) rather than a param name
            if re.search(r"per\s+(diode|device)|δ\s*=|T\s*[=<>]|I\s*=|V\s*=", pc, re.I):
                param_extra = clean_cond(pc)

    # KEY: if this row's primary cond has a temperature but no VR/bias, inject carry_vr
    # This handles IR table: row1 "Tj=25 degC, VR=200V" -> row2 "Tj=125 degC" (VR missing)
    if carry_vr and cond_a:
        has_vr = bool(re.search(r'VR\s*=|VF\s*=|V\s*=\s*[0-9]', cond_a, re.I))
        has_temp = bool(re.search(r'T[jcC]\s*[=<>]?\s*\d{2,3}', cond_a, re.I))
        if has_temp and not has_vr:
            cond_a = cond_a + "; " + carry_vr

    # Merge: primary ; secondary ; param_extra
    parts = [p for p in [cond_a, cond_b, param_extra] if p]
    # Deduplicate parts that are already contained in another
    unique = []
    for part in parts:
        if not any(part in existing for existing in unique):
            unique.append(part)
    cond = "; ".join(unique)

    _RATING_FALLBACK = {"VRRM","IF_avg","IF_rms","IFSM","IFRM"}
    # Explicit Typ/Max/Min qualifier stated INLINE in the row's text cells.
    # Some datasheets (e.g. ON Semi "Thermal Resistance, Junction to Case,
    # Max. 0.94") put the qualifier in the parameter-name/condition text instead
    # of a column header, so the value lands in a generic value column and would
    # otherwise be mislabelled.  An inline qualifier is authoritative and keeps
    # Typ and Max from ever being interchanged.
    _mtext = " ".join(str(gcell(row, st.get(k)) or "")
                      for k in ("sym", "param_col", "cond", "cond2")).lower()
    _has_max_word = re.search(r"\bmax(?:\.|imum)?\b", _mtext) is not None
    _has_typ_word = re.search(r"\btyp(?:\.|ical)?\b", _mtext) is not None
    if p.get("use_min"):
        typ = g("min") or g("val") or g("typ") or g("max"); mx = None
    else:
        typ = g("typ") or g("val"); mx = g("max")
        # A lone value explicitly qualified as Max in the row text is a MAX, not
        # a typical (and vice-versa) — the inline marker overrides the column it
        # happened to fall in, so Typ and Max are never interchanged.
        if typ is not None and mx is None and _has_max_word and not _has_typ_word:
            mx, typ = typ, None
        elif mx is not None and typ is None and _has_typ_word and not _has_max_word:
            typ, mx = mx, None
        if (typ is None and mx is not None and p["param"] in _RATING_FALLBACK
                and not _has_max_word):
            # absolute-maximum ratings: the single limit value is the rating
            # (unless the text explicitly tags it as a maximum)
            typ = mx; mx = None
        if mx == typ or is_blank(str(mx or "")): mx = None

    # KEY FIX: return the raw cond2 from THIS row (before carry-forward was applied)
    # so scan_tables knows what to carry forward to the next row.
    # cond2_raw was populated from this row's own column; cond_b may also include carry.
    own_cond2 = clean_cond(cond2_raw) if cond2_raw and not is_blank(cond2_raw) else None
    effective_carry = own_cond2 if own_cond2 is not None else carry_cond2
    return typ, mx, cond, effective_carry

def scan_tables(tables,p,page_texts=None,sec_hdrs=None,words_pp=None):
    records=[]; multi=p.get("multi_cond",False)
    for tbl in tables:
        rows=tbl["data"]
        if not rows: continue
        # Skip quick-reference summary tables (2-col Symbol/Value tables on page 1)
        # These contain shorthand like "2 x 30 A" and are unreliable for extraction.
        ncols_tbl = max((len(r) for r in rows[:3] if r), default=0)
        if ncols_tbl <= 2:
            # Only use if it has a proper Min/Typ/Max structure
            flat_hdr = " ".join(str(c or "").lower() for c in (rows[0] or []))
            if not any(k in flat_hdr for k in ("min","typ","max","value","limit")):
                continue  # skip this table
        st=detect_structure(rows)
        if not st: continue
        skip=st.get("skip",0); i=skip
        while i<len(rows):
            row=rows[i]
            if not row: i+=1; continue
            if row_matches(row,p,st):
                grp=[i]; j=i+1
                while j<len(rows):
                    nr=rows[j]
                    if not nr: j+=1; continue
                    sc=compact(gcell(nr,st.get("sym")))
                    if is_blank(sc) or row_matches(nr,p,st): grp.append(j);j+=1
                    else: break
                carry2 = None
                carry_vr = None  # carries VR/bias condition across temp rows
                # Get page text context for Infineon switching condition decoding
                pg_txt = ""
                preceding_txt = ""
                if page_texts and tbl.get("page"):
                    pg_idx = tbl["page"] - 1
                    if 0 <= pg_idx < len(page_texts):
                        pg_txt = page_texts[pg_idx]
                        # Find the most recent section header ABOVE this table (by Y position)
                        # This correctly assigns Tvj temperature from section headers
                        tbl_y = tbl.get("y_top")
                        pnum  = tbl.get("page")
                        page_sec_hdrs = sec_hdrs.get(pnum, []) if sec_hdrs else []
                        
                        if tbl_y is not None and page_sec_hdrs:
                            # Find headers ABOVE the table (y_top of header <= y_top of table)
                            # In Infineon PDFs, Y coordinates are negative (more negative = higher)
                            # "above" means y_header <= y_table (less negative)
                            above = [h for h in page_sec_hdrs if h["y"] <= tbl_y + 5]
                            # Sort by y descending → last before table
                            above.sort(key=lambda x: x["y"], reverse=True)
                            if above:
                                hdr = above[0]
                                # Build preceding_txt as a Tvj= string for the decoder
                                if hdr["t2"]:
                                    preceding_txt = f"Tvj={hdr['t1']}°C/{hdr['t2']}°C"
                                else:
                                    preceding_txt = f"Tvj={hdr['t1']}°C"
                            else:
                                preceding_txt = pg_txt
                        else:
                            preceding_txt = pg_txt
                # ── TABLE-LEVEL DECODED CONDITION ─────────────────────────────────────
                # Scan ALL rows of this table (not just the matched group) to find
                # the decoded condition. This is essential for Infineon switching tables
                # where trr row[0] has the garbled condition but Qrr row[1] is empty.
                # Both trr and Qrr share the SAME condition in the same table.
                table_decoded_cv = None
                for ri_scan in range(skip, len(rows)):
                    rr_scan = rows[ri_scan]
                    if not rr_scan: continue
                    # Get raw condition from any column that might have it
                    raw_cv = ""
                    if st.get("cond") is not None:
                        raw_cv = gcell(rr_scan, st["cond"])
                    if not raw_cv and st.get("cond2") is not None:
                        raw_cv = gcell(rr_scan, st["cond2"])
                    if not raw_cv: continue
                    # Decode Infineon garbled conditions
                    if len(raw_cv) > 15:
                        dec = _decode_infineon_switching_cond(raw_cv, pg_txt, preceding_txt)
                        if dec:
                            table_decoded_cv = dec; break
                    elif len(raw_cv) > 3 and not is_blank(raw_cv):
                        table_decoded_cv = clean_cond(raw_cv); break

                # ── GROUP-LEVEL PRE-SCAN (Infineon garbled conditions only) ──────────
                # Only set group_decoded_cv when the condition string is garbled/unreadable
                # (Infineon Symbol-font encoding). For normal readable conditions (STM, WeEN etc.)
                # each row must keep its own per-row condition (different IF/Temp per row).
                group_decoded_cv = None
                for ri_pre in grp:
                    rr_pre = rows[ri_pre]
                    if not rr_pre: continue
                    _, _, cv_pre, _ = row_values(rr_pre, st, p)
                    if cv_pre and len(cv_pre) > 15:
                        dec = _decode_infineon_switching_cond(cv_pre, pg_txt, preceding_txt)
                        if dec:
                            group_decoded_cv = dec; break
                    # NOTE: do NOT fall through to set group_decoded_cv from a normal readable
                    # condition string — that would override all rows in the group with one condition.

                # Use table-level decoded condition if group has none (Infineon garbled only)
                if not group_decoded_cv and table_decoded_cv:
                    # Only use if table_decoded_cv looks like a decoded Infineon condition
                    # (contains "Tvj=" or "diF/dt" — not just a plain temp string)
                    if re.search(r'Tvj\s*=|diF/dt|IKW|A/µs', table_decoded_cv, re.I):
                        group_decoded_cv = table_decoded_cv

                # ── PER-ROW PROCESSING ────────────────────────────────────────────────
                for ri in grp:
                    rr=rows[ri]
                    if not rr: continue
                    tv, mv, cv, carry2 = row_values(rr, st, p, carry_cond2=carry2, carry_vr=carry_vr)
                    # Update carry_vr from this row's merged condition
                    _vr_m = re.search(r'VR\s*=\s*[A-Za-z0-9.]+', cv, re.I)
                    if _vr_m:
                        carry_vr = _vr_m.group(0).strip()

                    # Apply Infineon decoded condition (garbled symbol-font only)
                    if group_decoded_cv:
                        cv = group_decoded_cv
                    elif cv and len(cv) > 20:
                        dec = _decode_infineon_switching_cond(cv, pg_txt, preceding_txt)
                        if dec:
                            cv = dec

                    # KEY FIX: extract temperature from the PRIMARY cond column only
                    # (not from the full merged cv string which may have carry-forward IF values)
                    # This ensures each row gets its own correct temperature
                    _primary_cond_raw = gcell(rr, st.get("cond")) if st.get("cond") is not None else ""
                    _primary_cond = clean_cond(_primary_cond_raw)

                    # ── Handle stacked cells (Infineon "1.60\n1.65" style) ──────────
                    # Read RAW cell (before gcell/collapse strip newlines) for stacking
                    # Use "typ" col first, fallback to "val" col (2-col tables like Max Ratings)
                    _typ_col = st.get("typ") if st.get("typ") is not None else st.get("val")
                    _max_col = st.get("max")
                    raw_typ = str(rr[_typ_col] if _typ_col is not None and _typ_col < len(rr) else "")
                    raw_max = str(rr[_max_col] if _max_col is not None and _max_col < len(rr) else "")
                    # Use raw cell if it contains \n (true stacking); else use extracted tv/mv
                    typ_vals = _split_stacked(raw_typ) if raw_typ and "\n" in raw_typ else _split_stacked(tv) if tv else [tv]
                    max_vals = _split_stacked(raw_max) if raw_max and "\n" in raw_max else _split_stacked(mv) if mv else [mv]
                    # Size the loop by the LONGEST of typ/max — a max-rating row
                    # often has a blank "-/-" typ cell (which collapses to one
                    # token) alongside two real max values ("18\n11"); using only
                    # len(typ_vals) would truncate the second max value.
                    n_rows   = max(len(typ_vals), len(max_vals), 1)
                    typ_vals = typ_vals + [None] * (n_rows - len(typ_vals))
                    max_vals_padded = max_vals + [None] * (n_rows - len(max_vals))

                    # Extract temperatures for stacked rows from condition string
                    stacked_temps = []
                    if n_rows > 1 or "\n" in str(tv or "") or "\n" in str(mv or ""):
                        # Try 0: de-spaced condition.  Infineon's stacked Note
                        # cells come through fragmented (e.g. "T = 2 5 ° C TC
                        # =100 °C"); collapsing the spaces and reading every
                        # "<digits>°C" recovers the per-row temperatures in order.
                        _cvds = re.sub(r"\s+", "", cv or "")
                        _t0 = [int(t) for t in re.findall(r"(\d{1,3})\s*°?C", _cvds)
                               if 0 <= int(t) <= 250]
                        if len(_t0) >= len(typ_vals):
                            stacked_temps = _t0[:len(typ_vals)]
                        # Try 1: clean condition string first
                        if not stacked_temps:
                            stacked_temps = [int(t) for t in re.findall(r"Tvj\s*=\s*(\d{2,3})", cv)]
                        if not stacked_temps:
                            stacked_temps = [int(t) for t in re.findall(r"T[jcC]\s*=\s*(\d{2,3})", cv)]
                        if not stacked_temps:
                            stacked_temps = [int(t) for t in re.findall(r"(\d{2,3})\s*°C", cv)
                                             if 0 <= int(t) <= 250]
                        # Try 2: extract from page header + Tvjmax
                        # Infineon stacks 2 values: [header_temp, tvjmax]
                        if not stacked_temps and len(typ_vals) == 2 and pg_txt:
                            pg_clean = "".join(ch for ch in pg_txt if ord(ch) >= 32)
                            # Page header temp: "at Tvj = 25°C"
                            hdr_m = re.search(r"at\s*Tvj\s*=\s*(\d{2,3})\s*°?C", pg_clean, re.I)
                            hdr_t = int(hdr_m.group(1)) if hdr_m else 25
                            # Tvjmax from same page or from page text
                            tvjmax_m = re.search(r"[+-]?\d+\.\.\.[+]?(\d{2,3})\s*°C", pg_clean)
                            tvjmax = int(tvjmax_m.group(1)) if tvjmax_m else None
                            if not tvjmax and page_texts:
                                # Search all pages for Tvjmax
                                for pt in page_texts:
                                    m = re.search(r"[+-]?\d+\.\.\.[+]?(\d{2,3})\s*°C",
                                                  "".join(ch for ch in pt if ord(ch) >= 32))
                                    if m: tvjmax = int(m.group(1)); break
                            if tvjmax:
                                stacked_temps = [hdr_t, tvjmax]

                        # Try 3: word-level temperature extraction
                        # For tables with TC conditions (IF_avg max ratings)
                        if not stacked_temps and words_pp and tbl.get("page"):
                            pnum_w = tbl["page"]
                            pg_words = words_pp.get(pnum_w, [])
                            # The table Y range; each stacked sub-row has its own Y band
                            tbl_y_top = tbl.get("y_top")
                            if tbl_y_top is not None and pg_words:
                                # Scan words in a wide range around the table for temperature patterns
                                # (table bbox y_top may be the table top, but condition text
                                # can span the full table height)
                                from collections import defaultdict
                                y_bands = defaultdict(list)
                                # Find actual y extent of table content (min/max y of all words near table)
                                tbl_words = [w for w in pg_words
                                             if tbl_y_top - 60 <= w["top"] <= tbl_y_top + 60]
                                y_min_w = min((w["top"] for w in tbl_words), default=tbl_y_top - 10)
                                y_max_w = max((w["top"] for w in tbl_words), default=tbl_y_top + 10)
                                for w in pg_words:
                                    if y_min_w - 5 <= w["top"] <= y_max_w + 5:
                                        ykey = round(w["top"] * 2) / 2  # 0.5px resolution
                                        y_bands[ykey].append(
                                            "".join(ch for ch in w["text"] if ord(ch) >= 32))
                                # For each Y band, try to find a temperature
                                band_temps = []
                                for yband in sorted(y_bands.keys()):
                                    band_text = " ".join(y_bands[yband])
                                    # Try direct temperature match
                                    tm = re.findall(r"T[Cj]\s*[=<>≤≥]\s*(\d{2,3})", band_text, re.I)
                                    if tm:
                                        band_temps.append(int(tm[0]))
                                    else:
                                        # Scattered digits: concatenate adjacent single digits
                                        toks = band_text.split()
                                        j_w = 0
                                        while j_w < len(toks):
                                            if len(toks[j_w]) == 1 and toks[j_w].isdigit():
                                                run = [toks[j_w]]
                                                k_w = j_w + 1
                                                while (k_w < len(toks) and len(toks[k_w]) == 1
                                                       and toks[k_w].isdigit()):
                                                    run.append(toks[k_w]); k_w += 1
                                                if len(run) >= 2:
                                                    try:
                                                        v = int("".join(run))
                                                        if 20 <= v <= 250:
                                                            band_temps.append(v)
                                                    except: pass
                                                j_w = k_w
                                            else:
                                                j_w += 1
                                if len(band_temps) >= 2:
                                    stacked_temps = band_temps[:len(typ_vals)]

                    for sub_i, (sv, smv) in enumerate(zip(typ_vals, max_vals_padded)):
                        sv_num  = first_num(str(sv  or ""))
                        smv_num = first_num(str(smv or "")) if smv else None
                        if sv_num is None and smv_num is None: continue
                        # read the row's unit FIRST and normalise to base unit so
                        # the range check sees e.g. 1000 mV → 1.0 V, 450 mA → 0.45 A
                        extracted_unit = None
                        if st.get("unit") is not None:
                            u_raw = gcell(rr, st["unit"])
                            if u_raw and not is_blank(u_raw):
                                extracted_unit = u_raw.strip()
                            else:
                                # Merged unit cell: datasheets often print the
                                # unit once for a group of rows (e.g. ON Semi
                                # lists "°C/W" only on the RthJC row, leaving the
                                # RthJA row's unit cell blank).  Inherit the
                                # nearest non-blank unit from the same column so
                                # the displayed unit matches the datasheet rather
                                # than the catalogue default.  Constrained to an
                                # accepted unit for this parameter so an unrelated
                                # neighbour can't corrupt it.
                                _acc = _UNIT_ACCEPT.get(p["param"], {})
                                for _dist in range(1, min(len(rows), 6)):
                                    for _ri2 in (ri - _dist, ri + _dist):
                                        if not (0 <= _ri2 < len(rows)):
                                            continue
                                        _u2 = gcell(rows[_ri2], st["unit"])
                                        if _u2 and not is_blank(_u2):
                                            _u2s = _u2.strip()
                                            if (not _acc) or _norm_unit(_u2s) in _acc \
                                                    or _u2s in _acc:
                                                extracted_unit = _u2s
                                            break
                                    if extracted_unit is not None:
                                        break
                        sv_base  = to_base(p["param"], sv_num,  extracted_unit)
                        smv_base = to_base(p["param"], smv_num, extracted_unit)
                        if not in_range(p["param"], sv_base):  sv_num  = None
                        if not in_range(p["param"], smv_base): smv_num = None
                        if sv_num is None and smv_num is None: continue

                        # ── Temperature for this sub-row ──────────────────────────────
                        if stacked_temps and sub_i < len(stacked_temps):
                            # Stacked cell: use the sub_i-th temperature from condition
                            tc = stacked_temps[sub_i]
                            ih = tc >= 90
                        else:
                            # Single value: extract temp from PRIMARY cond column first,
                            # then fall back to full merged condition or row text
                            # This is the KEY FIX: use per-row primary cond for temperature
                            all_temps = (get_temps(_primary_cond)
                                        or get_temps(cv)
                                        or get_temps(collapse(" ".join(str(c or "") for c in rr))))
                            if all_temps:
                                tc = all_temps[0]; ih = tc >= 90
                            else:
                                tc = None; ih = False

                        # When the condition is temperature-only (e.g. the ID
                        # max-rating "TC=25°C / TC=100°C"), give each stacked row
                        # its own clean condition instead of the merged, slightly
                        # fragmented two-line string.
                        row_cond = cv
                        if stacked_temps and sub_i < len(stacked_temps):
                            _ds = re.sub(r"\s+", "", cv or "")
                            _residual = re.sub(r"T?[CcJjVA]?=?-?\d{1,3}°?C", "", _ds)
                            if _ds and not re.search(r"[A-Za-z0-9]", _residual):
                                row_cond = f"TC = {stacked_temps[sub_i]} °C"

                        records.append({"typ": fmt_val(sv_num),
                                        "max": fmt_val(smv_num) if smv_num else "—",
                                        "cond": row_cond, "temp_c": tc, "is_high": ih,
                                        "extracted_unit": extracted_unit})
                        if not multi: return records
                i=j
            else: i+=1
    return records

def scan_text(text,p):
    records=[]; multi=p.get("multi_cond",False)
    for line in text.split("\n"):
        lc=collapse(line); lk=compact(line)
        if not (smatch(lk,p["symbol_res"]) or nmatch(lc,p["name_res"])): continue
        tmps=get_temps(lc); ih=any(t>=90 for t in tmps); tc=max(tmps) if tmps else None
        nums=[n for n in re.findall(r"[+-]?\d+\.?\d*(?:[eE][+-]?\d+)?",lc) if _vnum(n)]
        if not nums: continue
        tv=nums[1] if len(nums)>=3 else nums[0]
        mv=nums[2] if len(nums)>=3 else None
        if not in_range(p["param"],tv): continue
        # An inline "Max."/"Maximum" (or "Typ."/"typical") qualifier in the line
        # is authoritative: a lone value tagged Max is a maximum, not a typical,
        # so the two are never interchanged (e.g. ON Semi "Thermal Resistance,
        # Junction to Case, Max. 0.94").
        _low = lc.lower()
        _has_max_word = re.search(r"\bmax(?:\.|imum)?\b", _low) is not None
        _has_typ_word = re.search(r"\btyp(?:\.|ical)?\b", _low) is not None
        if mv is None and _has_max_word and not _has_typ_word:
            records.append({"typ":"—","max":fmt_val(tv),
                            "cond":lc[:130],"temp_c":tc,"is_high":ih,
                            "extracted_unit":None})
        else:
            records.append({"typ":fmt_val(tv),"max":fmt_val(mv) if mv else "—",
                            "cond":lc[:130],"temp_c":tc,"is_high":ih,
                            "extracted_unit":None})
        if not multi: break
    return records

def enrich_temps(recs, full):
    """
    Assign temperature to records that have none.

    Three-pass strategy (most-specific first):
    1. Pattern: "– – VALUE T≤TEMP °C" — Infineon max-ratings table format
       (value appears as a standalone result followed directly by temperature)
    2. Pattern: value immediately followed by temperature keyword
       BUT skip lines where value appears as a condition parameter (I = VALUE A)
    3. Fallback: any line with both value and temperature
    """
    if not recs: return recs
    lines = full.split("\n")

    for rec in recs:
        # Skip if temp already assigned (unless it suspiciously equals the typ value)
        if rec["temp_c"] is not None:
            # Sanity-check: if the assigned temp equals the typ value, it was likely a
            # mis-extraction (e.g. VRRM=200V → temp=200°C). Clear it so we can reassign.
            if rec["typ"] != "—":
                try:
                    if abs(float(rec["typ"]) - rec["temp_c"]) < 1:
                        rec["temp_c"] = None  # was incorrectly set to the param value
                    else:
                        continue
                except (ValueError, TypeError):
                    continue
            else:
                continue
        if rec["typ"] == "—":
            continue
        val = rec["typ"]
        found = False

        # Pass 0: "Tvj ≥ X°C" / "Tvj ≤ X°C" pattern — spec-at-temperature
        # e.g. VRRM line: "Repetitive peak reverse voltage, Tvj ≥ 25°C"
        # The temperature is the MINIMUM spec temperature, use it directly.
        for line in lines:
            # Strip ALL control chars (-) before matching
            lc = re.sub(r"[\x00-\x1f]", " ", line); lc = " ".join(lc.split())
            if val not in lc: continue
            m_ineq = re.search(r"T[vj]+\s*[\u2265\u2264\u2021\u2267\u2266><=\u2212]+\s*(\d{2,3})\s*°?\s*C", lc, re.I)
            if m_ineq:
                t = int(m_ineq.group(1))
                if 0 <= t <= 250:
                    rec["temp_c"] = t; rec["is_high"] = t >= 90
                    found = True; break

        if found: continue

        # Pass 1: Infineon/strict format — "– – VALUE T[≤<=]TEMP °C"
        # The value appears as a MAX/TYP result, directly followed by temp condition
        p1 = re.compile(
            r"(?:[\-\u2013\u2014]\s+[\-\u2013\u2014]\s+|"   # "– – " prefix
            r"[0-9.]+\s+[\-\u2013\u2014]\s+)"                   # "NUM – " prefix
            + re.escape(val) +
            r"\s+T[Cj]?\s*[\u2264\u2265<>=]+\s*(\d{2,3})\s*°?\s*C",
            re.I)
        for line in lines:
            lc = " ".join(line.split())
            m = p1.search(lc)
            if m:
                t = int(m.group(1)); rec["temp_c"] = t; rec["is_high"] = t >= 90
                found = True; break

        if found: continue

        # Pass 2: avoid false matches where value appears as a condition
        # e.g. skip "I = 16 A, T = 25 °C" for value "16"
        for line in lines:
            lc = " ".join(line.split())
            if val not in lc: continue
            # Skip lines where value appears after "I =", "V =", "IF =", etc.
            if re.search(r"[IVPQCR][A-Za-z_]*\s*=\s*" + re.escape(val) + r"\s*[AVWΩCµnp]", lc, re.I):
                continue
            tj = re.search(r"T[jcC]\s*[=\u2264<>]?\s*(\d{2,3})\s*°?\s*[Cc]", lc, re.I)
            if tj:
                t = int(tj.group(1)); rec["temp_c"] = t; rec["is_high"] = t >= 90
                found = True; break
            tmps = get_temps(lc)
            if tmps:
                t = tmps[0]
                # Extra sanity: temp should be in diode operating range
                if 0 <= t <= 250:
                    # Don't assign temperature if it equals the param value itself
                    try:
                        if abs(float(val) - t) < 1:
                            continue
                    except (ValueError, TypeError):
                        pass
                    rec["temp_c"] = t; rec["is_high"] = t >= 90
                    found = True; break

        if found: continue

        # Pass 3: any line with value and temperature (legacy fallback)
        for line in lines:
            lc = collapse(line)
            if val not in lc: continue
            tmps = get_temps(lc)
            if tmps and 0 <= tmps[0] <= 250:
                # Sanity guard: don't assign a temperature that equals the param value itself
                try:
                    if abs(float(val) - tmps[0]) < 1:
                        continue  # e.g. VRRM=200V should not get temp_c=200
                except (ValueError, TypeError):
                    pass
                rec["temp_c"] = tmps[0]; rec["is_high"] = tmps[0] >= 90; break

    return recs

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 4 — rd TEXT/TABLE FALLBACK (5 strategies)
# ═════════════════════════════════════════════════════════════════════════════

# Pre-compiled patterns — avoids escape/newline issues in function body
_PAT_VORS = re.compile(
    r"V[o0]?\s*=\s*([0-9]+\.?[0-9]*)\s*V"
    r"[;,\s]+"
    r"[Rr][sS]?\s*=\s*([0-9]+\.?[0-9]*)\s*[\u03a9\u2126Oo]",
    re.I)

_PAT_PEQN = re.compile(
    r"P[\w()\s]*=\s*([0-9]+\.?[0-9]*)\s*[x\xd7*]\s*I[^\n+]{0,30}"
    r"\+\s*([0-9]+\.?[0-9]*)\s*[x\xd7*]\s*I",
    re.I)

def _rd_text_fallback(all_tables, full_text):
    """
    5 strategies in priority order:
    S0  Vo/Rs model   "V = 0.993 V; R = 0.0660 Ω"    WeEN, Nexperia
    S1  Loss eqn      "P = 0.58 x I + 0.0037 x I 2"  STM, Vishay
    S2  Explicit RDIFF table row
    S3  Infineon RDIFF(Tj) polynomial
    S4  Two-point dVF/dIF
    """
    # S0 — Vo/Rs
    m = _PAT_VORS.search(full_text)
    if m:
        try:
            vf0=float(m.group(1)); rd=float(m.group(2))
            if in_range("rd",str(rd)) and 0.1<vf0<3.0:
                return [{"typ":f"{rd:.4f}","max":"—",
                         "cond":f"Vo={vf0}V Rs=rd={rd} Ohm [Vo/Rs model in datasheet text]",
                         "temp_c":25,"is_high":False,"source":"Vo/Rs linearisation (WeEN/Nexperia)"}]
        except (ValueError,TypeError): pass

    # S1 — P = a*IF + b*IF² conduction loss equation
    m = _PAT_PEQN.search(full_text)
    if m:
        try:
            vf0=float(m.group(1)); rd=float(m.group(2))
            if in_range("rd",str(rd)) and 0.1<vf0<3.0:
                return [{"typ":f"{rd:.4f}","max":"—",
                         "cond":f"P={vf0}*IF+{rd}*IF2 loss eq: rd=Rs={rd} Ohm [STM/Vishay style]",
                         "temp_c":25,"is_high":False,"source":"Conduction-loss equation (STM/Vishay)"}]
        except (ValueError,TypeError): pass

    # S2 — explicit RDIFF/rd table row
    p_rd=dict(param="rd",full="rd",unit="Ohm",section="",note="",
              symbol_res=[r"^RDIFF$",r"^R_?DIFF$",r"^rd$",r"^RD$",r"^rs$",r"^RS$"],
              name_res=[r"differential\s+res",r"dynamic\s+res",
                        r"slope\s+res",r"incremental\s+res",r"series\s+res"],
              use_min=False,multi_cond=False)
    recs=scan_tables(all_tables,p_rd) or scan_text(full_text,p_rd)
    if recs:
        r=recs[0]
        try:
            v=float(r["typ"])
            if v>1.5: v/=1000.0
            r["typ"]=f"{v:.4f}"
        except (ValueError,TypeError): pass
        r["source"]="Explicit RDIFF table row"; return recs

    # S3 — Infineon RDIFF(Tj) = A*Tj^2 + B*Tj + C
    lines=full_text.split("\n")
    for li,line in enumerate(lines):
        if not re.search(r"differential\s+res|RDIFF|R_?DIFF",line,re.I): continue
        win="\n".join(lines[max(0,li-2):li+22])
        win=re.sub(r"[\uf000-\uf0ff]"," ",win)
        coeff={}
        for L in ("A","B","C"):
            m2=re.search(r"(?<![A-Za-z])"+L+r"(?![A-Za-z])\s*[=:\s]\s*"
                r"([+-]?[0-9]+\.?[0-9]*)(?:\s*(?:[eE*]|10)\s*([+-]?\d+))?",win)
            if m2:
                mt=float(m2.group(1))
                exp=int(m2.group(2)) if m2.group(2) else 0
                if not m2.group(2):
                    tl=win[m2.end():]; em=re.match(r"\s*10([+-]\d+)",tl)
                    if em: exp=int(em.group(1))
                coeff[L]=mt*(10**exp)
        if "A" in coeff and "B" in coeff and "C" in coeff:
            A,B,C=coeff["A"],coeff["B"],coeff["C"]
            rd25=A*625+B*25+C
            if in_range("rd",str(rd25)):
                return [{"typ":f"{rd25:.4f}","max":"—",
                         "cond":f"RDIFF=A*Tj2+B*Tj+C at 25C [A={A:.3e} B={B:.3e} C={C:.3e}]",
                         "temp_c":25,"is_high":False,"source":"Infineon RDIFF polynomial"}]
        break

    # S4 — two-point dVF/dIF
    def tbl_pairs():
        vp=dict(param="VF",full="",unit="V",section="",note="",
                symbol_res=[r"^VF$",r"^V\(?F\)?$",r"^VD$",r"^VFM$"],
                name_res=[r"forward\s+voltage"],use_min=False,multi_cond=True)
        pairs=[]
        for rec in scan_tables(all_tables,vp):
            if rec["is_high"]: continue
            im=re.search(r"I[FD]?\s*=\s*([0-9.]+)",rec["cond"],re.I)
            if im and rec["typ"]!="—":
                try:
                    I=float(im.group(1)); V=float(rec["typ"])
                    if 0.1<V<5.0 and 0<I<5000: pairs.append((I,V))
                except: pass
        seen=set(); out=[]
        for p in sorted(pairs,key=lambda x:x[0]):
            if p not in seen: seen.add(p); out.append(p)
        return out

    def txt_pairs():
        pairs=[]
        pt=re.compile(r"I[FD]\s*=\s*([0-9.]+)\s*A.*?V[FD]\s*=\s*([0-9.]+)\s*V"
                      r"|V[FD]\s*=\s*([0-9.]+)\s*V.*?I[FD]\s*=\s*([0-9.]+)\s*A",re.I)
        for mx in pt.finditer(full_text):
            if mx.group(1) and mx.group(2): I,V=float(mx.group(1)),float(mx.group(2))
            else: V,I=float(mx.group(3)),float(mx.group(4))
            if 0.1<V<5.0 and 0<I<5000: pairs.append((I,V))
        seen=set(); out=[]
        for p in sorted(set(pairs),key=lambda x:x[0]):
            if p not in seen: seen.add(p); out.append(p)
        return out

    for lbl,pairs in [("table",tbl_pairs()),("text",txt_pairs())]:
        if len(pairs)>=2:
            (I1,V1),(I2,V2)=pairs[0],pairs[-1]
            if I2!=I1:
                rd=(V2-V1)/(I2-I1)
                if in_range("rd",str(rd)):
                    return [{"typ":f"{rd:.4f}","max":"—",
                             "cond":f"rd=(V2-V1)/(I2-I1)=({V2}-{V1})/({I2}-{I1}) 25C [{lbl}]",
                             "temp_c":25,"is_high":False,"source":f"Two-point dVF/dIF [{lbl}]"}]
    return []

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 5 — DEVICE INFO
# ═════════════════════════════════════════════════════════════════════════════

_PKG={"TO220","TO247","TO252","TO263","TO268","TO3P","TO3PF","D2PAK","D3PAK","DPAK",
      "SOT227","SOT23","DFN","QFN","SOP","TO220F","TO220AB","TO247AC","TO247N","TO263AB",
      "TO220FP","TO220F2L","TO247N"}

def get_device_info(text):
    info={"part":"Unknown","mfr":"Unknown"}
    SKIP={"THE","AND","FOR","MAX","MIN","TYP","MOSFET","NMOS","PMOS","SIC","GAN","IGBT",
          "COOLSIC","PUBLIC","TABLE","FINAL","REV","CHANNEL","POWER","DATA","SHEET","CASE",
          "JEDEC","PARAMETER","SYMBOL","DIODE","SCHOTTKY","NOTE","FIGURE","VALUE","UNIT",
          "FEATURES","PRODUCT","DESCRIPTION","GENERAL","PURPOSE","APPLICATION",
          "SEMICONDUCTOR","DATASHEET","PRELIMINARY"}
    SKIP.update(p.upper() for p in _PKG)
    for pat in [r"(?:Part\s*(?:No|Number|#)[.:\s]+)([A-Z0-9][A-Z0-9\-]{3,24})",
                r"(?:Ordering\s*(?:code|info)[.:\s]+)([A-Z0-9][A-Z0-9\-]{3,24})",
                r"(?:Device|Type|Model)[.:\s]+([A-Z0-9][A-Z0-9\-]{3,24})",
                r"\b([A-Z]{2,6}\d{1,3}[A-Z]{1,2}\d{3,8}[A-Z0-9\-]{0,8})\b",
                r"\b([A-Z]{2,6}\d{2,8}[A-Z0-9\-]{0,12})\b"]:
        m=re.search(pat,text[:5000])
        if m:
            c=m.group(1).strip("-")
            if c.upper() not in SKIP and len(c)>=5 and re.search(r"\d",c):
                info["part"]=c; break
    MFR={"Infineon Technologies":["Infineon","INFINEON","CoolSiC","CoolMOS"],
         "STMicroelectronics":["STMicro","STB","STP","STW","STPOWER","STPSC"],
         "WeEN Semiconductors":["WeEN","WEEN","WeEn"],
         "Diotec":["Diotec","DIOTEC"],
         "Taiwan Semiconductor":["Taiwan Semiconductor","TSC"],
         "Wolfspeed / Cree":["Wolfspeed","Cree","C3D","C2D","C4D"],
         "ROHM":["ROHM","Rohm"],
         "ON Semiconductor":["ON Semiconductor","onsemi","Fairchild"],
         "Vishay / Siliconix":["Vishay","Siliconix","VS-"],
         "Nexperia":["Nexperia","NEXPERIA"],
         "GeneSiC":["GeneSiC"],"SemiQ":["SemiQ"],
         "IXYS / Littelfuse":["IXYS","Littelfuse"],
         "Microchip / Microsemi":["Microsemi","Microchip"],
         "Toshiba":["Toshiba","TOSHIBA"]}
    t=text[:6000]
    for mfr,kws in MFR.items():
        if any(k in t for k in kws): info["mfr"]=mfr; break
    return info

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 6 — MAIN EXTRACTION
# ═════════════════════════════════════════════════════════════════════════════

def _skey(row):
    sec=SECTION_ORDER.get(row["section"],9)
    try: idx=next(i for i,p in enumerate(PARAMS) if p["param"]==row["param"])
    except StopIteration: idx=999
    return (sec,idx)

def extract(pdf_path, debug_png=None, interactive=True):
    print(f"\n[1/4] Reading: {pdf_path}")
    all_text, all_tables = [], []
    section_headers = {}  # section_headers[page_num] = list of (y, t1, t2)
    words_per_page = {}  # words_per_page[pnum] = list of word dicts

    with pdfplumber.open(pdf_path) as pdf:
        for pnum, page in enumerate(pdf.pages, 1):
            txt = page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            txt = _decode_cid_text(txt)
            all_text.append(txt)

            # Extract section headers + all words (for stacked-cell temperature lookup)
            try:
                page_words = page.extract_words()
                words_per_page[pnum] = page_words
                page_headers = []
                for w in page_words:
                    wt = "".join(ch for ch in w["text"] if ord(ch) >= 32)
                    if "Tvj" in wt and any(c.isdigit() for c in wt):
                        m = re.search(r"Tvj\s*[=\s]+(\d{2,3})(?:\s*°?C?\s*/\s*(\d{2,3}))?", wt)
                        if m:
                            page_headers.append({"y": w["top"],
                                                "t1": m.group(1),
                                                "t2": m.group(2) if m.group(2) else None})
                if page_headers:
                    section_headers[pnum] = page_headers
            except Exception:
                pass

            # Extract tables with bounding boxes
            try:
                _pw = words_per_page.get(pnum) or page.extract_words()
                for ti, tbl_obj in enumerate(page.find_tables()):
                    tbl_data = tbl_obj.extract()
                    if tbl_data:
                        # Recover dropped Parameter / Note-condition columns.
                        tbl_data = _augment_table_sidecols(tbl_obj, tbl_data, _pw)
                        # Skip pseudo-tables that pdfplumber lifts out of the
                        # diagram grids — they carry "Diagram N" titles and
                        # "X=f(Y)" axis captions (in cells or harvested
                        # conditions), not real parameter rows.
                        _flat = " ".join(str(c) for r in tbl_data for c in r if c)
                        if (re.search(r"Diagram\s*\d", _flat)
                                or re.search(r"=\s*f\s*\(", _flat)):
                            continue
                        y_top = tbl_obj.bbox[1] if hasattr(tbl_obj, "bbox") else None
                        all_tables.append({"page": pnum, "data": tbl_data,
                                           "y_top": y_top, "table_index": ti})
            except Exception:
                for tbl in page.extract_tables():
                    if tbl:
                        all_tables.append({"page": pnum, "data": tbl,
                                           "y_top": None, "table_index": None})
    full_text="\n".join(all_text)
    print(f"      {len(full_text):,} chars, {len(all_tables)} tables")

    print("[2/4] Identifying device ...")
    info=get_device_info(full_text)
    print(f"      Part: {info['part']}  |  Mfr: {info['mfr']}")

    print("[3/4] Extracting parameters ...")
    _blines=baseline_lines(pdf_path)
    _family_devs, _dev_variants = find_device_columns(all_tables)
    if _family_devs:
        print(f"      Multi-device family table: {', '.join(_family_devs[:6])}")
        if info.get("part") in (None, "", "Unknown"):
            info["part"] = "/".join(_family_devs[:6])
    output_rows=[]
    for p in PARAMS:
        recs=scan_tables(all_tables,p,page_texts=all_text,sec_hdrs=section_headers,words_pp=words_per_page)
        # ── family datasheets: one row per device when values differ ────────
        _dv=None
        for _sym,_vals in _dev_variants.items():
            if smatch(_sym,p["symbol_res"]):
                _dv=_vals; break
        if _dv:
            _vrecs=[]
            for _dname,_v in _dv:
                if is_blank(_v): continue
                _n=first_num(_v)
                if _n is None or not in_range(p["param"],str(_n)): continue
                _vrecs.append({"typ":fmt_val(str(_n)),"max":"—",
                               "cond":f"Device: {_dname}","temp_c":25,
                               "is_high":False,"extracted_unit":None,
                               "_devvar":True})
            if len(_vrecs)>=2:
                recs=_vrecs
        if not recs and p["param"]=="VF": recs=scan_diotec_vf(_blines,p)
        if not recs: recs=scan_baseline_lines(_blines,p)
        if not recs: recs=scan_text(full_text,p)
        # Drop spurious records harvested from the diagram section — a condition
        # carrying a "Diagram N" title or "X=f(Y)" axis caption is a plot axis
        # value (e.g. an Eoss axis max), never a tabulated parameter.
        recs=[r for r in recs
              if not re.search(r"Diagram\s*\d|=\s*f\s*\(", str(r.get("cond") or ""))]
        recs=enrich_temps(recs,full_text)
        # Deduplicate: same typ/max/temp extracted from multiple tables
        # (e.g. quick-reference table on page 1 repeats the characteristics
        # table). Keeps every DISTINCT condition row — important for IR where
        # typ AND max at each temperature must all survive.
        _seen=set(); _uniq=[]
        for r in recs:
            key=(r.get("typ"),r.get("max"),r.get("temp_c"))
            if key in _seen: continue
            _seen.add(key); _uniq.append(r)
        recs=_uniq
        lo=[r for r in recs if not r["is_high"]]
        hi=[r for r in recs if r["is_high"]]
        if not p.get("multi_cond",False) and not any(r.get("_devvar") for r in recs):
            lo=lo[:1]; hi=hi[:1]
        found=bool(recs); total=len(lo)+len(hi)
        print(f"      {'✓' if found else '–'} {p['param']:10s}  {total} row(s)")
        def mk(rec,_p=p):
            # Unit priority: table unit column → catalogue default
            _eu = rec.get("extracted_unit")
            unit = _norm_unit(_eu) if _eu else _p["unit"]
            return dict(param=_p["param"],name=_p["full"],unit=unit,
                        section=_p["section"],note=_p["note"],
                        typ=rec["typ"],max=rec["max"],temp_c=rec["temp_c"],
                        cond=rec["cond"] or "—",is_high=rec["is_high"],found=True)
        for rec in (lo+hi): output_rows.append(mk(rec))
        if not recs:
            output_rows.append(dict(param=p["param"],name=p["full"],unit=p["unit"],
                                    section=p["section"],note=p["note"],
                                    typ="—",max="—",temp_c=None,cond="—",
                                    is_high=False,found=False))

    # (R_DS(on) is a normal table parameter for MOSFETs — no graph-derived
    #  dynamic-resistance step is needed, unlike the diode rd extraction.)

    output_rows.sort(key=_skey)

    # ── Derating-curve digitisation (IF vs T  /  Ptot vs T) ──────────────────
    derating=[]
    if _HAS_GRAPH:
        print("      [graphs] Scanning for ENERGY (Eon/Eoff/Etot) and TEMPERATURE-axis figures ...")
        if not _HAS_OCR:
            print("      [graphs] pytesseract/tesseract not found — vector-glyph axis "
                  "labels (WeEN/Diotec style) cannot be auto-read.")
            print("                 Install with:  pip install pytesseract  "
                  "+  apt/choco install tesseract-ocr")
        try:
            derating=extract_derating_curves(pdf_path, interactive=interactive)
        except Exception as e:
            print(f"      [graphs] extraction error: {e}")
        for f in derating:
            rel = " (RELATIVE ratio)" if f.get("relative") else ""
            fid = f"Figure {f['fig_no']}" if f.get("fig_no") else "Graph"
            print(f"      \u2713 {fid} (p{f['page']}): {f['kind']}{rel} — "
                  f"{len(f['curves'])} curve(s), {len(f['temps'])} samples "
                  f"[calib: {f['calib']}]")
        if not derating:
            print("      \u2013 no temperature-axis graphs found (or figures are raster images)")

    return info,output_rows,derating

# ═════════════════════════════════════════════════════════════════════════════
# SECTION 7 — EXCEL OUTPUT
# ═════════════════════════════════════════════════════════════════════════════

def _bdr(s="thin",c="CCCCCC"):
    x=Side(style=s,color=c); return Border(left=x,right=x,top=x,bottom=x)
def _hdr(cell,bg,fg="FFFFFF",sz=10,bold=True,wrap=False):
    cell.font=Font(name="Calibri",bold=bold,size=sz,color=fg)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=wrap)
    cell.border=_bdr("medium","888888")
def _dat(cell,bg,fg="000000",bold=False,align="center",wrap=False):
    cell.font=Font(name="Calibri",size=10,bold=bold,color=fg)
    cell.fill=PatternFill("solid",fgColor=bg)
    cell.alignment=Alignment(horizontal=align,vertical="center",wrap_text=wrap)
    cell.border=_bdr("thin","CCCCCC")

CW={"A":14,"B":34,"C":8,"D":12,"E":12,"F":11,"G":54,"H":40,"I":13}
MC=("A","B","C","H","I")
SC={"Required Parameters":"D6E4F0","Loss Analysis Parameters":"EAD8F7",
    "Thermal Parameters":"D5F0E4"}

def _write_graph_sheet(wb, info, figs, sheet_title, intro):
    """Generic graph sheet: digitised curve data + embedded figure images.
    Used for both the Energy-graph and Temperature-graph sheets."""
    import tempfile
    from openpyxl.utils import get_column_letter
    try:
        from openpyxl.drawing.image import Image as _XLImage
        _can_embed = True
    except Exception:
        _can_embed = False

    ws = wb.create_sheet(sheet_title)
    ws.column_dimensions["A"].width = 12
    for ci in range(2, 16):
        ws.column_dimensions[get_column_letter(ci)].width = 17

    ws.merge_cells("A1:J1"); c = ws["A1"]
    c.value = _xl_safe(f"{sheet_title.upper()} — digitised from PDF graphics  |  "
               f"Part: {info['part']}")
    c.font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2"); c = ws["A2"]
    c.value = intro
    c.font = Font(name="Calibri", size=9, italic=True, color="555555")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    _tmpdir = tempfile.mkdtemp(prefix="mosfet_figs_")
    cur = 4
    for fi, f in enumerate(figs):
        n_curves = len(f["curves"])
        img_col = n_curves + 3                       # leave one blank column
        blk_w = max(img_col + 3, 10)

        ws.merge_cells(f"A{cur}:{get_column_letter(blk_w)}{cur}"); c = ws[f"A{cur}"]
        cap_body = re.sub(r"^fig(?:ure)?\.?\s*\d+\.?\s*[-\u2013:]?\s*", "", f["caption"], flags=re.I)
        fig_id = f"Figure {f['fig_no']}" if f.get("fig_no") else "Graph"
        c.value = _xl_safe(f"  \u25b6  {fig_id}  \u2014  {cap_body}"
                   f"   (page {f['page']})")
        c.font = Font(name="Calibri", bold=True, size=10, color="1F3864")
        kind_bg = {"current": "D6E4F0", "power": "D5F0E4",
                   "relative": "EAD8F7", "energy": "FCE4D6",
                   "resistance": "FFF2CC"}.get(f["kind"], "E2EFDA")
        c.fill = PatternFill("solid", fgColor=kind_bg)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[cur].height = 16; cur += 1

        ws.merge_cells(f"A{cur}:{get_column_letter(blk_w)}{cur}"); c = ws[f"A{cur}"]
        note = (f"X axis: {f['x_name']}   |   Y axis: {f['y_name']}   |   "
                f"Axis calibration \u2014 {f['calib']}")
        if f.get("conditions"):
            note += f"   |   Conditions (from figure): {f['conditions']}"
        if (f["kind"] == "current" and len(f["curves"]) > 2
                and all(cv["label"].startswith("Curve") for cv in f["curves"])):
            note += ("   |   Curves ordered top\u2192bottom as drawn; match the \u03b4 "
                     "(duty-cycle) labels from the embedded figure (top = smallest \u03b4)")
        if "NORMALISED" in f["calib"]:
            note += "   |   \u26a0 the flagged axis is in 0\u20131 fractions of its span"
        c.value = _xl_safe(note)
        c.font = Font(name="Calibri", size=9,
                      color="7B2D00" if "NORMALISED" in f["calib"] else "333333")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[cur].height = 26; cur += 1

        # unit for the curve (Y) columns = the LAST parenthesis in the y-axis
        # name (e.g. "RDS(on) (m\u03a9)" \u2192 "m\u03a9", not the "(on)" subscript).
        _yms = re.findall(r"\(([^)]+)\)", f.get("y_name", "") or "")
        y_unit = _yms[-1].strip() if _yms else ""
        x_unit = f.get("x_unit") or "\u00b0C"
        x_head = f["x_name"].split(" ")[0]
        hdr = [f"{x_head} ({x_unit})"]
        for cv in f["curves"]:
            h = cv["label"]
            if y_unit:
                h += f" ({y_unit})"
            if cv["style"] != "solid" and "dash" not in h.lower():
                h += f"  [{cv['style']}]"
            hdr.append(h)
        for ci, h in enumerate(hdr, 1):
            c = ws.cell(row=cur, column=ci, value=_xl_safe(h)); _hdr(c, "2E75B6", sz=9, wrap=True)
        ws.row_dimensions[cur].height = 30
        tbl_top = cur; cur += 1

        # Limit table to 15 rows: pick evenly-spaced indices over the full range
        _all_temps = f["temps"]
        _all_curves = f["curves"]
        if len(_all_temps) > 15:
            _idx = sorted({int(round(i * (len(_all_temps) - 1) / 14))
                           for i in range(15)})
            _disp_temps = [_all_temps[i] for i in _idx]
            _disp_curves = [{"label": cv["label"], "style": cv["style"],
                             "vals": [cv["vals"][i] for i in _idx]}
                            for cv in _all_curves]
        else:
            _disp_temps, _disp_curves = _all_temps, _all_curves
        for ti, t in enumerate(_disp_temps):
            c = ws.cell(row=cur, column=1, value=_xl_safe(t)); _dat(c, "EBF3FB", bold=True)
            for ci, cv in enumerate(_disp_curves, 2):
                v = cv["vals"][ti]
                c = ws.cell(row=cur, column=ci, value=v if v is not None else "")
                _dat(c, "FFFFFF" if ti % 2 else "F5F9FD")
            cur += 1

        # ── embed the original figure next to the table ─────────────────────
        if _can_embed and f.get("png"):
            try:
                pth = os.path.join(_tmpdir, f"fig_{fi}_{f['fig_no']}.png")
                with open(pth, "wb") as fh:
                    fh.write(f["png"])
                img = _XLImage(pth)
                scale = 340.0 / max(f["png_w"], 1)
                img.width  = int(f["png_w"] * scale)
                img.height = int(f["png_h"] * scale)
                ws.add_image(img, f"{get_column_letter(img_col)}{tbl_top}")
            except Exception:
                pass
        cur += 2

def write_excel(info,rows,out_path,derating=None):
    print("[4/4] Writing Excel ...")
    wb=Workbook(); ws=wb.active; ws.title="MOSFET Parameters"
    for col,w in CW.items(): ws.column_dimensions[col].width=w

    ws.merge_cells("A1:I1"); c=ws["A1"]
    c.value="POWER MOSFET PARAMETER EXTRACTION  —  LOSS-ANALYSIS REFERENCE  (v1)"
    c.font=Font(name="Calibri",bold=True,size=14,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor="1F3864")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=30

    ws.merge_cells("A2:I2"); c=ws["A2"]
    c.value=(f"Part: {info['part']}   |   Mfr: {info['mfr']}   |   "
             "Symbols: Infineon\u00b7ST\u00b7ON Semi\u00b7Toshiba\u00b7Nexperia\u00b7Vishay\u00b7ROHM\u00b7Wolfspeed\u00b7IXYS   |   "
             "Typ + Max at every stated temperature; any K/W or \u00b0C/W parameter captured")
    c.font=Font(name="Calibri",size=10,color="FFFFFF")
    c.fill=PatternFill("solid",fgColor="2E75B6")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=17; ws.row_dimensions[3].height=5

    hdrs=["Symbol","Parameter Name","Unit","Typ","Max","Temp\n(\u00b0C)",
          "Test Conditions","Design Note / Loss Relevance","Status"]
    for ci,h in enumerate(hdrs,1):
        c=ws.cell(row=4,column=ci,value=h); _hdr(c,"1F3864",wrap=True)
    ws.row_dimensions[4].height=36

    PL="EBF3FB"; WH="FFFFFF"; GR="E2EFDA"; AM="FFF2CC"; MX="FCE4D6"; HB="FFF9E6"
    cur=5; csec=None; par=0

    grps=[]
    for _,gi in itertools.groupby(rows,key=lambda r:(r["section"],r["param"])):
        grps.append(list(gi))

    for grp in grps:
        first=grp[0]
        if first["section"]!=csec:
            csec=first["section"]
            ws.merge_cells(f"A{cur}:I{cur}"); c=ws[f"A{cur}"]
            c.value=f"  \u25b6  {csec}"
            c.font=Font(name="Calibri",bold=True,size=9,color="1F3864")
            c.fill=PatternFill("solid",fgColor=SC.get(csec,"D6E4F0"))
            c.alignment=Alignment(horizontal="left",vertical="center")
            ws.row_dimensions[cur].height=13; cur+=1; par=0
        nc=len(grp); sr=cur; fd=first["found"]; pb=PL if par%2==0 else WH; par+=1
        for si,r in enumerate(grp):
            rb=HB if r["is_high"] else pb
            tb=GR if (r["found"] and r["typ"]!="—") else AM
            mb=MX if r["max"]!="—" else rb
            tf="1E5631" if (r["found"] and r["typ"]!="—") else "7D6608"
            mf="7B2D00" if r["max"]!="—" else "888888"
            ts=str(int(r["temp_c"])) if r["temp_c"] is not None else "25*"
            for cl,val,bg,fg,bo,al in [
                ("D",r["typ"],tb,tf,True,"center"),("E",r["max"],mb,mf,False,"center"),
                ("F",ts,rb,"333333",False,"center"),("G",r["cond"],rb,"333366",False,"left")]:
                c=ws[f"{cl}{cur}"]; c.value=val
                _dat(c,bg,fg=fg,bold=bo,align=al,wrap=(cl=="G"))
            if si==0:
                ss="Found \u2713" if fd else "Not Found"
                for cl,val,bg,fg,bo,al in [
                    ("A",first["param"],pb,"2E75B6",True,"center"),
                    ("B",first["name"],pb,"000000",False,"left"),
                    ("C",first["unit"],pb,"555555",False,"center"),
                    ("H",first["note"],pb,"333333",False,"left"),
                    ("I",ss,GR if fd else AM,"1E5631" if fd else "7D6608",True,"center")]:
                    c=ws[f"{cl}{cur}"]; c.value=val
                    _dat(c,bg,fg=fg,bold=bo,align=al,wrap=(cl in ("B","H")))
            ws.row_dimensions[cur].height=22; cur+=1
        if nc>1:
            for cl in MC:
                ws.merge_cells(f"{cl}{sr}:{cl}{cur-1}")
                tc=ws[f"{cl}{sr}"]
                tc.alignment=Alignment(horizontal=tc.alignment.horizontal,
                                       vertical="center",wrap_text=tc.alignment.wrap_text)

    cur+=1; ws.merge_cells(f"A{cur}:I{cur}"); c=ws[f"A{cur}"]
    c.value=("Legend: Green=found | Yellow=not found | Orange=max value | "
             "Warm-yellow=high-temp row | 25*=temperature not stated in datasheet | "
             "V_DS covers V(BR)DSS/VDSS | R_DS(on) in m\u03a9 | V_SD covers VSD/VF | "
             "Energy & Temperature graphs digitised on separate sheets")
    c.font=Font(name="Calibri",size=9,italic=True,color="555555")
    c.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
    ws.row_dimensions[cur].height=28
    ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:I{cur-2}"
    if derating:
        energy_figs  = [f for f in derating if f.get("kind") == "energy"]
        thermal_figs = [f for f in derating if f.get("kind") == "thermal_z"]
        cap_figs     = [f for f in derating if f.get("kind") == "capacitance"]
        temp_figs    = [f for f in derating if f.get("kind")
                        not in ("energy", "thermal_z", "capacitance")]
        if energy_figs:
            _write_graph_sheet(
                wb, info, energy_figs, "Energy Graphs",
                "Switching-energy figures (Eon / Eoff / Etot, Eoss) digitised "
                "point-by-point. The X axis is whatever the datasheet plots against "
                "(gate resistance RG, drain current ID, or junction temperature Tj) \u2014 "
                "its printed unit is shown in the first column header. The original "
                "figure is embedded next to each table for checking.")
        if temp_figs:
            _write_graph_sheet(
                wb, info, temp_figs, "Temperature Graphs",
                "Every figure whose variable is a temperature (Tj / Tc / Ta / Th): "
                "RDS(on) vs Tj, ID/Ptot derating vs Tc, SOA, etc. Values are read "
                "point-by-point (accuracy ~1-2% of full scale). The original figure is "
                "embedded next to each table for checking.")
        if thermal_figs:
            _write_graph_sheet(
                wb, info, thermal_figs, "Thermal Impedance Graphs",
                "Transient thermal impedance ZthJC = f(tp) with its duty-cycle "
                "(D = tp/T) curve family. Both axes are logarithmic (tp in seconds, "
                "ZthJC in K/W), so digitised values are approximate \u2014 treat the "
                "embedded figure as the authoritative source and use the table for "
                "quick interpolation of the single-pulse / duty-cycle envelope.")
        if cap_figs:
            _write_graph_sheet(
                wb, info, cap_figs, "Capacitance Graphs",
                "Junction capacitances C = f(VDS): Ciss, Coss and Crss versus "
                "drain-source voltage (VGS = 0 V). The Y axis is logarithmic (pF); "
                "values are read point-by-point and the original figure is embedded "
                "next to each table for checking.")
    wb.save(out_path)
    fc=len(set(r["param"] for r in rows if r["found"]))
    tc=len(set(r["param"] for r in rows))
    print(f"\n\u2705  Saved: {out_path}")
    print(f"    {fc}/{tc} parameters found.\n")

# ═════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════

def main():
    ap=argparse.ArgumentParser(
        description="Universal Power-MOSFET Parameter Extractor v1 [Single File]\n"
                    "Mfrs: Infineon|ST|ON Semi|Toshiba|Nexperia|Vishay|ROHM|Wolfspeed|IXYS\n"
                    "\nUsage: python mosfet_extractor_v1.py [optional: path_to_pdf]",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("pdf", nargs="?", default=None,
                    help="Path to MOSFET datasheet PDF (optional — prompted if not given)")
    ap.add_argument("--output","-o",help="Output .xlsx path")
    ap.add_argument("--debug-png",help="Save annotated VF-IF graph PNG",default=None)
    ap.add_argument("--no-prompt",action="store_true",
                    help="Never ask interactive questions (use defaults)")
    args=ap.parse_args()

    # Interactive file path input if not provided on command line
    pdf_path = args.pdf
    if not pdf_path:
        print("=" * 60)
        print("  Universal Power-MOSFET Parameter Extractor  (v1)")
        print("  Infineon · ST · ON Semi · Toshiba · Nexperia · ...")
        print("=" * 60)
        while True:
            pdf_path = input("\nEnter path to MOSFET datasheet PDF: ").strip().strip('"').strip("'")
            if not pdf_path:
                print("  [!] No path entered. Please try again.")
                continue
            if not os.path.exists(pdf_path):
                print(f"  [!] File not found: {pdf_path}")
                retry = input("  Try again? (y/n): ").strip().lower()
                if retry != 'y':
                    sys.exit(1)
            else:
                break

    if not os.path.exists(pdf_path):
        print(f"ERROR: file not found -> {pdf_path}"); sys.exit(1)

    # ── Output file name: ask the user (unless given with --output) ─────────
    out = args.output
    if not out:
        default_out = os.path.splitext(pdf_path)[0] + "_mosfet_params.xlsx"
        if args.no_prompt or not sys.stdin.isatty():
            out = default_out
        else:
            try:
                entered = input(f"\nEnter output Excel file name "
                                f"[press Enter for: {os.path.basename(default_out)}]: ").strip().strip('"').strip("'")
            except EOFError:
                entered = ""
            out = entered or default_out
            # if user typed just a name (no folder), save next to the PDF
            if entered and not os.path.dirname(out):
                out = os.path.join(os.path.dirname(default_out) or ".", out)
        if not out.lower().endswith(".xlsx"):
            out += ".xlsx"
    print(f"  Output will be saved to: {out}")
    if _HAS_GRAPH:
        print("[INFO] Graph extraction ENABLED (PyMuPDF + OpenCV + Pillow) \u2713")
    else:
        print("[WARN] Graph extraction disabled — install: pip install pymupdf opencv-python Pillow")
    print()
    info,rows,derating=extract(pdf_path,debug_png=args.debug_png,
                               interactive=not args.no_prompt)
    write_excel(info,rows,out,derating=derating)

if __name__=="__main__":
    main()
